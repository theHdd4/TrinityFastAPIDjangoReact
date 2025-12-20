# app/routes.py - API Routes
from fastapi import APIRouter, Depends, HTTPException, File, Form, UploadFile, Query, Request, Response
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any, Optional
import base64
import json
import pandas as pd
import polars as pl
import io
import os
import openpyxl
import pyarrow as pa
import pyarrow.ipc as ipc
from time import perf_counter
from app.core.utils import get_env_vars
from pathlib import Path
import fastexcel
from urllib.parse import unquote

# Add this line with your other imports
from datetime import datetime, timezone
import logging
import uuid
from collections import Counter


# app/routes.py - Add this import
from app.features.data_upload_validate.app.schemas import (
    # Create validator schemas
    CreateValidatorResponse,

    # Column types schemas
    UpdateColumnTypesResponse,
    MongoDBUpdateStatus,
    
    # Business dimensions schemas
    DefineDimensionsResponse,
    BusinessDimensionItem,
    
    # Assignment schemas
    AssignIdentifiersResponse,
    AssignmentSummary,
    
    # Validation schemas
    ValidateResponse,
    FileValidationResult,
    ValidationSummary,
    MinIOUploadResult,ConditionFailure,
    ProcessDataframeRequest,
    ProcessDataframeResponse,
)

# Add to your existing imports in app/routes.py
from app.features.data_upload_validate.app.database import get_validation_config_from_mongo  # ✅ ADD THIS

from app.features.data_upload_validate.app.database import save_validation_config_to_mongo
from app.features.data_upload_validate.app.schemas import ConfigureValidationConfigResponse
from app.features.data_upload_validate.app.database import get_validator_atom_from_mongo, update_validator_atom_in_mongo

from app.features.data_upload_validate.app.database import (
    save_business_dimensions_to_mongo,
    get_business_dimensions_from_mongo,
    update_business_dimensions_assignments_in_mongo,
    save_validation_units_to_mongo,
    get_validation_units_from_mongo,
)

from app.redis_cache import cache_master_config
from app.core.observability import timing_dependency_factory
from app.core.task_queue import celery_task_client, format_task_response

import re
import csv
from datetime import datetime

# Allowed characters for file keys (alphanumeric, underscores, hyphens, periods)
FILE_KEY_RE = re.compile(r"^[A-Za-z0-9_.-]+$")




from app.features.data_upload_validate.app.database import (
    get_validator_atom_from_mongo,  # Fallback function
    save_validation_log_to_mongo,
    log_operation_to_mongo,
    mark_operation_log_deleted,
)

# Add this import
from app.features.data_upload_validate.app.database import save_validator_atom_to_mongo

# Import column classifier functions for auto-classification
from app.features.column_classifier.database import (
    save_classifier_config_to_mongo,
    get_classifier_config_from_mongo,
)

# Initialize router
timing_dependency = timing_dependency_factory("app.features.data_upload_validate")

router = APIRouter(dependencies=[Depends(timing_dependency)])

logger = logging.getLogger(__name__)





from app.features.data_upload_validate.app.validators.custom_validator import perform_enhanced_validation
from app.features.data_upload_validate import service as data_upload_service
from app.features.data_upload_validate.file_ingestion import RobustFileReader

# Config directory
CUSTOM_CONFIG_DIR = data_upload_service.CUSTOM_CONFIG_DIR
MONGODB_DIR = data_upload_service.MONGODB_DIR
extraction_results = data_upload_service.extraction_results

CSV_READ_KWARGS = data_upload_service.CSV_READ_KWARGS


def _smart_csv_parse(content: bytes, csv_kwargs: dict) -> tuple[pl.DataFrame, list[str], dict]:
    return data_upload_service._smart_csv_parse(content, csv_kwargs)




# app/routes.py - Add MinIO imports and configuration

from minio import Minio
from minio.error import S3Error
from minio.commonconfig import CopySource
from app.core.feature_cache import feature_cache
from app.DataStorageRetrieval.db import (
    fetch_client_app_project,
    record_arrow_dataset,
    rename_arrow_dataset,
    delete_arrow_dataset,
    arrow_dataset_exists,
)
from app.DataStorageRetrieval.arrow_client import upload_dataframe, download_table_bytes
from app.DataStorageRetrieval.flight_registry import (
    set_ticket,
    get_ticket_by_key,
    get_latest_ticket_for_basename,
    get_original_csv,
    rename_arrow_object,
    remove_arrow_object,
    get_flight_path_for_csv,
    get_arrow_for_flight_path,
    CSV_TO_FLIGHT,
    FILEKEY_TO_CSV,
)
from app.DataStorageRetrieval.minio_utils import (
    ensure_minio_bucket,
    save_arrow_table,
    upload_to_minio,
    get_client,
    ARROW_DIR,
    get_arrow_dir,
)
from app.features.data_upload_validate.app.minio_sheet_utils import (
    extract_all_sheets_from_excel,
    get_sheet_data,
    list_upload_session_sheets,
    normalize_sheet_name,
    convert_session_sheet_to_arrow,
    list_upload_folders,
)
from pathlib import Path
import asyncio
import os


redis_client = feature_cache.router("data_upload_validate")

# ✅ MINIO CONFIGURATION - values come from docker-compose/.env
# Default to the development MinIO service if not explicitly configured
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "minio:9000")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "minio")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "minio123")
MINIO_BUCKET = os.getenv("MINIO_BUCKET", "trinity")


def _metadata_object_name(object_name: str) -> str:
    return f"{object_name}.meta.json"


def _save_workbook_metadata(object_name: str, payload: Dict[str, Any]) -> None:
    if not object_name:
        return
    data = json.dumps(payload, indent=2).encode("utf-8")
    buffer = io.BytesIO(data)
    minio_client.put_object(
        MINIO_BUCKET,
        _metadata_object_name(object_name),
        buffer,
        len(data),
        content_type="application/json",
    )


def _remove_workbook_metadata(object_name: str) -> None:
    if not object_name:
        return
    meta_name = _metadata_object_name(object_name)
    try:
        minio_client.remove_object(MINIO_BUCKET, meta_name)
    except S3Error as exc:
        code = getattr(exc, "code", "")
        if code not in {"NoSuchKey", "NoSuchBucket"}:
            logger.warning("Failed to remove workbook metadata %s: %s", meta_name, exc)
    except Exception:
        logger.warning("Failed to remove workbook metadata %s", meta_name)


def _load_workbook_metadata(object_name: str) -> Dict[str, Any]:
    if not object_name:
        raise FileNotFoundError("Missing object name for metadata lookup")
    meta_name = _metadata_object_name(object_name)
    try:
        response = minio_client.get_object(MINIO_BUCKET, meta_name)
    except S3Error as exc:
        code = getattr(exc, "code", "")
        if code in {"NoSuchKey", "NoSuchBucket"}:
            raise FileNotFoundError(meta_name) from exc
        raise
    with response as obj:
        data = obj.read()
    return json.loads(data.decode("utf-8"))


def _copy_workbook_artifacts(old_object: str, new_object: str) -> None:
    if not old_object or not new_object or old_object == new_object:
        return
    old_meta = _metadata_object_name(old_object)
    new_meta = _metadata_object_name(new_object)
    try:
        minio_client.copy_object(
            MINIO_BUCKET,
            new_meta,
            CopySource(MINIO_BUCKET, old_meta),
        )
        try:
            minio_client.remove_object(MINIO_BUCKET, old_meta)
        except S3Error:
            pass
    except S3Error as exc:
        code = getattr(exc, "code", "")
        if code not in {"NoSuchKey", "NoSuchBucket"}:
            logger.warning("Failed to copy workbook metadata %s -> %s: %s", old_meta, new_meta, exc)
        return
    except Exception as exc:
        logger.warning("Failed to copy workbook metadata %s -> %s: %s", old_meta, new_meta, exc)
        return

    try:
        metadata = _load_workbook_metadata(new_object)
    except FileNotFoundError:
        return

    workbook_path = metadata.get("workbook_path")
    if workbook_path:
        new_workbook_path = f"{new_object}.workbook{Path(workbook_path).suffix}"
        try:
            minio_client.copy_object(
                MINIO_BUCKET,
                new_workbook_path,
                CopySource(MINIO_BUCKET, workbook_path),
            )
            try:
                minio_client.remove_object(MINIO_BUCKET, workbook_path)
            except S3Error:
                pass
            metadata["workbook_path"] = new_workbook_path
            _save_workbook_metadata(new_object, metadata)
        except S3Error as exc:
            code = getattr(exc, "code", "")
            if code not in {"NoSuchKey", "NoSuchBucket"}:
                logger.warning("Failed to copy workbook file %s -> %s: %s", workbook_path, new_workbook_path, exc)
        except Exception as exc:
            logger.warning("Failed to copy workbook file %s -> %s: %s", workbook_path, new_workbook_path, exc)


def _delete_workbook_artifacts(object_name: str) -> None:
    try:
        metadata = _load_workbook_metadata(object_name)
    except FileNotFoundError:
        return
    workbook_path = metadata.get("workbook_path")
    if workbook_path:
        try:
            minio_client.remove_object(MINIO_BUCKET, workbook_path)
        except S3Error as exc:
            code = getattr(exc, "code", "")
            if code not in {"NoSuchKey", "NoSuchBucket"}:
                logger.warning("Failed to remove workbook %s: %s", workbook_path, exc)
        except Exception:
            logger.warning("Failed to remove workbook %s", workbook_path)
    _remove_workbook_metadata(object_name)


def _parse_numeric_id(value: str | int | None) -> int:
    """Return the numeric component of an ID string like "name_123"."""
    if value is None:
        return 0
    try:
        return int(str(value).split("_")[-1])
    except Exception:
        return 0

async def get_object_prefix(
    client_id: str = "",
    app_id: str = "",
    project_id: str = "",
    *,
    client_name: str = "",
    app_name: str = "",
    project_name: str = "",
    include_env: bool = False,
) -> str | tuple[str, dict[str, str], str]:
    """Return the MinIO prefix for the current client/app/project.

    When ``include_env`` is True a tuple of ``(prefix, env, source)`` is
    returned where ``source`` describes where the environment variables were
    loaded from.
    """
    USER_ID = _parse_numeric_id(os.getenv("USER_ID"))
    PROJECT_ID = _parse_numeric_id(project_id or os.getenv("PROJECT_ID", "0"))
    # If explicit names are provided, avoid using potentially stale identifier
    # values from ``os.environ``. This ensures that when the frontend sends the
    # current ``client_name/app_name/project_name`` combo, we resolve the
    # environment for that namespace rather than whatever IDs may have been set
    # previously.
    if client_name or app_name or project_name:
        client_id_env = client_id or ""
        app_id_env = app_id or ""
        project_id_env = project_id or ""
    else:
        client_id_env = client_id or os.getenv("CLIENT_ID", "")
        app_id_env = app_id or os.getenv("APP_ID", "")
        project_id_env = project_id or os.getenv("PROJECT_ID", "")

    # Resolve environment variables using ``get_env_vars`` which consults the
    # Redis cache keyed by ``<client>/<app>/<project>`` and falls back to
    # Postgres when missing.  This ensures we always load the latest names for
    # the currently selected namespace instead of defaulting to
    # ``default_client/default_app/default_project``.
    env: dict[str, str] = {}
    env_source = "unknown"
    fresh = await get_env_vars(
        client_id_env,
        app_id_env,
        project_id_env,
        client_name=client_name or os.getenv("CLIENT_NAME", ""),
        app_name=app_name or os.getenv("APP_NAME", ""),
        project_name=project_name or os.getenv("PROJECT_NAME", ""),
        use_cache=True,
        return_source=True,
    )
    if isinstance(fresh, tuple):
        env, env_source = fresh
    else:
        env, env_source = fresh, "unknown"

    # print(f"🔧 fetched env {env} (source={env_source})")  # Disabled
    client = env.get("CLIENT_NAME", os.getenv("CLIENT_NAME", "default_client"))
    app = env.get("APP_NAME", os.getenv("APP_NAME", "default_app"))
    project = env.get("PROJECT_NAME", os.getenv("PROJECT_NAME", "default_project"))

    if PROJECT_ID and (client == "default_client" or app == "default_app" or project == "default_project"):
        try:
            client_db, app_db, project_db = await fetch_client_app_project(
                USER_ID if USER_ID else None, PROJECT_ID
            )
            client = client_db or client
            app = app_db or app
            project = project_db or project
        except Exception as exc:  # pragma: no cover - database unreachable
            print(f"⚠️ Failed to load names from DB: {exc}")

    os.environ["CLIENT_NAME"] = client
    os.environ["APP_NAME"] = app
    os.environ["PROJECT_NAME"] = project
    prefix = f"{client}/{app}/{project}/"
    # print(
    #     f"📦 prefix {prefix} (CLIENT_ID={client_id or os.getenv('CLIENT_ID','')} APP_ID={app_id or os.getenv('APP_ID','')} PROJECT_ID={PROJECT_ID})"
    # )  # Disabled
    if include_env:
        return prefix, env, env_source
    return prefix


read_minio_object = data_upload_service.read_minio_object


@router.get("/get_object_prefix")
async def get_object_prefix_endpoint(
    client_id: str = "",
    app_id: str = "",
    project_id: str = "",
    client_name: str = "",
    app_name: str = "",
    project_name: str = "",
) -> dict:
    """Expose ``get_object_prefix`` as an API endpoint.

    The endpoint resolves the MinIO prefix for the provided client/app/project
    combination. Environment variables are sourced from Redis when available
    and otherwise retrieved from Postgres' ``registry_environment`` table.
    
    Can accept either IDs or names (or both). If names are provided, they take precedence.
    If only IDs are provided, names will be resolved dynamically from the database.
    """

    prefix, env, env_source = await get_object_prefix(
        client_id=client_id,
        app_id=app_id,
        project_id=project_id,
        client_name=client_name,
        app_name=app_name,
        project_name=project_name,
        include_env=True,
    )
    return {"prefix": prefix, "environment": env, "source": env_source}

# Initialize MinIO client
minio_client = get_client()
ensure_minio_bucket()

# MongoDB directory setup
MONGODB_DIR = Path("mongodb")
MONGODB_DIR.mkdir(exist_ok=True)

def save_non_validation_data(validator_atom_id: str, data_type: str, data: dict):
    """
    Save non-validation data to separate JSON files in mongodb folder
    data_type: 'business_dimensions', 'identifier_assignments'
    """
    try:
        file_path = MONGODB_DIR / f"{validator_atom_id}_{data_type}.json"
        
        # Load existing data if file exists
        existing_data = {}
        if file_path.exists():
            with open(file_path, "r") as f:
                existing_data = json.load(f)
        
        # Merge with new data
        existing_data.update(data)
        
        # Save updated data
        with open(file_path, "w") as f:
            json.dump(existing_data, f, indent=2)
        
        print(f"✅ Saved {data_type} for {validator_atom_id} to mongodb folder")
        return True
    except Exception as e:
        print(f"❌ Error saving {data_type}: {str(e)}")
        return False

load_all_non_validation_data = data_upload_service.load_all_non_validation_data
get_validator_from_memory_or_disk = data_upload_service.get_validator_from_memory_or_disk
load_existing_configs = data_upload_service.load_existing_configs


# Upload arbitrary file to MinIO and return its path
# Supports large files up to 2GB with chunked reading
@router.post("/upload-file")
async def upload_file(
    file: UploadFile = File(...),
    client_id: str = Form(""),
    app_id: str = Form(""),
    project_id: str = Form(""),
    client_name: str = Form(""),
    app_name: str = Form(""),
    project_name: str = Form(""),
    sheet_name: str = Form(""),
):
    import gc
    
    start_time = perf_counter()
    logger.info(
        "data_upload.temp_upload.start file=%s client_id=%s app_id=%s project_id=%s",
        file.filename,
        client_id or "",
        app_id or "",
        project_id or "",
    )
    if client_id:
        os.environ["CLIENT_ID"] = client_id
    if app_id:
        os.environ["APP_ID"] = app_id
    if project_id:
        os.environ["PROJECT_ID"] = project_id
    if client_name:
        os.environ["CLIENT_NAME"] = client_name
    if app_name:
        os.environ["APP_NAME"] = app_name
    if project_name:
        os.environ["PROJECT_NAME"] = project_name

    prefix = await get_object_prefix()
    tmp_prefix = prefix + "tmp/"
    
    # Maximum file size: 2GB
    MAX_FILE_SIZE = 2 * 1024 * 1024 * 1024  # 2 GB
    CHUNK_SIZE = 10 * 1024 * 1024  # 10 MB chunks
    
    try:
        # Read file in chunks to handle large files
        content_buffer = io.BytesIO()
        total_size = 0
        
        while True:
            chunk = await file.read(CHUNK_SIZE)
            if not chunk:
                break
            total_size += len(chunk)
            
            # Check file size limit
            if total_size > MAX_FILE_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File exceeds maximum size of 2GB. Current size: {total_size / (1024*1024*1024):.2f}GB"
                )
            
            content_buffer.write(chunk)
        
        content = content_buffer.getvalue()
        content_buffer.close()
        gc.collect()
        
        logger.info(
            "data_upload.temp_upload.read_bytes file=%s size=%s prefix=%s",
            file.filename,
            len(content),
            tmp_prefix,
        )

        # 🔥 SAVE DATA SUMMARY TO PIPELINE MONGO BEFORE CELERY PROCESSING
        if client_name and app_name and project_name:
            try:
                # Quick parse to get schema (don't store full dataframe)
                import polars as pl
                import io as io_module
                
                df_sample = None
                if file.filename.lower().endswith('.csv'):
                    df_sample = pl.read_csv(io_module.BytesIO(content), n_rows=1)
                elif file.filename.lower().endswith(('.xlsx', '.xls')):
                    import openpyxl
                    df_sample = pl.read_excel(io_module.BytesIO(content), sheet_name=sheet_name or 0, read_csv_options={"n_rows": 1})
                
                if df_sample is not None:
                    columns = list(df_sample.columns)
                    dtypes = {col: str(df_sample[col].dtype) for col in columns}
                    
                    # Build the file_key - matches the actual arrow file path
                    # Format: client_name/app_name/project_name/filename.arrow
                    prefix_without_tmp = tmp_prefix.replace("tmp/", "")
                    file_key = f"{prefix_without_tmp}{file.filename.rsplit('.', 1)[0]}.arrow"
                    
                    # Import and call async save_data_summary
                    from app.features.pipeline.service import save_data_summary
                    
                    summary_result = await save_data_summary(
                        client_name=client_name,
                        app_name=app_name,
                        project_name=project_name,
                        file_key=file_key,
                        columns=columns,
                        dtypes=dtypes,
                        mode="laboratory"
                    )
                    
                    if summary_result.get("status") == "success":
                        logger.info(f"✅ Saved data summary for {file.filename} ({len(columns)} columns) to pipeline MongoDB")
                    else:
                        logger.warning(f"⚠️ Failed to save data summary: {summary_result.get('error', 'Unknown')}")
                    
                    # Clean up sample
                    del df_sample
                    gc.collect()
            except Exception as e:
                logger.warning(f"⚠️ Could not save data summary (non-critical): {e}")

        submission = celery_task_client.submit_callable(
            name="data_upload_validate.upload_file",
            dotted_path="app.features.data_upload_validate.service.process_temp_upload",
            kwargs={
                "file_b64": base64.b64encode(content).decode("utf-8"),
                "filename": file.filename,
                "tmp_prefix": tmp_prefix,
                "sheet_name": sheet_name or None,
                "client_name": client_name or "",
                "app_name": app_name or "",
                "project_name": project_name or "",
            },
            metadata={
                "feature": "data_upload_validate",
                "operation": "upload_file",
                "filename": file.filename,
                "prefix": tmp_prefix,
                "file_size_mb": total_size / (1024 * 1024),
            },
        )
        
        # Free memory after submitting task
        del content
        gc.collect()

        if submission.status == "failure":  # pragma: no cover - defensive programming
            logger.error(
                "data_upload.temp_upload.failed task_id=%s file=%s",
                submission.task_id,
                file.filename,
            )
            raise HTTPException(status_code=400, detail=submission.detail or "Upload failed")

        duration_ms = (perf_counter() - start_time) * 1000
        logger.info(
            "data_upload.temp_upload.queued file=%s task_id=%s duration_ms=%.2f size_mb=%.2f",
            file.filename,
            submission.task_id,
            duration_ms,
            total_size / (1024 * 1024),
        )
        return format_task_response(submission, embed_result=True)
        
    except HTTPException:
        raise
    except MemoryError as e:
        logger.error(
            "data_upload.temp_upload.memory_error file=%s error=%s",
            file.filename,
            str(e),
        )
        gc.collect()
        raise HTTPException(
            status_code=507,
            detail="Server ran out of memory processing file. Try a smaller file or contact support."
        )
    except Exception as e:
        logger.exception(
            "data_upload.temp_upload.error file=%s",
            file.filename,
        )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to upload file: {str(e)}"
        )


@router.post("/upload-excel-multi-sheet")
async def upload_excel_multi_sheet(
    file: UploadFile = File(...),
    client_id: str = Form(""),
    app_id: str = Form(""),
    project_id: str = Form(""),
    client_name: str = Form(""),
    app_name: str = Form(""),
    project_name: str = Form(""),
):
    """
    Upload an Excel file and extract all sheets automatically.
    Each sheet is stored separately in MinIO as a Parquet file.
    Supports large files up to 2GB with chunked reading.
    
    Returns:
        {
            "upload_session_id": "uuid",
            "sheets": ["Sheet1", "Sheet2", ...],
            "sheet_details": [...],
            "original_file_path": "..."
        }
    """
    import tempfile
    import gc
    
    start_time = perf_counter()
    
    # Validate file type
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Only Excel files (.xlsx, .xls) are supported for multi-sheet upload"
        )
    
    # Set environment variables
    if client_id:
        os.environ["CLIENT_ID"] = client_id
    if app_id:
        os.environ["APP_ID"] = app_id
    if project_id:
        os.environ["PROJECT_ID"] = project_id
    if client_name:
        os.environ["CLIENT_NAME"] = client_name
    if app_name:
        os.environ["APP_NAME"] = app_name
    if project_name:
        os.environ["PROJECT_NAME"] = project_name
    
    # Get MinIO prefix
    prefix = await get_object_prefix()
    
    # Generate unique upload session ID
    upload_session_id = str(uuid.uuid4())
    
    # Maximum file size: 2GB
    MAX_EXCEL_SIZE = 2 * 1024 * 1024 * 1024  # 2 GB
    
    # Use chunked reading for large files to avoid memory issues
    CHUNK_SIZE = 10 * 1024 * 1024  # 10 MB chunks
    
    try:
        # Read file in chunks to handle large files
        # Use a temporary file for very large files to avoid memory issues
        content_buffer = io.BytesIO()
        total_size = 0
        
        while True:
            chunk = await file.read(CHUNK_SIZE)
            if not chunk:
                break
            total_size += len(chunk)
            
            # Check file size limit
            if total_size > MAX_EXCEL_SIZE:
                raise HTTPException(
                    status_code=413,
                    detail=f"File exceeds maximum size of 2GB. Current size: {total_size / (1024*1024*1024):.2f}GB"
                )
            
            content_buffer.write(chunk)
        
        # Get the complete content
        content = content_buffer.getvalue()
        content_buffer.close()
        
        # Force garbage collection to free memory from chunks
        gc.collect()
        
        logger.info(
            "data_upload.multi_sheet.start file=%s size=%s session_id=%s",
            file.filename,
            len(content),
            upload_session_id,
        )
        
        # Extract all sheets and store in MinIO
        result = extract_all_sheets_from_excel(
            excel_content=content,
            upload_session_id=upload_session_id,
            prefix=prefix
        )
        
        # Free memory after processing
        del content
        gc.collect()
        
        duration_ms = (perf_counter() - start_time) * 1000
        logger.info(
            "data_upload.multi_sheet.completed session_id=%s sheets=%s duration_ms=%.2f",
            upload_session_id,
            len(result["sheets"]),
            duration_ms,
        )
        
        # Return UI-friendly response format
        response_data = {
            "status": "success",
            "upload_session_id": result["upload_session_id"],
            "session_id": result["upload_session_id"],  # Alias for compatibility
            "file_name": file.filename,
            "file_size_mb": total_size / (1024 * 1024),
            "sheet_count": len(result["sheets"]),
            "sheets": result["sheets"],
            "sheet_details": result.get("sheet_details", []),
            "original_file_path": result.get("original_file_path", ""),
        }
        logger.info(
            "data_upload.multi_sheet.response session_id=%s sheets=%s sheet_details_count=%s",
            result["upload_session_id"],
            result["sheets"],
            len(result.get("sheet_details", [])),
        )
        return response_data
        
    except HTTPException:
        raise
    except ValueError as e:
        logger.error(
            "data_upload.multi_sheet.failed session_id=%s error=%s",
            upload_session_id,
            str(e),
        )
        raise HTTPException(status_code=400, detail=str(e))
    except MemoryError as e:
        logger.error(
            "data_upload.multi_sheet.memory_error session_id=%s error=%s",
            upload_session_id,
            str(e),
        )
        gc.collect()
        raise HTTPException(
            status_code=507,
            detail="Server ran out of memory processing file. Try a smaller file or contact support."
        )
    except Exception as e:
        logger.exception(
            "data_upload.multi_sheet.error session_id=%s",
            upload_session_id,
        )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process Excel file: {str(e)}"
        )


@router.delete("/temp-uploads")
async def clear_temp_uploads(
    client_name: str = "",
    app_name: str = "",
    project_name: str = "",
):
    """Remove any temporary uploads for the given environment."""
    start_time = perf_counter()
    prefix, env, env_source = await get_object_prefix(
        client_name=client_name,
        app_name=app_name,
        project_name=project_name,
        include_env=True,
    )
    tmp_prefix = prefix + "tmp/"
    try:
        objects = list(
            minio_client.list_objects(MINIO_BUCKET, prefix=tmp_prefix, recursive=True)
        )
        duration_ms = (perf_counter() - start_time) * 1000
        logger.info(
            "data_upload.temp_cleanup.completed prefix=%s deleted=%s duration_ms=%.2f",
            tmp_prefix,
            len(objects),
            duration_ms,
        )
        for obj in objects:
            minio_client.remove_object(MINIO_BUCKET, obj.object_name)
        return {
            "deleted": len(objects),
            "prefix": tmp_prefix,
            "environment": env,
            "env_source": env_source,
        }
    except S3Error as e:
        logger.warning(
            "data_upload.temp_cleanup.error prefix=%s error=%s",
            tmp_prefix,
            str(e),
        )
        return {
            "deleted": 0,
            "error": str(e),
            "prefix": tmp_prefix,
            "environment": env,
            "env_source": env_source,
        }


# POST: CREATE_NEW - Create validator atom with column preprocessing
@router.post("/create_new", status_code=202, response_model=CreateValidatorResponse)
async def create_new(
    validator_atom_id: str = Form(..., description="Unique ID for your validator atom"),
    files: List[UploadFile] = File(...),
    file_keys: str = Form(...)
) -> Dict[str, Any]:
    """
    Create new validator atom by uploading files and generating validation rules
    """
    # ✅ ADD COLUMN PREPROCESSING FUNCTION
    def preprocess_column_name(col_name: str) -> str:
        """
        Preprocess column name:
        - Strip leading/trailing spaces
        - Lowercase
        - Remove spaces inside the name but preserve underscores
        """
        col_name = col_name.strip().lower()
        # Remove spaces but keep underscores
        col_name = re.sub(r'(?<!_)\s+(?!_)', '', col_name)
        return col_name

    # Parse file_keys JSON
    try:
        keys = json.loads(file_keys)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON format for file_keys")

    # Basic validations
    if not validator_atom_id or not validator_atom_id.strip():
        raise HTTPException(status_code=400, detail="validator_atom_id cannot be empty")
    if len(files) != len(keys):
        raise HTTPException(status_code=400, detail="Number of keys must match number of files")
    if len(files) > 3:
        raise HTTPException(status_code=400, detail="Maximum 3 files supported")

    schemas = {}
    dataframes = {}  # Store DataFrames for data type extraction
    
    # Process each file
    for file, key in zip(files, keys):
        # Read file
        try:
            content = await file.read()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading file {file.filename}: {str(e)}")

        # Parse file to DataFrame using RobustFileReader (uses fixed CSVReader/ExcelReader)
        try:
            # Use RobustFileReader which handles column preservation automatically
            df_result, file_metadata = RobustFileReader.read_file_to_pandas(
                content=content,
                filename=file.filename,
                auto_detect_header=True,
                return_raw=False,
            )
            
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
            
            # Convert to Polars for consistency with existing code
            df_pl = pl.from_pandas(df)

            # Attempt to convert object columns that look like dates or datetimes
            date_pat = re.compile(
                r"^(?:\d{4}[-/]\d{2}[-/]\d{2}|\d{2}[-/]\d{2}[-/]\d{4})(?:[ T]\d{2}:\d{2}:\d{2}(?:\.\d+)?)?$"
            )
            for col in df.columns:
                if df[col].dtype == object:
                    sample = df[col].dropna().astype(str).head(5)
                    if not sample.empty and all(date_pat.match(v.strip()) for v in sample):
                        parsed = pd.to_datetime(df[col], errors="coerce", infer_datetime_format=True)
                        if parsed.notna().sum() >= len(df[col]) * 0.8:
                            df[col] = parsed
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parsing file {file.filename}: {str(e)}")

        # ✅ PREPROCESS COLUMN NAMES - Remove spaces but preserve underscores
        df.columns = [preprocess_column_name(col) for col in df.columns]

        # Store DataFrame for data type extraction
        dataframes[key] = df

        # Extract schema
        schema = []
        for col in df.columns:
            dtype = str(df[col].dtype)
            if "int" in dtype or "float" in dtype:
                col_type = "numeric"
            elif "datetime" in dtype:
                col_type = "date"
            elif "bool" in dtype:
                col_type = "boolean"
            else:
                col_type = "string"
            
            schema.append({"column": col, "type": col_type})

        # Store schema info
        schemas[key] = {
            "columns": schema,
            "sample_rows": df.head(3).astype(str).to_dict(orient="records"),  # ✅ ONLY THIS LINE CHANGED
            "total_rows": len(df),
            "total_columns": len(df.columns)
        }


    # Extract data types for each column in each file
    column_types = {}
    for key, df in dataframes.items():
        types = {}
        for col in df.columns:
            dtype = str(df[col].dtype)
            if "int" in dtype:
                types[col] = "integer"
            elif "float" in dtype:
                types[col] = "numeric"
            elif "datetime" in dtype:
                types[col] = "date"
            elif "bool" in dtype:
                types[col] = "boolean"
            else:
                types[col] = "string"
        column_types[key] = types

    # Also include column_types in the schemas
    for key in schemas:
        schemas[key]["column_types"] = column_types.get(key, {})

    # Generate validation config
    validation_config = {
        "validator_atom_id": validator_atom_id,
        "created_from_files": [f.filename for f in files],
        "file_keys": keys,
        "schemas": schemas,
        "column_types": column_types,
        "validation_mode": "simple"
    }

    # Save to file
    try:
        config_path = CUSTOM_CONFIG_DIR / f"{validator_atom_id}.json"
        with open(config_path, "w") as f:
            json.dump(validation_config, f, indent=2)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving config: {str(e)}")

    # ✅ ADD: Save to MongoDB
    mongo_result = save_validator_atom_to_mongo(validator_atom_id, validation_config)

    # Store in memory for GET endpoint
    extraction_results[validator_atom_id] = {
        "validator_atom_id": validator_atom_id,
        "schemas": schemas,
        "column_types": column_types,
        "config_saved": True,
        "config_path": str(config_path)
    }

    # ✅ MINIMAL POST RESPONSE - Only success confirmation
    return {
        "status": "success",
        "message": "Validator atom created successfully", 
        "validator_atom_id": validator_atom_id,
        "config_saved": True
    }


@router.post("/convert-session-sheet-to-arrow")
async def convert_session_sheet_to_arrow_endpoint(
    upload_session_id: str = Form(...),
    sheet_name: str = Form(...),
    original_filename: str = Form(...),
    use_folder_structure: str = Form("true"),
    client_name: str = Form(""),
    app_name: str = Form(""),
    project_name: str = Form(""),
):
    """
    Convert a sheet from an upload session to Arrow format and save it.
    
    Args:
        upload_session_id: The upload session ID
        sheet_name: The normalized sheet name
        original_filename: Original Excel filename
        use_folder_structure: "true" or "false" - whether to use folder structure
        client_name, app_name, project_name: Environment context
        
    Returns:
        {
            "file_path": "path/to/file.arrow",
            "file_name": "filename (sheet_name)",
            "file_key": "file_key"
        }
    """
    # Validate inputs
    if not upload_session_id or not upload_session_id.strip():
        raise HTTPException(status_code=400, detail="upload_session_id is required and cannot be empty")
    
    if not sheet_name or not sheet_name.strip():
        raise HTTPException(status_code=400, detail="sheet_name is required and cannot be empty")
    
    if not original_filename or not original_filename.strip():
        raise HTTPException(status_code=400, detail="original_filename is required and cannot be empty")
    
    # Set environment variables
    if client_name:
        os.environ["CLIENT_NAME"] = client_name
    if app_name:
        os.environ["APP_NAME"] = app_name
    if project_name:
        os.environ["PROJECT_NAME"] = project_name
    
    prefix = await get_object_prefix()
    use_folder = use_folder_structure.lower() == "true"
    
    try:
        logger.info(
            "convert_session_sheet_to_arrow_endpoint called: session_id=%s sheet_name=%s original_filename=%s use_folder=%s",
            upload_session_id,
            sheet_name,
            original_filename,
            use_folder
        )
        result = convert_session_sheet_to_arrow(
            upload_session_id=upload_session_id,
            sheet_name=sheet_name,
            original_filename=original_filename,
            prefix=prefix,
            use_folder_structure=use_folder
        )
        logger.info(
            "convert_session_sheet_to_arrow_endpoint success: file_path=%s file_name=%s",
            result.get("file_path"),
            result.get("file_name")
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception(f"Error converting sheet {sheet_name} from session {upload_session_id}")
        raise HTTPException(status_code=500, detail=f"Failed to convert sheet: {str(e)}")


# POST: UPDATE_COLUMN_TYPES - Allow user to change column data types
@router.post("/update_column_types", response_model=UpdateColumnTypesResponse)
async def update_column_types(
    validator_atom_id: str = Form(...),
    file_key: str = Form(...),
    column_types: str = Form(...)
):
    """
    Update column data types for a specific validator atom and file key
    """
    # Parse column_types JSON
    try:
        submitted_column_types = json.loads(column_types)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON format for column_types")

    # Check if validator atom exists (MongoDB first)
    validator_data = get_validator_atom_from_mongo(validator_atom_id)
    if not validator_data:
        # Fallback to old method for backward compatibility
        validator_data = get_validator_from_memory_or_disk(validator_atom_id)

    if not validator_data:
        raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")

    # Check if file_key exists
    if file_key not in validator_data["schemas"]:
        raise HTTPException(status_code=400, detail=f"File key '{file_key}' not found in validator")

    # Get current schema
    current_schema = validator_data["schemas"][file_key]
    available_columns = [col["column"] for col in current_schema.get("columns", [])]

    # Validate that all columns in submitted_column_types exist in the schema
    invalid_columns = [col for col in submitted_column_types.keys() if col not in available_columns]
    if invalid_columns:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid columns: {invalid_columns}. Available columns: {available_columns}"
        )

    # Validate column type values
    valid_types = ["string", "integer", "numeric", "date", "boolean", "number"]
    invalid_types = {col: typ for col, typ in submitted_column_types.items() if typ not in valid_types and typ not in ["", None, "not_defined"]}
    if invalid_types:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid column types: {invalid_types}. Valid types: {valid_types}"
        )

    # Normalize and update column_types in memory
    current_column_types = extraction_results[validator_atom_id]["column_types"].get(file_key, {})
    for col in available_columns:
        val = submitted_column_types.get(col)
        if val in ["", None, "not_defined"]:
            current_column_types.pop(col, None)
        else:
            normalized = "numeric" if val == "number" else val
            current_column_types[col] = normalized
    extraction_results[validator_atom_id]["column_types"][file_key] = current_column_types

    # Also update in schemas for consistency
    extraction_results[validator_atom_id]["schemas"][file_key]["column_types"] = current_column_types

    # Update the columns array with new types
    updated_columns = []
    for col_info in current_schema["columns"]:
        col_name = col_info["column"]
        if col_name in submitted_column_types and submitted_column_types.get(col_name) not in ["", None, "not_defined"]:
            normalized = "numeric" if submitted_column_types[col_name] == "number" else submitted_column_types[col_name]
            col_info["type"] = normalized
        updated_columns.append(col_info)
    
    extraction_results[validator_atom_id]["schemas"][file_key]["columns"] = updated_columns

    # Update JSON config file
    try:
        config_path = CUSTOM_CONFIG_DIR / f"{validator_atom_id}.json"
        if config_path.exists():
            with open(config_path, "r") as f:
                config = json.load(f)
            
            # Update column_types in config
            if "column_types" not in config:
                config["column_types"] = {}
            config["column_types"][file_key] = current_column_types
            
            # Update schemas in config
            if "schemas" in config and file_key in config["schemas"]:
                config["schemas"][file_key]["column_types"] = current_column_types
                config["schemas"][file_key]["columns"] = updated_columns
            
            # Save updated config
            with open(config_path, "w") as f:
                json.dump(config, f, indent=2)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating config file: {str(e)}")

    # ✅ MongoDB Update - INSERT THIS NEW CODE
    mongo_update_data = {
        f"schemas.{file_key}.column_types": current_column_types,
        f"schemas.{file_key}.columns": updated_columns,
        f"column_types.{file_key}": current_column_types
    }
    mongo_result = update_validator_atom_in_mongo(validator_atom_id, mongo_update_data)

    # Save datatype validation units
    datatype_units = [
        {"column": col, "validation_type": "datatype", "expected": typ}
        for col, typ in current_column_types.items()
    ]
    existing_units = get_validation_units_from_mongo(validator_atom_id, file_key)
    other_units = []
    if existing_units and "validations" in existing_units:
        other_units = [
            u for u in existing_units["validations"] if u.get("validation_type") != "datatype"
        ]
    save_validation_units_to_mongo(
        validator_atom_id,
        file_key,
        other_units + datatype_units,
    )

    # Optional: Log MongoDB result
    if mongo_result["status"] == "success":
        print(f"✅ Validator atom updated in MongoDB")
    else:
        print(f"⚠️ MongoDB update failed: {mongo_result['error']}")

    return {
        "status": "success",
        "message": "Column types updated successfully",
        "validator_atom_id": validator_atom_id,
        "file_key": file_key,
        "updated_column_types": submitted_column_types,
        "current_all_column_types": current_column_types,
        "updated_columns_count": len([c for c in submitted_column_types.values() if c not in ["", None, "not_defined"]]),
        # ✅ ADD: MongoDB update status
        "mongodb_update": {
            "status": mongo_result["status"],
            "modified": mongo_result.get("modified_count", 0) > 0 if mongo_result["status"] == "success" else False,
            "details": mongo_result.get("error", "Update successful") if mongo_result["status"] == "error" else f"Matched: {mongo_result.get('matched_count', 0)}, Modified: {mongo_result.get('modified_count', 0)}"
        }
    }

    
    

# POST: DEFINE_DIMENSIONS - Complete fixed version for both validator types
@router.post("/define_dimensions", response_model=DefineDimensionsResponse)
async def define_dimensions(
    validator_atom_id: str = Form(...),
    file_key: str = Form(...),
    dimensions: str = Form(...)
):
    """
    Endpoint to define business dimensions for a specific file key in a validator atom.
    Maximum of 4 dimensions allowed per file key.
    Works for both regular validator atoms (from /create_new) and template validator atoms (from /validate_*)
    """
    # Parse dimensions JSON
    try:
        dims = json.loads(dimensions)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON format for dimensions: {str(e)}")

    # Validate dims structure
    if not isinstance(dims, list):
        raise HTTPException(status_code=400, detail="Dimensions must be a list of objects")

    # ✅ ENFORCE MAX 4 DIMENSIONS
    if len(dims) > 4:
        raise HTTPException(status_code=400, detail="Maximum of 4 dimensions allowed")
    
    if len(dims) == 0:
        raise HTTPException(status_code=400, detail="At least 1 dimension must be provided")

    # Validate each dimension structure
    required_fields = ['id', 'name']
    dimension_ids = []
    dimension_names = []
    
    for i, dim in enumerate(dims):
        if not isinstance(dim, dict):
            raise HTTPException(status_code=400, detail=f"Dimension {i+1} must be an object")
        
        # Check required fields
        for field in required_fields:
            if field not in dim:
                raise HTTPException(status_code=400, detail=f"Dimension {i+1} missing required field: '{field}'")
            if not dim[field] or not isinstance(dim[field], str):
                raise HTTPException(status_code=400, detail=f"Dimension {i+1} field '{field}' must be a non-empty string")
        
        # Check for duplicate IDs and names
        if dim['id'] in dimension_ids:
            raise HTTPException(status_code=400, detail=f"Duplicate dimension ID: '{dim['id']}'")
        if dim['name'] in dimension_names:
            raise HTTPException(status_code=400, detail=f"Duplicate dimension name: '{dim['name']}'")
        
        dimension_ids.append(dim['id'])
        dimension_names.append(dim['name'])

    # ✅ UPDATED: Check if validator atom exists (MongoDB first)
    validator_data = get_validator_atom_from_mongo(validator_atom_id)
    if not validator_data:
        # Fallback to old method for backward compatibility
        validator_data = get_validator_from_memory_or_disk(validator_atom_id)

    if not validator_data:
        raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")

    # Check if file_key exists
    if file_key not in validator_data["schemas"]:
        available_keys = list(validator_data["schemas"].keys())
        raise HTTPException(
            status_code=400, 
            detail=f"File key '{file_key}' not found in validator. Available file keys: {available_keys}"
        )

    # Store dimensions for this specific file key
    dims_dict = {dim['id']: dim for dim in dims}
    
    # Add metadata
    dimension_data = {
        "dimensions": dims_dict,
        "file_key": file_key,
        "validator_atom_id": validator_atom_id,
        "timestamp": datetime.now().isoformat(),
        "validator_type": validator_data.get("template_type", "custom"),
        "dimensions_count": len(dims)
    }

    # ✅ REPLACE: Save to MongoDB instead of file
    try:
        mongo_result = save_business_dimensions_to_mongo(validator_atom_id, file_key, dims_dict)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save dimensions to MongoDB: {str(e)}")

    # ✅ FIXED: Safe update in memory (works for both validator types)
    try:
        # Initialize extraction_results entry if it doesn't exist (for template validator atoms)
        if validator_atom_id not in extraction_results:
            extraction_results[validator_atom_id] = {}
        
        if "business_dimensions" not in extraction_results[validator_atom_id]:
            extraction_results[validator_atom_id]["business_dimensions"] = {}
        
        extraction_results[validator_atom_id]["business_dimensions"][file_key] = dims_dict
        in_memory_status = "success"
    except Exception as e:
        # Log but don't fail - MongoDB save is what matters
        print(f"Warning: Could not update in-memory results for {validator_atom_id}: {e}")
        in_memory_status = "warning"

    return {
        "status": "success",
        "message": f"Business dimensions defined successfully for file key '{file_key}' ({len(dims)} dimensions)",
        "validator_atom_id": validator_atom_id,
        "file_key": file_key,
        "validator_type": validator_data.get("template_type", "custom"),
        "dimensions": dims_dict,
        "dimensions_count": len(dims),
        "max_allowed": 4,
        "dimension_details": {
            "dimension_ids": dimension_ids,
            "dimension_names": dimension_names,
            "created_at": datetime.now().isoformat()
        },
        "mongodb_saved": mongo_result.get("status") == "success",
        "in_memory_saved": in_memory_status,
        "next_steps": {
            "assign_identifiers": f"POST /assign_identifiers_to_dimensions with validator_atom_id: {validator_atom_id}",
            "view_assignments": f"GET /get_identifier_assignments/{validator_atom_id}/{file_key}"
        }
    }




# # POST: ASSIGN_IDENTIFIERS_TO_DIMENSIONS - Save assignments in business dimensions structure
# @router.post("/assign_identifiers_to_dimensions", response_model=AssignIdentifiersResponse)
# async def assign_identifiers_to_dimensions(
#     validator_atom_id: str = Form(...),
#     file_key: str = Form(...),
#     identifier_assignments: str = Form(...)
# ):
#     """
#     Assign identifiers to dimensions and save within business dimensions structure
#     """
#     # Parse identifier_assignments JSON
#     try:
#         assignments = json.loads(identifier_assignments)
#     except json.JSONDecodeError:
#         raise HTTPException(status_code=400, detail="Invalid JSON format for identifier_assignments")

#     # Validate assignments structure
#     if not isinstance(assignments, dict):
#         raise HTTPException(status_code=400, detail="identifier_assignments must be a JSON object")

#     # ✅ Check if validator atom exists (MongoDB first)
#     validator_data = get_validator_atom_from_mongo(validator_atom_id)
#     if not validator_data:
#         # Fallback to old method for backward compatibility
#         validator_data = get_validator_from_memory_or_disk(validator_atom_id)

#     if not validator_data:
#         raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")

#     # Check if file_key exists
#     if file_key not in validator_data.get("schemas", {}):
#         raise HTTPException(status_code=400, detail=f"File key '{file_key}' not found in validator")

#     # ✅ Get business dimensions from MongoDB first with correct structure handling
#     mongo_dimensions = get_business_dimensions_from_mongo(validator_atom_id, file_key)

#     if mongo_dimensions:
#         # MongoDB format: extract from dimensions array
#         dimensions_array = mongo_dimensions.get("dimensions", [])
#         available_dimension_ids = [dim.get("dimension_id") for dim in dimensions_array]
#         business_dimensions = {dim["dimension_id"]: dim for dim in dimensions_array}
#     elif validator_data.get("business_dimensions", {}).get(file_key, {}):
#         # Old format: dictionary of dimensions
#         business_dimensions = validator_data.get("business_dimensions", {}).get(file_key, {})
#         available_dimension_ids = list(business_dimensions.keys())
#     else:
#         raise HTTPException(status_code=400, detail=f"No business dimensions defined for file key '{file_key}'. Define dimensions first.")


# POST: ASSIGN_IDENTIFIERS_TO_DIMENSIONS - Complete fixed version for both validator types
@router.post("/assign_identifiers_to_dimensions", response_model=AssignIdentifiersResponse)
async def assign_identifiers_to_dimensions(
    validator_atom_id: str = Form(...),
    file_key: str = Form(...),
    identifier_assignments: str = Form(...)
):
    """
    Assign identifiers to dimensions and save within business dimensions structure.
    Works for both regular validator atoms (from /create_new) and template validator atoms (from /validate_*)
    """
    # Parse identifier_assignments JSON
    try:
        assignments = json.loads(identifier_assignments)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"Invalid JSON format for identifier_assignments: {str(e)}")

    # Validate assignments structure
    if not isinstance(assignments, dict):
        raise HTTPException(status_code=400, detail="identifier_assignments must be a JSON object")
    
    if not assignments:
        raise HTTPException(status_code=400, detail="identifier_assignments cannot be empty")

    # ✅ Check if validator atom exists (MongoDB first)
    validator_data = get_validator_atom_from_mongo(validator_atom_id)
    if not validator_data:
        # Fallback to old method for backward compatibility
        validator_data = get_validator_from_memory_or_disk(validator_atom_id)

    if not validator_data:
        raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")

    # Check if file_key exists
    if file_key not in validator_data.get("schemas", {}):
        available_keys = list(validator_data.get("schemas", {}).keys())
        raise HTTPException(
            status_code=400, 
            detail=f"File key '{file_key}' not found in validator. Available file keys: {available_keys}"
        )

    # ✅ Get business dimensions from MongoDB first with correct structure handling
    mongo_dimensions = get_business_dimensions_from_mongo(validator_atom_id, file_key)

    if mongo_dimensions:
        # MongoDB format: extract from dimensions array
        dimensions_array = mongo_dimensions.get("dimensions", [])
        available_dimension_ids = [dim.get("dimension_id") for dim in dimensions_array]
        business_dimensions = {dim["dimension_id"]: dim for dim in dimensions_array}
    elif validator_data.get("business_dimensions", {}).get(file_key, {}):
        # Old format: dictionary of dimensions
        business_dimensions = validator_data.get("business_dimensions", {}).get(file_key, {})
        available_dimension_ids = list(business_dimensions.keys())
    else:
        raise HTTPException(
            status_code=400, 
            detail=f"No business dimensions defined for file key '{file_key}'. Define dimensions first using /define_dimensions."
        )

    # Validate dimension IDs
    invalid_dimensions = [dim_id for dim_id in assignments.keys() if dim_id not in available_dimension_ids]
    if invalid_dimensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid dimension IDs: {invalid_dimensions}. Available dimensions: {available_dimension_ids}"
        )

    # Validate assignments
    all_assigned_identifiers = []
    for dim_id, identifiers in assignments.items():
        if not isinstance(identifiers, list):
            raise HTTPException(status_code=400, detail=f"Identifiers for dimension '{dim_id}' must be a list")
        
        if not identifiers:
            raise HTTPException(status_code=400, detail=f"Identifiers list for dimension '{dim_id}' cannot be empty")
        
        all_assigned_identifiers.extend(identifiers)

    # Check for unique assignment
    if len(all_assigned_identifiers) != len(set(all_assigned_identifiers)):
        duplicates = [ident for ident in set(all_assigned_identifiers) if all_assigned_identifiers.count(ident) > 1]
        raise HTTPException(status_code=400, detail=f"Identifiers cannot be assigned to multiple dimensions: {duplicates}")

    # ✅ UPDATE BUSINESS DIMENSIONS STRUCTURE WITH ASSIGNMENTS
    updated_business_dimensions = business_dimensions.copy()
    for dim_id, identifiers in assignments.items():
        if dim_id in updated_business_dimensions:
            updated_business_dimensions[dim_id]["assigned_identifiers"] = identifiers
            updated_business_dimensions[dim_id]["assignment_timestamp"] = datetime.now().isoformat()

    # ✅ Save to MongoDB
    try:
        mongo_result = update_business_dimensions_assignments_in_mongo(validator_atom_id, file_key, assignments)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save assignments to MongoDB: {str(e)}")

    # ✅ FIXED: Safe update in memory (works for both validator types)
    try:
        # Initialize extraction_results entry if it doesn't exist (for template validator atoms)
        if validator_atom_id not in extraction_results:
            extraction_results[validator_atom_id] = {}
        
        if "business_dimensions" not in extraction_results[validator_atom_id]:
            extraction_results[validator_atom_id]["business_dimensions"] = {}
        
        extraction_results[validator_atom_id]["business_dimensions"][file_key] = updated_business_dimensions
        in_memory_status = "success"
    except Exception as e:
        # Log but don't fail - MongoDB save is what matters
        print(f"Warning: Could not update in-memory results for {validator_atom_id}: {e}")
        in_memory_status = "warning"

    # In this simplified version identifiers are not validated,
    # so all provided identifiers are considered assigned
    unassigned_identifiers: list = []

    return {
        "status": "success",
        "message": f"Identifiers assigned to dimensions and saved in business dimensions structure for file key '{file_key}'",
        "validator_atom_id": validator_atom_id,
        "file_key": file_key,
        "validator_type": validator_data.get("template_type", "custom"),
        "updated_business_dimensions": updated_business_dimensions,
        "assignment_summary": {
            "total_identifiers": len(all_assigned_identifiers),
            "dimensions_with_assignments": len(assignments),
            "assignment_timestamp": datetime.now().isoformat()
        },
        "unassigned_identifiers": unassigned_identifiers,
        "dimension_breakdown": {dim_id: len(identifiers) for dim_id, identifiers in assignments.items()},
        "mongodb_updated": mongo_result.get("status") == "success",
        "in_memory_updated": in_memory_status,
        "next_steps": {
            "view_complete_setup": f"GET /get_validator_atom_summary/{validator_atom_id}",
            "export_configuration": f"GET /export_validator_atom/{validator_atom_id}"
        }
    }



# Add this constant at the top of routes.py (after imports)
VALIDATION_OPERATORS = {
    "greater_than": ">",
    "greater_than_or_equal": ">=", 
    "less_than": "<",
    "less_than_or_equal": "<=",
    "equal_to": "==",
    "not_equal_to": "!=",
    "between": "BETWEEN",
    "contains": "CONTAINS",
    "not_contains": "NOT_CONTAINS",
    "starts_with": "STARTS_WITH",
    "ends_with": "ENDS_WITH",
    "regex_match": "REGEX",
    "in_list": "IN",
    "not_in_list": "NOT_IN",
    "date_before": "DATE_BEFORE",
    "date_after": "DATE_AFTER",
    "date_between": "DATE_BETWEEN"
}

# ✅ ADD: Valid frequency options
VALID_FREQUENCIES = ["daily", "weekly", "monthly"]

# Updated endpoint with data frequency per column support
@router.post("/configure_validation_config", response_model=ConfigureValidationConfigResponse)
async def configure_validation_config(request: Request):
    """
    Configure custom validation config for specific columns with optional date frequency per column
    """
    data = await request.json()
    validator_atom_id = data.get("validator_atom_id")
    file_key = data.get("file_key")
    column_conditions = data.get("column_conditions")
    column_frequencies = data.get("column_frequencies", {})  # ✅ NEW: Optional dict of column to frequency

    if not validator_atom_id:
        raise HTTPException(status_code=400, detail="validator_atom_id is required")
    if not file_key:
        raise HTTPException(status_code=400, detail="file_key is required")
    if column_conditions is None:
        raise HTTPException(status_code=400, detail="column_conditions is required")

    if not isinstance(column_conditions, dict):
        raise HTTPException(status_code=400, detail="column_conditions must be a dictionary of column to list of conditions")

    if not isinstance(column_frequencies, dict):
        raise HTTPException(status_code=400, detail="column_frequencies must be a dictionary of column to frequency strings")

    # ✅ NEW: Validate frequencies if provided
    for col, freq in column_frequencies.items():
        if freq.lower() not in VALID_FREQUENCIES:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid frequency '{freq}' for column '{col}'. Valid options: {VALID_FREQUENCIES}"
            )

    validator_data = get_validator_atom_from_mongo(validator_atom_id)
    if not validator_data:
        validator_data = get_validator_from_memory_or_disk(validator_atom_id)

    if not validator_data:
        raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")

    if file_key not in validator_data.get("schemas", {}):
        raise HTTPException(status_code=400, detail=f"File key '{file_key}' not found in validator")

    available_columns = [col["column"] for col in validator_data["schemas"][file_key].get("columns", [])]

    total_conditions = 0
    for col, cond_list in column_conditions.items():
        if col not in available_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Column '{col}' not found in validator schema. Available columns: {available_columns}"
            )
        if not isinstance(cond_list, list):
            raise HTTPException(status_code=400, detail=f"Conditions for column '{col}' must be a list")
        
        for i, cond in enumerate(cond_list):
            if not isinstance(cond, dict):
                raise HTTPException(status_code=400, detail=f"Condition {i+1} for column '{col}' must be a dictionary")
            
            required_fields = ['operator', 'value', 'error_message']
            missing_fields = [field for field in required_fields if field not in cond]
            if missing_fields:
                raise HTTPException(
                    status_code=400,
                    detail=f"Condition {i+1} for column '{col}' missing required fields: {missing_fields}"
                )
            
            if cond['operator'] not in VALIDATION_OPERATORS:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid operator '{cond['operator']}' for column '{col}'. Valid operators: {list(VALIDATION_OPERATORS.keys())}"
                )
            
            if 'severity' not in cond:
                cond['severity'] = 'error'
            
            if cond['severity'] not in ['error', 'warning']:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid severity '{cond['severity']}' for column '{col}'. Must be 'error' or 'warning'"
                )
            
            total_conditions += 1

    # ✅ NEW: Build config data with optional column frequencies
    config_data = {
        "column_conditions": column_conditions,
        "column_frequencies": column_frequencies
    }

    mongo_result = save_validation_config_to_mongo(validator_atom_id, file_key, config_data)

    # Build validation units and save
    range_units = []
    regex_units = []
    null_units = []
    ref_units = []
    for col, conds in column_conditions.items():
        min_val = None
        max_val = None
        for cond in conds:
            op = cond.get("operator")
            if op in ["greater_than_or_equal", "greater_than"]:
                min_val = cond.get("value")
            elif op in ["less_than_or_equal", "less_than"]:
                max_val = cond.get("value")
            elif op == "regex_match":
                regex_units.append({
                    "column": col,
                    "validation_type": "regex",
                    "pattern": cond.get("value"),
                })
            elif op == "null_percentage":
                null_units.append({
                    "column": col,
                    "validation_type": "null_percentage",
                    "value": cond.get("value"),
                })
            elif op == "in_list":
                ref_units.append({
                    "column": col,
                    "validation_type": "in_list",
                    "value": cond.get("value"),
                })
        if (min_val not in [None, ""] or max_val not in [None, ""]):
            range_units.append({
                "column": col,
                "validation_type": "range",
                "min": min_val,
                "max": max_val,
            })

    periodicity_units = [
        {
            "column": col,
            "validation_type": "periodicity",
            "periodicity": freq,
        }
        for col, freq in column_frequencies.items()
    ]

    existing_units = get_validation_units_from_mongo(validator_atom_id, file_key)
    other_units = []
    if existing_units and "validations" in existing_units:
        other_units = [
            u
            for u in existing_units["validations"]
            if u.get("validation_type")
            not in ["range", "periodicity", "regex", "null_percentage", "in_list"]
        ]
    save_validation_units_to_mongo(
        validator_atom_id,
        file_key,
        other_units
        + range_units
        + periodicity_units
        + regex_units
        + null_units
        + ref_units,
    )

    client_id = os.getenv("CLIENT_ID", "")
    app_id = os.getenv("APP_ID", "")
    project_id = os.getenv("PROJECT_ID", "")
    cache_master_config(client_id, app_id, project_id, file_key, config_data)
    print(
        f"📦 Stored in redis namespace {client_id}:{app_id}:{project_id}:{file_key}"
    )

    message = f"Validation config configured successfully for file key '{file_key}' with {total_conditions} conditions"
    if column_frequencies:
        message += f" and frequencies specified for columns: {list(column_frequencies.keys())}"

    return {
        "status": "success",
        "message": message,
        "validator_atom_id": validator_atom_id,
        "file_key": file_key,
        "mongo_id": mongo_result.get("mongo_id", ""),
        "operation": mongo_result.get("operation", "unknown"),
        "total_conditions": total_conditions,
        "columns_configured": list(column_conditions.keys()),
        "columns_with_frequencies": list(column_frequencies.keys()),  # ✅ NEW
        "mongodb_saved": mongo_result["status"] == "success"
    }




# POST: VALIDATE - Enhanced validation with auto-correction, custom conditions, and MongoDB logging
@router.post("/validate", response_model=ValidateResponse)
async def validate(
    validator_atom_id: str = Form(...),
    files: List[UploadFile] | None = File(None),
    file_keys: str = Form(...),
    file_paths: str = Form(default=""),
    date_frequency: str = Form(default=None),
    user_id: str = Form(""),
    client_id: str = Form("")
):
    """
    Enhanced validation: mandatory columns + type check + auto-correction + custom conditions + MongoDB logging
    """
    start_time = perf_counter()
    try:
        keys = json.loads(file_keys)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON format for file_keys")
    if not isinstance(keys, list):
        raise HTTPException(status_code=400, detail="file_keys must be a JSON array")

    logger.info(
        "data_upload.validate.start validator=%s file_count=%s has_inline_files=%s",
        validator_atom_id,
        len(keys),
        bool(files),
    )

    try:
        paths = json.loads(file_paths) if file_paths else []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON format for file_paths")

    files_list = files or []

    if files_list and len(files_list) != len(keys):
        raise HTTPException(status_code=400, detail="Number of files must match number of keys")
    if paths and len(paths) != len(keys):
        raise HTTPException(status_code=400, detail="Number of file paths must match number of keys")
    if not files_list and not paths:
        raise HTTPException(status_code=400, detail="No files or file paths provided")

    if len(set(keys)) != len(keys):
        raise HTTPException(status_code=400, detail="Duplicate file keys are not allowed")
    for k in keys:
        if not isinstance(k, str) or not k.strip() or not FILE_KEY_RE.match(k):
            raise HTTPException(status_code=400, detail=f"Malformed file key: {k}")

    if files_list and len(files_list) > 3:
        raise HTTPException(status_code=400, detail="Maximum 3 files allowed")
    
    file_payloads: List[Dict[str, str]] | None = None
    if files_list:
        file_payloads = []
        for file, key in zip(files_list, keys):
            try:
                content = await file.read()
            except Exception as exc:  # pragma: no cover - defensive programming
                logger.exception("Failed to read uploaded file %s", file.filename)
                raise HTTPException(status_code=400, detail=f"Error reading file {file.filename}: {exc}")
            file_payloads.append(
                {
                    "key": key,
                    "filename": file.filename,
                    "content_b64": base64.b64encode(content).decode("utf-8"),
                }
            )

    submission = celery_task_client.submit_callable(
        name="data_upload_validate.validate",
        dotted_path="app.features.data_upload_validate.service.run_validation",
        kwargs={
            "validator_atom_id": validator_atom_id,
            "file_payloads": file_payloads,
            "file_paths": paths,
            "keys": keys,
            "date_frequency": date_frequency,
            "user_id": user_id,
            "client_id": client_id,
        },
        metadata={
            "feature": "data_upload_validate",
            "operation": "validate",
            "validator_atom_id": validator_atom_id,
            "file_count": len(keys),
        },
    )

    if submission.status == "failure":  # pragma: no cover - defensive
        logger.error(
            "data_upload.validate.failed validator=%s task_id=%s",
            validator_atom_id,
            submission.task_id,
        )
        raise HTTPException(status_code=400, detail=submission.detail or "Validation failed")

    duration_ms = (perf_counter() - start_time) * 1000
    logger.info(
        "data_upload.validate.queued validator=%s task_id=%s files=%s duration_ms=%.2f",
        validator_atom_id,
        submission.task_id,
        len(keys),
        duration_ms,
    )

    return format_task_response(submission, embed_result=True)

    
    



# DELETE: DELETE_VALIDATOR_ATOM - Delete a custom validator atom completely
@router.delete("/delete_validator_atom/{validator_atom_id}")
async def delete_validator_atom(validator_atom_id: str):
    """
    Delete a custom validator atom completely
    - Removes from custom_validations folder
    - Removes from mongodb folder (all related files)
    - Clears from memory
    """
    
    def delete_validator_atom_files(validator_atom_id: str):
        """Delete all files related to a validator atom from disk and memory"""
        # Paths
        custom_dir = Path("custom_validations")
        mongo_dir = Path("mongodb")
        deleted_files = []

        # Delete from custom_validations
        custom_file = custom_dir / f"{validator_atom_id}.json"
        if custom_file.exists():
            custom_file.unlink()
            deleted_files.append(str(custom_file))

        # Delete from mongodb folder - dimensions, assignments
        for suffix in ["business_dimensions", "identifier_assignments"]:
            mongo_file = mongo_dir / f"{validator_atom_id}_{suffix}.json"
            if mongo_file.exists():
                mongo_file.unlink()
                deleted_files.append(str(mongo_file))

        # Clear from memory
        if validator_atom_id in extraction_results:
            del extraction_results[validator_atom_id]
            deleted_files.append("memory_cleared")

        return deleted_files
    
    # Validate validator_atom_id
    if not validator_atom_id or not validator_atom_id.strip():
        raise HTTPException(status_code=400, detail="validator_atom_id cannot be empty")
    
    # Check if validator atom exists
    validator_exists = False
    custom_file = CUSTOM_CONFIG_DIR / f"{validator_atom_id}.json"
    mongo_dimensions = MONGODB_DIR / f"{validator_atom_id}_business_dimensions.json"
    mongo_assignments = MONGODB_DIR / f"{validator_atom_id}_identifier_assignments.json"
    
    if (custom_file.exists() or
        mongo_dimensions.exists() or
        mongo_assignments.exists() or
        validator_atom_id in extraction_results):
        validator_exists = True
    
    if not validator_exists:
        raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")
    
    # Delete all files and clear memory
    try:
        deleted_files = delete_validator_atom_files(validator_atom_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting validator atom: {str(e)}")
    
    return {
        "status": "success",
        "message": f"Validator atom '{validator_atom_id}' deleted completely",
        "validator_atom_id": validator_atom_id,
        "deleted_files": deleted_files,
        "deletion_summary": {
            "custom_validations_removed": any("custom_validations" in f for f in deleted_files),
            "mongodb_files_removed": any("mongodb" in f for f in deleted_files),
            "memory_cleared": "memory_cleared" in deleted_files,
            "total_files_deleted": len([f for f in deleted_files if f != "memory_cleared"])
        }
    }


# GET: GET_VALIDATOR_CONFIG - return validator setup with MongoDB details
@router.get("/get_validator_config/{validator_atom_id}")
async def get_validator_config(validator_atom_id: str):
    """Retrieve stored validator atom configuration along with any
    dimension information."""

    validator_data = get_validator_atom_from_mongo(validator_atom_id)
    if not validator_data:
        validator_data = get_validator_from_memory_or_disk(validator_atom_id)

    if not validator_data:
        raise HTTPException(status_code=404, detail=f"Validator atom '{validator_atom_id}' not found")

    extra = load_all_non_validation_data(validator_atom_id)

    validations = {}
    for key in validator_data.get("file_keys", []):
        units = get_validation_units_from_mongo(validator_atom_id, key)
        if units:
            validations[key] = units.get("validations", [])

    return {**validator_data, **extra, "validations": validations}


# Call this function when the module loads
load_existing_configs()

# --- New endpoints for saving and listing validated dataframes ---
# Accept both trailing and non-trailing slash variants and explicitly
# handle CORS preflight OPTIONS requests so browsers or proxies never
# receive a 405/500 before the actual POST is issued.
@router.options("/save_dataframes")
@router.options("/save_dataframes/")
async def save_dataframes_options() -> Response:
    return Response(status_code=204)

@router.post("/save_dataframes")
@router.post("/save_dataframes/")
async def save_dataframes(
    validator_atom_id: str = Form(...),
    files: List[UploadFile] | None = File(None),
    file_keys: str = Form(...),
    file_paths: str = Form(default=""),
    workbook_paths: str = Form(default=""),
    sheet_metadata: str = Form(default=""),
    overwrite: bool = Form(False),
    client_id: str = Form(""),
    user_id: str = Form(""),
    app_id: str = Form(""),
    project_id: str = Form(""),
    client_name: str = Form(""),
    app_name: str = Form(""),
    project_name: str = Form(""),
    user_name: str = Form(""),
):
    """Save validated dataframes as Arrow tables and upload via Flight."""
    start_time = perf_counter()
    logger.info(
        "data_upload.save_dataframes.start validator=%s overwrite=%s",
        validator_atom_id,
        overwrite,
    )
    logger.debug("raw file_keys=%s", file_keys)
    logger.debug("raw file_paths=%s", file_paths)

    # --- Parse and validate inputs -------------------------------------------------
    try:
        key_inputs = json.loads(file_keys) if file_keys else []
    except json.JSONDecodeError:
        logger.exception("Invalid JSON for file_keys")
        raise HTTPException(status_code=400, detail="Invalid JSON format for file_keys")
    if not isinstance(key_inputs, list):
        logger.error("file_keys not list: %s", type(key_inputs))
        raise HTTPException(status_code=400, detail="file_keys must be a JSON array")

    try:
        paths = json.loads(file_paths) if file_paths else []
    except json.JSONDecodeError:
        logger.exception("Invalid JSON for file_paths")
        raise HTTPException(status_code=400, detail="Invalid JSON format for file_paths")
    if paths and (
        not isinstance(paths, list)
        or any(not isinstance(p, str) or not p for p in paths)
    ):
        logger.error("file_paths malformed: %s", paths)
        raise HTTPException(
            status_code=400, detail="file_paths must be a JSON array of non-empty strings"
        )

    try:
        workbook_path_inputs = json.loads(workbook_paths) if workbook_paths else []
    except json.JSONDecodeError:
        logger.exception("Invalid JSON for workbook_paths")
        raise HTTPException(status_code=400, detail="Invalid JSON format for workbook_paths")
    if workbook_path_inputs and not isinstance(workbook_path_inputs, list):
        raise HTTPException(status_code=400, detail="workbook_paths must be a JSON array")

    try:
        sheet_metadata_inputs = json.loads(sheet_metadata) if sheet_metadata else []
    except json.JSONDecodeError:
        logger.exception("Invalid JSON for sheet_metadata")
        raise HTTPException(status_code=400, detail="Invalid JSON format for sheet_metadata")
    if sheet_metadata_inputs and not isinstance(sheet_metadata_inputs, list):
        raise HTTPException(status_code=400, detail="sheet_metadata must be a JSON array")

    files_list = files or []
    source_count = len(files_list) if files_list else len(paths)
    if source_count == 0:
        logger.error("No files or file paths provided")
        raise HTTPException(status_code=400, detail="No files or file paths provided")

    # Fallback to filenames when keys are missing or empty
    fallback_names = (
        [f.filename for f in files_list]
        if files_list
        else [Path(p).name for p in paths]
    )
    if len(key_inputs) == 0:
        keys = fallback_names
    else:
        if len(key_inputs) != source_count:
            logger.error(
                "Mismatched file key count: %s keys for %s sources",
                len(key_inputs),
                source_count,
            )
            raise HTTPException(
                status_code=400,
                detail="Number of file keys must match number of files or paths",
            )
        keys = []
        for i, k in enumerate(key_inputs):
            if not isinstance(k, str) or not k.strip():
                k = fallback_names[i]
            keys.append(k)

    if len(set(keys)) != len(keys):
        logger.error("Duplicate file keys: %s", keys)
        raise HTTPException(status_code=400, detail="Duplicate file keys are not allowed")

    # Validate file key format
    for k in keys:
        if not FILE_KEY_RE.match(k):
            logger.error("Malformed file key: %s", k)
            raise HTTPException(
                status_code=400,
                detail=f"Malformed file key: {k}",
            )

    normalized_workbook_paths = (
        [p if isinstance(p, str) else "" for p in workbook_path_inputs]
        if isinstance(workbook_path_inputs, list)
        else []
    )
    normalized_sheet_metadata = (
        [m if isinstance(m, dict) else {} for m in sheet_metadata_inputs]
        if isinstance(sheet_metadata_inputs, list)
        else []
    )

    uploads = []
    flights = []
    if client_id:
        os.environ["CLIENT_ID"] = client_id
    if app_id:
        os.environ["APP_ID"] = app_id
    if project_id:
        os.environ["PROJECT_ID"] = project_id
    if client_name:
        os.environ["CLIENT_NAME"] = client_name
    if app_name:
        os.environ["APP_NAME"] = app_name
    if project_name:
        os.environ["PROJECT_NAME"] = project_name
    prefix = await get_object_prefix()
    numeric_pid = _parse_numeric_id(project_id or os.getenv("PROJECT_ID", "0"))
    print(f"📤 saving to prefix {prefix}")

    tmp_prefix = prefix + "tmp/"
    iter_sources: List[Dict[str, Any]] = []
    if files_list:
        for idx, (k, f) in enumerate(zip(keys, files_list)):
            iter_sources.append(
                {
                    "key": k,
                    "filename": f.filename,
                    "fileobj": f.file,
                    "orig_path": None,
                    "workbook_path": normalized_workbook_paths[idx] if idx < len(normalized_workbook_paths) else "",
                    "sheet_meta": normalized_sheet_metadata[idx] if idx < len(normalized_sheet_metadata) else {},
                }
            )
    else:
        for idx, (k, p) in enumerate(zip(keys, paths)):
            data = read_minio_object(p)
            iter_sources.append(
                {
                    "key": k,
                    "filename": Path(p).name,
                    "fileobj": io.BytesIO(data),
                    "orig_path": p,
                    "workbook_path": normalized_workbook_paths[idx] if idx < len(normalized_workbook_paths) else "",
                    "sheet_meta": normalized_sheet_metadata[idx] if idx < len(normalized_sheet_metadata) else {},
                }
            )

    MAX_FILE_SIZE = 2 * 1024 * 1024 * 1024  # 2 GB - increased for large datasets
    STATUS_TTL = 3600

    for source in iter_sources:
        key = source["key"]
        filename = source["filename"]
        fileobj = source["fileobj"]
        orig_path = source.get("orig_path")
        workbook_path = source.get("workbook_path") or ""
        sheet_meta = source.get("sheet_meta") or {}
        if not isinstance(sheet_meta, dict):
            sheet_meta = {}
        logger.info("Processing file %s with key %s", filename, key)
        progress_key = f"upload_status:{validator_atom_id}:{key}"
        redis_client.set(progress_key, "uploading", ex=STATUS_TTL)

        arrow_name = Path(filename).stem + ".arrow"
        exists = await arrow_dataset_exists(numeric_pid, validator_atom_id, filename)
        if exists and not overwrite:
            uploads.append({"file_key": key, "already_saved": True})
            flights.append({"file_key": key})
            redis_client.set(progress_key, "saved", ex=STATUS_TTL)
            continue

        fileobj.seek(0, os.SEEK_END)
        size = fileobj.tell()
        fileobj.seek(0)
        if size > MAX_FILE_SIZE:
            redis_client.set(progress_key, "rejected", ex=STATUS_TTL)
            raise HTTPException(status_code=413, detail=f"{filename} exceeds 2GB limit (current size: {size / (1024*1024*1024):.2f}GB)")

        redis_client.set(progress_key, "parsing", ex=STATUS_TTL)

        if filename.lower().endswith(".csv"):
            csv_path = getattr(fileobj, "name", None)
            if csv_path and os.path.exists(csv_path):
                # CRITICAL: For batched reading, we need to scan first to find max columns
                # Read first batch to determine schema, then use that schema for all batches
                from app.features.data_upload_validate.file_ingestion.readers.csv_reader import CSVReader
                from app.features.data_upload_validate.file_ingestion.detectors.encoding_detector import EncodingDetector
                
                # Read file content to find max columns
                with open(csv_path, 'rb') as f:
                    content = f.read()
                encoding = EncodingDetector.detect(content)
                delimiter = CSVReader._detect_delimiter(content, encoding)
                max_cols = CSVReader._find_max_columns(content, encoding, delimiter, sample_rows=0)
                
                # Create schema with all columns
                if max_cols > 0:
                    schema = {f"col_{i}": pl.Utf8 for i in range(max_cols)}
                    batched_kwargs = CSV_READ_KWARGS.copy()
                    batched_kwargs["schema"] = schema
                    batched_kwargs["truncate_ragged_lines"] = False
                    batched_kwargs["ignore_errors"] = True  # Handle mixed dtype columns gracefully
                else:
                    batched_kwargs = CSV_READ_KWARGS.copy()
                    batched_kwargs["ignore_errors"] = True  # Handle mixed dtype columns gracefully
                
                reader = pl.read_csv_batched(
                    csv_path, batch_size=1_000_000, **batched_kwargs
                )
                try:
                    first_chunk = next(reader)
                except StopIteration:
                    uploads.append(
                        {
                            "file_key": key,
                            "already_saved": False,
                            "error": "empty file",
                        }
                    )
                    flights.append({"file_key": key})
                    continue
                # Normalize column names in first chunk
                first_chunk_normalized = data_upload_service._normalize_column_names(first_chunk)
                normalized_cols = first_chunk_normalized.columns
                original_cols = first_chunk.columns
                # Build rename mapping if columns changed
                rename_map = dict(zip(original_cols, normalized_cols)) if original_cols != normalized_cols else None
                
                arrow_buf = io.BytesIO()
                # Use PyArrow conversion to avoid "string_view" byte-range errors
                first_arrow = first_chunk_normalized.to_arrow(use_pyarrow=True)
                with pa.ipc.new_file(arrow_buf, first_arrow.schema) as writer:
                    writer.write(first_arrow)
                    for chunk in reader:
                        # Apply same normalization to subsequent chunks
                        if rename_map:
                            chunk = chunk.rename(rename_map)
                        else:
                            chunk = data_upload_service._normalize_column_names(chunk)
                        writer.write(chunk.to_arrow(use_pyarrow=True))
                arrow_bytes = arrow_buf.getvalue()
                df_pl = None
            else:
                data_bytes = fileobj.read()
                # Use pl.read_csv with CSV_READ_KWARGS for proper dtype inference
                # This matches the old routes behavior and preserves numeric types
                # Add ignore_errors to handle mixed dtype columns (e.g., "allpacksize" in numeric column)
                read_kwargs = CSV_READ_KWARGS.copy()
                read_kwargs["ignore_errors"] = True  # Convert unparseable values to null instead of failing
                df_pl = pl.read_csv(io.BytesIO(data_bytes), **read_kwargs)
                df_pl = data_upload_service._normalize_column_names(df_pl)
                arrow_buf = io.BytesIO()
                df_pl.write_ipc(arrow_buf)
                arrow_bytes = arrow_buf.getvalue()
        elif filename.lower().endswith((".xls", ".xlsx")):
            data_bytes = fileobj.read()
            reader = fastexcel.read_excel(data_bytes)
            sheet = reader.load_sheet_by_idx(0)
            df_pl = sheet.to_polars()
            df_pl = data_upload_service._normalize_column_names(df_pl)
            if df_pl.height == 0:
                uploads.append({"file_key": key, "already_saved": False, "error": "empty file"})
                flights.append({"file_key": key})
                continue
            arrow_buf = io.BytesIO()
            df_pl.write_ipc(arrow_buf)
            arrow_bytes = arrow_buf.getvalue()
        elif filename.lower().endswith(".arrow"):
            arrow_bytes = fileobj.read()
            df_pl = pl.read_ipc(io.BytesIO(arrow_bytes))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")

        # Check if this is a multi-sheet Excel file and should be saved in folder structure
        use_folder_structure = False
        excel_folder_name = None
        if workbook_path and sheet_meta:
            sheet_names_meta = sheet_meta.get("sheet_names", [])
            has_multiple = len(sheet_names_meta) > 1 if isinstance(sheet_names_meta, list) else False
            if has_multiple:
                # Extract Excel filename without extension for folder name
                original_filename_meta = sheet_meta.get("original_filename") or filename
                excel_folder_name = Path(original_filename_meta).stem.replace(' ', '_').replace('.', '_')
                use_folder_structure = True
                # Modify arrow_name to use folder structure: {excel_folder_name}/sheets/{sheet_name}.arrow
                selected_sheet_meta = sheet_meta.get("selected_sheet") or (sheet_names_meta[0] if sheet_names_meta else "")
                if selected_sheet_meta:
                    # Normalize sheet name for path
                    normalized_sheet_name = normalize_sheet_name(selected_sheet_meta)
                    arrow_name = f"{excel_folder_name}/sheets/{normalized_sheet_name}.arrow"
        
        result = upload_to_minio(arrow_bytes, arrow_name, prefix)
        saved_name = Path(result.get("object_name", "")).name or arrow_name
        flight_path = f"{validator_atom_id}/{saved_name}"
        logger.info("Uploaded %s as %s (folder_structure=%s)", filename, result.get("object_name", ""), use_folder_structure)

        # If df_pl is None (chunked csv), upload via polars scan
        if filename.lower().endswith(".csv"):
            if df_pl is None:
                reader_for_flight = pl.read_ipc(io.BytesIO(arrow_bytes))
                upload_dataframe(reader_for_flight.to_pandas(), flight_path)
            else:
                upload_dataframe(df_pl.to_pandas(), flight_path)
        else:
            upload_dataframe(df_pl.to_pandas(), flight_path)

        set_ticket(
            key,
            result.get("object_name", ""),
            flight_path,
            filename,
        )
        redis_client.set(f"flight:{flight_path}", result.get("object_name", ""))

        await record_arrow_dataset(
            numeric_pid,
            validator_atom_id,
            key,
            result.get("object_name", ""),
            flight_path,
            filename,
        )

        if workbook_path:
            sheet_names_meta = sheet_meta.get("sheet_names")
            if not isinstance(sheet_names_meta, list):
                sheet_names_meta = []
            selected_sheet_meta = sheet_meta.get("selected_sheet")
            if not isinstance(selected_sheet_meta, str) or not selected_sheet_meta:
                selected_sheet_meta = sheet_names_meta[0] if sheet_names_meta else ""
            original_filename_meta = sheet_meta.get("original_filename") or filename
            extension = Path(workbook_path).suffix if isinstance(workbook_path, str) else ""
            workbook_dest = f"{result.get('object_name', '')}.workbook{extension}"
            try:
                minio_client.copy_object(
                    MINIO_BUCKET,
                    workbook_dest,
                    CopySource(MINIO_BUCKET, workbook_path),
                )
                metadata_payload = {
                    "workbook_path": workbook_dest,
                    "sheet_names": sheet_names_meta,
                    "selected_sheet": selected_sheet_meta,
                    "has_multiple_sheets": len(sheet_names_meta) > 1,
                    "validator_atom_id": validator_atom_id,
                    "file_key": key,
                    "flight_path": flight_path,
                    "original_filename": original_filename_meta,
                }
                _save_workbook_metadata(result.get("object_name", ""), metadata_payload)
                if isinstance(workbook_path, str) and workbook_path.startswith(tmp_prefix):
                    try:
                        minio_client.remove_object(MINIO_BUCKET, workbook_path)
                    except Exception:
                        logger.warning("Failed to remove temp workbook %s", workbook_path)
            except S3Error as exc:
                logger.warning("Failed to persist workbook %s: %s", workbook_path, exc)
            except Exception as exc:
                logger.warning("Failed to persist workbook %s: %s", workbook_path, exc)
        else:
            _remove_workbook_metadata(result.get("object_name", ""))

        redis_client.set(progress_key, "saved", ex=STATUS_TTL)
        # Remove temporary upload if it exists
        if orig_path and orig_path.startswith(tmp_prefix):
            try:
                minio_client.remove_object(MINIO_BUCKET, orig_path)
            except Exception:
                logger.warning("Failed to remove temp object %s", orig_path)

        uploads.append({
            "file_key": key,
            "filename": arrow_name,
            "minio_upload": result,
            "already_saved": False,
        })
        flights.append({"file_key": key, "flight_path": flight_path})

    env = {
        "CLIENT_NAME": os.getenv("CLIENT_NAME"),
        "APP_NAME": os.getenv("APP_NAME"),
        "PROJECT_NAME": os.getenv("PROJECT_NAME"),
    }
    logger.info("save_dataframes completed: %s files", len(uploads))
    log_operation_to_mongo(
        user_id=user_id,
        client_id=client_id,
        validator_atom_id=validator_atom_id,
        operation="save_dataframes",
        details={"files_saved": uploads, "prefix": prefix},
        user_name=user_name,
        client_name=client_name,
        app_id=app_id,
        app_name=app_name,
        project_id=project_id,
        project_name=project_name,
    )
    response_payload = {
        "minio_uploads": uploads,
        "flight_uploads": flights,
        "prefix": prefix,
        "environment": env,
    }
    duration_ms = (perf_counter() - start_time) * 1000
    logger.info(
        "data_upload.save_dataframes.completed validator=%s uploads=%s flights=%s duration_ms=%.2f",
        validator_atom_id,
        len(uploads),
        len(flights),
        duration_ms,
    )
    return response_payload


@router.get("/upload-status/{validator_atom_id}/{file_key}")
async def get_upload_status(validator_atom_id: str, file_key: str) -> dict:
    progress_key = f"upload_status:{validator_atom_id}:{file_key}"
    status = redis_client.get(progress_key)
    if isinstance(status, bytes):
        status = status.decode()
    logger.info(
        "data_upload.upload_status validator=%s file_key=%s status=%s",
        validator_atom_id,
        file_key,
        status or "unknown",
    )
    return {"status": status}


_TIMESTAMP_PATTERN = re.compile(r"(\d{8})_(\d{6})")


def _normalize_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        try:
            return value.replace(tzinfo=timezone.utc)
        except Exception:
            return value
    try:
        return value.astimezone(timezone.utc)
    except Exception:
        return value


def _extract_timestamp_from_string(value: str | None) -> datetime | None:
    if not value:
        return None
    match = _TIMESTAMP_PATTERN.search(Path(value).name)
    if not match:
        return None
    try:
        parsed = datetime.strptime(
            f"{match.group(1)}{match.group(2)}", "%Y%m%d%H%M%S"
        )
    except ValueError:
        return None
    return parsed.replace(tzinfo=timezone.utc)


def _stat_object_metadata(object_name: str) -> tuple[datetime | None, int | None]:
    try:
        stat = minio_client.stat_object(MINIO_BUCKET, object_name)
    except S3Error as exc:
        code = getattr(exc, "code", "")
        if code in {"NoSuchKey", "NoSuchBucket"}:
            return None, None
        logger.warning("stat_object failed for %s: %s", object_name, exc)
        return None, None
    except Exception as exc:
        logger.warning("stat_object error for %s: %s", object_name, exc)
        return None, None
    last_modified = _normalize_datetime(getattr(stat, "last_modified", None))
    size = getattr(stat, "size", None)
    return last_modified, size if isinstance(size, int) else None


def _choose_newest_candidate(
    current: dict[str, Any] | None, candidate: dict[str, Any]
) -> dict[str, Any]:
    if current is None:
        return candidate
    cand_ts = candidate.get("timestamp")
    curr_ts = current.get("timestamp")
    if cand_ts and curr_ts:
        if cand_ts > curr_ts:
            return candidate
        if cand_ts < curr_ts:
            return current
    elif cand_ts and not curr_ts:
        return candidate
    elif curr_ts and not cand_ts:
        return current
    cand_mod = candidate.get("last_modified")
    curr_mod = current.get("last_modified")
    if cand_mod and curr_mod:
        if cand_mod > curr_mod:
            return candidate
        if cand_mod < curr_mod:
            return current
    elif cand_mod and not curr_mod:
        return candidate
    elif curr_mod and not cand_mod:
        return current
    cand_priority = candidate.get("priority", 0)
    curr_priority = current.get("priority", 0)
    if cand_priority != curr_priority:
        return candidate if cand_priority > curr_priority else current
    cand_name = candidate.get("object_name", "")
    curr_name = current.get("object_name", "")
    return candidate if cand_name > curr_name else current


@router.get("/latest_project_dataframe")
async def latest_project_dataframe(
    client_name: str = "",
    app_name: str = "",
    project_name: str = "",
    client_id: str = "",
    app_id: str = "",
    project_id: str = "",
) -> dict[str, Any]:
    prefix, env, env_source = await get_object_prefix(
        client_id=client_id,
        app_id=app_id,
        project_id=project_id,
        client_name=client_name,
        app_name=app_name,
        project_name=project_name,
        include_env=True,
    )

    best: dict[str, Any] | None = None
    arrow_names: set[str] = set()
    arrow_names.update(name for name in CSV_TO_FLIGHT.keys() if isinstance(name, str))
    arrow_names.update(
        value for value in FILEKEY_TO_CSV.values() if isinstance(value, str)
    )

    for arrow_name in arrow_names:
        original = get_original_csv(arrow_name)
        if not arrow_name.startswith(prefix) and not (
            original and original.startswith(prefix)
        ):
            continue
        flight_path = CSV_TO_FLIGHT.get(arrow_name) or get_flight_path_for_csv(
            arrow_name
        )
        timestamp = _extract_timestamp_from_string(arrow_name) or _extract_timestamp_from_string(
            flight_path
        )
        last_modified, size = _stat_object_metadata(arrow_name)
        candidate = {
            "object_name": arrow_name,
            "csv_name": original or Path(arrow_name).name,
            "flight_path": flight_path,
            "timestamp": timestamp or last_modified,
            "last_modified": last_modified,
            "size": size,
            "priority": 2 if flight_path else 1,
            "source": "flight_registry" if flight_path else "registry",
        }
        best = _choose_newest_candidate(best, candidate)

    list_error: str | None = None
    objects: list[Any] = []
    try:
        objects = list(
            minio_client.list_objects(
                MINIO_BUCKET, prefix=prefix, recursive=True
            )
        )
    except S3Error as exc:
        if getattr(exc, "code", "") == "NoSuchBucket":
            objects = []
        else:
            list_error = str(exc)
            objects = []
    except Exception as exc:
        list_error = str(exc)
        objects = []

    tmp_prefix = prefix + "tmp/"
    for obj in objects:
        object_name = getattr(obj, "object_name", "")
        if not object_name.endswith(".arrow"):
            continue
        if object_name.startswith(tmp_prefix):
            continue
        last_modified = _normalize_datetime(getattr(obj, "last_modified", None))
        size = obj.size if isinstance(obj.size, int) else None
        flight_path = get_flight_path_for_csv(object_name)
        timestamp = _extract_timestamp_from_string(object_name) or _extract_timestamp_from_string(
            flight_path
        )
        candidate = {
            "object_name": object_name,
            "csv_name": get_original_csv(object_name) or Path(object_name).name,
            "flight_path": flight_path,
            "timestamp": timestamp or last_modified,
            "last_modified": last_modified,
            "size": size,
            "priority": 2 if flight_path else 1,
            "source": "minio_flight" if flight_path else "minio",
        }
        previous = best
        best = _choose_newest_candidate(best, candidate)
        if previous is best and best and best.get("object_name") == object_name:
            if best.get("last_modified") is None and last_modified:
                best["last_modified"] = last_modified
            if best.get("size") is None and size is not None:
                best["size"] = size
            if best.get("flight_path") is None and flight_path:
                best["flight_path"] = flight_path
            if best.get("csv_name") in (None, "", Path(object_name).name):
                original = get_original_csv(object_name)
                if original:
                    best["csv_name"] = original

    if best and (best.get("last_modified") is None or best.get("size") is None):
        stat_modified, stat_size = _stat_object_metadata(best["object_name"])
        if best.get("last_modified") is None and stat_modified:
            best["last_modified"] = stat_modified
        if best.get("size") is None and stat_size is not None:
            best["size"] = stat_size

    if best:
        logger.info(
            "latest_project_dataframe resolved %s via %s",
            best.get("object_name"),
            best.get("source"),
        )

    response: dict[str, Any] = {
        "bucket": MINIO_BUCKET,
        "prefix": prefix,
        "environment": env,
        "env_source": env_source,
    }
    if best:
        response["object_name"] = best.get("object_name")
        response["csv_name"] = best.get("csv_name") or Path(
            best["object_name"]
        ).name
        if best.get("flight_path"):
            response["flight_path"] = best.get("flight_path")
        if best.get("last_modified"):
            response["last_modified"] = best["last_modified"].isoformat()
        if best.get("timestamp"):
            response["timestamp"] = best["timestamp"].isoformat()
        if best.get("size") is not None:
            response["size"] = best.get("size")
        if best.get("source"):
            response["source"] = best.get("source")
    else:
        response["object_name"] = None
        if list_error:
            response["error"] = list_error

    return response


# =============================================================================
# AUTO-CLASSIFICATION HELPER FUNCTIONS
# =============================================================================

async def _file_has_classification(
    object_name: str,
    client_name: str,
    app_name: str,
    project_name: str,
) -> tuple[bool, dict | None]:
    """Check if a file already has classification in MongoDB. Returns (has_classification, config_dict)."""
    try:
        config = get_classifier_config_from_mongo(
            client_name, app_name, project_name, file_name=object_name
        )
        return (config is not None, config)
    except Exception as e:
        logger.warning(f"Error checking classification for {object_name}: {e}")
        return (False, None)


def _classify_column(
    col: str,
    col_type: str,
    identifier_keywords: list[str],
    measure_keywords: list[str],
) -> str:
    """Classify a single column. Returns 'identifiers', 'measures', or 'unclassified'.
    
    Logic:
    1. If column name matches identifier keywords → identifiers
    2. If column name matches measure keywords → measures
    3. If datetime type → identifiers
    4. If categorical/string/object type → identifiers
    5. If numerical type → measures
    6. Else → unclassified
    """
    col_lower = col.lower()
    
    # Check keyword matches first
    if any(keyword in col_lower for keyword in identifier_keywords):
        return "identifiers"
    elif any(keyword in col_lower for keyword in measure_keywords):
        return "measures"
    
    # If no keyword match, classify by data type
    # Datetime → identifiers
    elif "datetime" in col_type.lower() or col_type in ["datetime64[ns]", "datetime64", "date"]:
        return "identifiers"
    # Categorical/string/object → identifiers
    elif col_type in ["object", "category", "string"]:
        return "identifiers"
    # Numerical → measures
    elif "int" in col_type.lower() or "float" in col_type.lower() or col_type in ["numeric", "integer", "float64", "float32", "int64", "int32"]:
        return "measures"
    else:
        return "unclassified"


async def _auto_classify_and_save_file(
    object_name: str,
    client_name: str,
    app_name: str,
    project_name: str,
    project_id: int | None = None,
    existing_config: dict | None = None,
) -> None:
    """Auto-classify a file and save to MongoDB (same as column classifier atom).
    
    If existing_config is provided, only classifies new columns and merges with existing.
    """
    try:
        # Load dataframe from MinIO
        if not object_name.endswith(".arrow"):
            return
        
        try:
            response = minio_client.get_object(MINIO_BUCKET, object_name)
            content = response.read()
        except S3Error as e:
            code = getattr(e, "code", "")
            if code in {"NoSuchKey", "NoSuchBucket"}:
                logger.warning(f"File not found for classification: {object_name}")
                return
            raise
        
        # Parse Arrow file
        reader = ipc.RecordBatchFileReader(pa.BufferReader(content))
        df = reader.read_all().to_pandas()
        
        # Normalize column names to lowercase
        df.columns = [str(c).strip().lower() for c in df.columns]
        all_columns = df.columns.tolist()
        column_types = {c: str(df[c].dtype) for c in df.columns}
        
        # AUTO-CLASSIFY keywords (same as classify_columns endpoint)
        # identifier_keywords = [
        #     'id', 'name', 'brand', 'market', 'category', 'region', 'channel', 
        #     'date', 'time', 'year', 'week', 'month', 'variant', 'ppg', 'type', 
        #     'code', 'packsize', 'packtype'
        # ]
        # measure_keywords = [
        #     'sales', 'revenue', 'volume', 'amount', 'value', 'price', 'cost', 
        #     'profit', 'units', 'd1', 'd2', 'd3', 'd4', 'd5', 'd6', 
        #     'salesvalue', 'baseprice', 'promoprice'
        # ]

        identifier_keywords = [
            'id', 'name', 'brand', 'market', 'category', 'region', 'channel', 
            'date', 'time', 'year', 'week', 'month', 'variant', 'ppg', 'type', 
            'code', 'packsize', 'packtype',"sku","product",
            "segment","subsegment","subchannel","zone","state","city","cluster","store","retailer","distributor","partner","account",
            "customer","consumer","household","respondent","wave","period","quarter","day"]


        measure_keywords = [
            'sales', 'revenue', 'volume', 'amount', 'value', 'price', 'cost', 
            'profit', 'units', 'd1', 'd2', 'd3', 'd4', 'd5', 'd6', 
            'salesvalue', 'baseprice', 'promoprice',
            "sale","qty","quantity", "mrp","nrv","margin","loss","rate","spend","impressions","clicks","carts","orders","views","shares","likes",
            "comments","ratings","scores","awareness","consideration","preference","nps","penetration","frequency","reach","trps","grps","weight","index","share"]
        
        # If existing config exists, merge with existing classification
        if existing_config:
            # Normalize existing classification columns to lowercase (same as column classifier atom)
            existing_identifiers = set([str(c).strip().lower() for c in existing_config.get("identifiers", [])])
            existing_measures = set([str(c).strip().lower() for c in existing_config.get("measures", [])])
            # Unclassified might not exist in old configs, so default to empty set
            existing_unclassified = set([str(c).strip().lower() for c in existing_config.get("unclassified", [])])
            existing_all = existing_identifiers | existing_measures | existing_unclassified
            
            # Find new columns (all_columns are already normalized to lowercase)
            all_columns_set = set(all_columns)
            new_columns = all_columns_set - existing_all
            
            # Filter existing classification to only include columns that still exist in the file
            # This removes columns that were deleted from the file
            existing_identifiers = existing_identifiers & all_columns_set
            existing_measures = existing_measures & all_columns_set
            existing_unclassified = existing_unclassified & all_columns_set
            
            # Classify only new columns (if any)
            new_identifiers = []
            new_measures = []
            new_unclassified = []
            
            if new_columns:
                for col in new_columns:
                    col_type = column_types.get(col, "string")
                    classification = _classify_column(col, col_type, identifier_keywords, measure_keywords)
                    
                    if classification == "identifiers":
                        new_identifiers.append(col)
                    elif classification == "measures":
                        new_measures.append(col)
                    else:
                        new_unclassified.append(col)
            
            # Merge with existing classification (existing columns are already filtered to only those in current file)
            final_identifiers = list(existing_identifiers | set(new_identifiers))
            final_measures = list(existing_measures | set(new_measures))
            final_unclassified = list(existing_unclassified | set(new_unclassified))
        else:
            # Full classification for new file
            final_identifiers = []
            final_measures = []
            final_unclassified = []
            
            for col in all_columns:
                col_type = column_types.get(col, "string")
                classification = _classify_column(col, col_type, identifier_keywords, measure_keywords)
                
                if classification == "identifiers":
                    final_identifiers.append(col)
                elif classification == "measures":
                    final_measures.append(col)
                else:
                    final_unclassified.append(col)
        
        # Save to MongoDB using same pattern as column classifier atom
        config_data = {
            "project_id": project_id,
            "client_name": client_name,
            "app_name": app_name,
            "project_name": project_name,
            "identifiers": final_identifiers,
            "measures": final_measures,
            "unclassified": final_unclassified,
            "dimensions": {},  # Empty dimensions object (same as column classifier)
            "file_name": object_name,
        }
        
        # Get environment variables (same as save_config endpoint)
        env = await get_env_vars(
            client_name=client_name,
            app_name=app_name,
            project_name=project_name,
        )
        if env:
            config_data["env"] = env
        
        # Save to MongoDB
        mongo_result = save_classifier_config_to_mongo(config_data)
        logger.info(f"✅ Auto-classified and saved: {object_name} | {len(final_identifiers)} identifiers, {len(final_measures)} measures, {len(final_unclassified)} unclassified")
            
    except Exception as e:
        logger.error(f"❌ Auto-classification failed for {object_name}: {e}", exc_info=True)


async def _background_auto_classify_files(
    files: list[dict],
    env: dict,
    client_name: str,
    app_name: str,
    project_name: str,
) -> None:
    """Background task to auto-classify files that don't have classification."""
    try:
        project_id = None
        try:
            project_id_str = env.get("PROJECT_ID")
            if project_id_str:
                project_id = int(project_id_str)
        except (ValueError, TypeError):
            pass
        
        logger.info(f"🔄 Background auto-classification started for {len(files)} files")
        
        for file_entry in files:
            object_name = file_entry.get("object_name")
            if not object_name:
                continue
            
            # Check if file already has classification
            has_classification, existing_config = await _file_has_classification(
                object_name, client_name, app_name, project_name
            )
            
            if has_classification:
                logger.info(f"🔍 Re-classifying (merge mode): {object_name}")
                # Check for new columns and classify only those
                await _auto_classify_and_save_file(
                    object_name,
                    client_name,
                    app_name,
                    project_name,
                    project_id,
                    existing_config=existing_config,
                )
            else:
                logger.info(f"🆕 Auto-classifying (new file): {object_name}")
                # Auto-classify in background (full classification)
                await _auto_classify_and_save_file(
                    object_name,
                    client_name,
                    app_name,
                    project_name,
                    project_id,
                    existing_config=None,
                )
        
        logger.info(f"✅ Background auto-classification completed for {len(files)} files")
    except Exception as e:
        logger.error(f"❌ Background auto-classification failed: {e}", exc_info=True)


@router.get("/list_saved_dataframes")
async def list_saved_dataframes(
    client_name: str = "",
    app_name: str = "",
    project_name: str = "",
) -> dict:
    """List all objects stored under the client/app/project prefix.

    Previously this endpoint returned only the latest ``.arrow`` file for each
    dataset which meant any additional files or nested directories inside the
    user's namespace were ignored by the UI. The Saved DataFrames panel now
    expects a complete listing so it can render a tree view of folders and
    files. To support this we simply return every object MinIO reports for the
    resolved prefix.
    """

    prefix, env, env_source = await get_object_prefix(
        client_name=client_name,
        app_name=app_name,
        project_name=project_name,
        include_env=True,
    )

    try:
        # print(
        #     f"🪣 listing from bucket '{MINIO_BUCKET}' prefix '{prefix}' (source={env_source})"
        # )  # Disabled
        objects = list(
            minio_client.list_objects(
                MINIO_BUCKET, prefix=prefix, recursive=True
            )
        )
        tmp_prefix = prefix + "tmp/"
        uploads_prefix = prefix + "uploads/"
        files = []
        excel_folders: Dict[str, Dict[str, Any]] = {}
        
        for obj in sorted(objects, key=lambda o: o.object_name):
            if not obj.object_name.endswith(".arrow"):
                continue
            if obj.object_name.startswith(tmp_prefix):
                continue
            if obj.object_name.startswith(uploads_prefix):
                continue  # Skip temporary upload folders
            
            last_modified = getattr(obj, "last_modified", None)
            if last_modified is not None:
                try:
                    modified_iso = last_modified.isoformat()
                except Exception:
                    modified_iso = None
            else:
                modified_iso = None
            
            # Check if this is part of an Excel folder structure: {prefix}{excel_name}/sheets/{sheet_name}.arrow
            rel_path = obj.object_name[len(prefix):] if obj.object_name.startswith(prefix) else obj.object_name
            path_parts = rel_path.split("/")
            
            if len(path_parts) >= 3 and path_parts[1] == "sheets":
                # This is a sheet inside an Excel folder
                excel_folder_name = path_parts[0]
                sheet_name = Path(path_parts[-1]).stem  # Remove .arrow extension
                
                if excel_folder_name not in excel_folders:
                    excel_folders[excel_folder_name] = {
                        "name": excel_folder_name,
                        "path": f"{prefix}{excel_folder_name}/",
                        "type": "excel_folder",
                        "sheets": []
                    }
                
                # Get original filename from flight registry to preserve original name
                original_csv_name = get_original_csv(obj.object_name)
                display_name = original_csv_name if original_csv_name else Path(obj.object_name).name
                # If display_name ends with .arrow, try to get original from registry
                if display_name.endswith('.arrow'):
                    if original_csv_name:
                        # We have original_csv_name from registry, use it
                        display_name = Path(original_csv_name).name
                    else:
                        # Try to get original from FILEKEY_TO_CSV mapping
                        from app.DataStorageRetrieval.flight_registry import FILEKEY_TO_CSV
                        original_from_registry = FILEKEY_TO_CSV.get(obj.object_name)
                        if original_from_registry:
                            display_name = Path(original_from_registry).name
                        else:
                            # Remove .arrow extension for display as fallback
                            display_name = Path(obj.object_name).stem
                
                sheet_entry = {
                    "object_name": obj.object_name,
                    "sheet_name": sheet_name,
                    "arrow_name": Path(obj.object_name).name,
                    "csv_name": display_name,  # Use original filename for display
                }
                if modified_iso:
                    sheet_entry["last_modified"] = modified_iso
                size = getattr(obj, "size", None)
                if isinstance(size, int):
                    sheet_entry["size"] = size
                
                excel_folders[excel_folder_name]["sheets"].append(sheet_entry)
            else:
                # Regular file (not in Excel folder structure)
                # Get original filename from flight registry to preserve original name
                original_csv_name = get_original_csv(obj.object_name)
                display_name = original_csv_name if original_csv_name else Path(obj.object_name).name
                # If display_name ends with .arrow, try to get original from registry
                if display_name.endswith('.arrow'):
                    if original_csv_name:
                        # We have original_csv_name from registry, use it
                        display_name = Path(original_csv_name).name
                    else:
                        # Try to get original from FILEKEY_TO_CSV mapping
                        from app.DataStorageRetrieval.flight_registry import FILEKEY_TO_CSV
                        original_from_registry = FILEKEY_TO_CSV.get(obj.object_name)
                        if original_from_registry:
                            display_name = Path(original_from_registry).name
                        else:
                            # Remove .arrow extension for display as fallback
                            display_name = Path(obj.object_name).stem
                
                entry = {
                    "object_name": obj.object_name,
                    "arrow_name": Path(obj.object_name).name,
                    "csv_name": display_name,  # Use original filename for display
                }
                if modified_iso:
                    entry["last_modified"] = modified_iso
                size = getattr(obj, "size", None)
                if isinstance(size, int):
                    entry["size"] = size
                files.append(entry)
        
        # Trigger background auto-classification for files
        asyncio.create_task(
            _background_auto_classify_files(
                files=files,
                env=env,
                client_name=client_name,
                app_name=app_name,
                project_name=project_name,
            )
        )
        
        result = {
            "bucket": MINIO_BUCKET,
            "prefix": prefix,
            "files": files,
            "excel_folders": list(excel_folders.values()),
            "environment": env,
            "env_source": env_source,
        }
        logger.info(
            "list_saved_dataframes result: files=%s excel_folders=%s",
            len(files),
            len(excel_folders)
        )
        return result
    except S3Error as e:
        if getattr(e, "code", "") == "NoSuchBucket":
            return {
                "bucket": MINIO_BUCKET,
                "prefix": prefix,
                "files": [],
                "excel_folders": [],
                "environment": env,
            }
        return {
            "bucket": MINIO_BUCKET,
            "prefix": prefix,
            "files": [],
            "excel_folders": [],
            "error": str(e),
            "environment": env,
        }
    except Exception as e:  # pragma: no cover - unexpected error
        return {
            "bucket": MINIO_BUCKET,
            "prefix": prefix,
            "files": [],
            "excel_folders": [],
            "error": str(e),
            "environment": env,
        }


@router.get("/latest_ticket/{file_key}")
async def latest_ticket(file_key: str):
    path, arrow_name = get_ticket_by_key(file_key)
    if path is None:
        path, arrow_name = get_latest_ticket_for_basename(file_key)
    if path is None:
        raise HTTPException(status_code=404, detail="Ticket not found")
    original = get_original_csv(arrow_name) or arrow_name
    return {
        "flight_path": path,
        "arrow_name": arrow_name,
        "csv_name": original,
    }


@router.get("/download_dataframe")
async def download_dataframe(object_name: str):
    """Return a presigned URL to download a dataframe"""
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    try:
        url = minio_client.presigned_get_object(MINIO_BUCKET, object_name)
        return {"url": url}
    except S3Error as e:
        if getattr(e, "code", "") in {"NoSuchKey", "NoSuchBucket"}:
            redis_client.delete(object_name)
            raise HTTPException(status_code=404, detail="File not found")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/download_dataframe_direct")
async def download_dataframe_direct(object_name: str):
    """
    Stream the dataframe file directly from MinIO (avoids presigned host resolution issues).
    """
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")

    try:
        obj = minio_client.get_object(MINIO_BUCKET, object_name)
        content = obj.read()
        obj.close()
        obj.release_conn()

        filename = Path(object_name).name
        return StreamingResponse(
          iter([content]),
          media_type="application/octet-stream",
          headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
          },
        )
    except S3Error as e:
        if getattr(e, "code", "") in {"NoSuchKey", "NoSuchBucket"}:
            redis_client.delete(object_name)
            raise HTTPException(status_code=404, detail="File not found")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/load-saved-dataframe")
async def load_saved_dataframe(
    request: Request,
):
    """
    Load a saved dataframe from MinIO into the data upload atom.
    This endpoint validates the file exists and returns metadata needed for the UI.
    This works alongside the existing file upload functionality - users can either:
    1. Upload files from their own source (existing functionality)
    2. Load files from MinIO using this endpoint (new functionality)
    
    Request body:
    {
        "object_name": "full/path/to/file.arrow",
        "client_name": "optional",
        "app_name": "optional",
        "project_name": "optional"
    }
    """
    from urllib.parse import unquote
    
    # Parse request body
    try:
        body = await request.json()
        object_name = body.get("object_name", "")
        client_name = body.get("client_name", "")
        app_name = body.get("app_name", "")
        project_name = body.get("project_name", "")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid request body: {str(e)}")
    
    if not object_name:
        raise HTTPException(status_code=400, detail="object_name is required")
    
    # Decode URL-encoded object_name
    object_name = unquote(object_name)
    
    # Set environment context if provided
    if client_name:
        os.environ["CLIENT_NAME"] = client_name
    if app_name:
        os.environ["APP_NAME"] = app_name
    if project_name:
        os.environ["PROJECT_NAME"] = project_name
    
    # Get the current prefix to validate object_name
    prefix = await get_object_prefix()
    
    # Validate object_name is within the allowed prefix
    if not object_name.startswith(prefix):
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid object name: file must be within the current project prefix"
        )
    
    # Check if file exists in MinIO
    try:
        # Try to get object metadata
        stat = minio_client.stat_object(MINIO_BUCKET, object_name)
        file_size = stat.size if hasattr(stat, 'size') else 0
        
        # Extract display name from object_name
        display_name = Path(object_name).name
        # Remove .arrow extension if present for display
        if display_name.endswith('.arrow'):
            csv_name = display_name[:-6]  # Remove .arrow
        else:
            csv_name = display_name
        
        # Get file metadata
        last_modified = None
        if hasattr(stat, 'last_modified'):
            try:
                last_modified = stat.last_modified.isoformat()
            except Exception:
                pass
        
        logger.info(
            "data_upload.load_saved_dataframe.success object_name=%s size=%s",
            object_name,
            file_size
        )
        
        return {
            "success": True,
            "object_name": object_name,
            "display_name": display_name,
            "csv_name": csv_name,
            "size": file_size,
            "last_modified": last_modified,
            "message": f"File '{csv_name}' is ready to be loaded into the data upload atom"
        }
        
    except S3Error as e:
        if getattr(e, "code", "") in {"NoSuchKey", "NoSuchBucket"}:
            logger.warning(
                "data_upload.load_saved_dataframe.not_found object_name=%s",
                object_name
            )
            raise HTTPException(status_code=404, detail=f"File not found: {object_name}")
        logger.error(
            "data_upload.load_saved_dataframe.error object_name=%s error=%s",
            object_name,
            str(e)
        )
        raise HTTPException(status_code=500, detail=f"Error accessing file: {str(e)}")
    except Exception as e:
        logger.error(
            "data_upload.load_saved_dataframe.exception object_name=%s error=%s",
            object_name,
            str(e)
        )
        raise HTTPException(status_code=400, detail=f"Error loading file: {str(e)}")


@router.get("/export_csv")
async def export_csv(object_name: str):
    """Export the saved dataframe as CSV file."""
    from urllib.parse import unquote
    import pyarrow.ipc as ipc
    
    object_name = unquote(object_name)
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    
    print(f"➡️ data_upload_validate export_csv request: {object_name}")
    
    try:
        # Try to get from Redis first
        content = redis_client.get(object_name)
        if content is None:
            # Get from MinIO
            try:
                response = minio_client.get_object(MINIO_BUCKET, object_name)
                content = response.read()
                redis_client.setex(object_name, 3600, content)
            except S3Error as e:
                if getattr(e, "code", "") in {"NoSuchKey", "NoSuchBucket"}:
                    raise HTTPException(status_code=404, detail="File not found")
                raise

        # Convert to DataFrame
        if object_name.endswith(".arrow"):
            reader = ipc.RecordBatchFileReader(pa.BufferReader(content))
            df = reader.read_all().to_pandas()
        else:
            # Use RobustFileReader to preserve all columns
            filename = Path(object_name).name
            df_result, _ = RobustFileReader.read_file_to_pandas(
                content=content,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
        
        # Convert to CSV bytes
        csv_bytes = df.to_csv(index=False).encode("utf-8")
        
        filename = object_name.split('/')[-1].replace('.arrow', '').replace('.csv', '') + '.csv'
        
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"⚠️ data_upload_validate export_csv error for {object_name}: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/export_excel")
async def export_excel(object_name: str):
    """Export the saved dataframe as Excel file."""
    from urllib.parse import unquote
    import pyarrow.ipc as ipc
    
    object_name = unquote(object_name)
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    
    print(f"➡️ data_upload_validate export_excel request: {object_name}")
    
    try:
        # Try to get from Redis first
        content = redis_client.get(object_name)
        if content is None:
            # Get from MinIO
            try:
                response = minio_client.get_object(MINIO_BUCKET, object_name)
                content = response.read()
                redis_client.setex(object_name, 3600, content)
            except S3Error as e:
                if getattr(e, "code", "") in {"NoSuchKey", "NoSuchBucket"}:
                    raise HTTPException(status_code=404, detail="File not found")
                raise

        # Convert to DataFrame
        if object_name.endswith(".arrow"):
            reader = ipc.RecordBatchFileReader(pa.BufferReader(content))
            df = reader.read_all().to_pandas()
        else:
            # Use RobustFileReader to preserve all columns
            filename = Path(object_name).name
            df_result, _ = RobustFileReader.read_file_to_pandas(
                content=content,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
        
        # Convert to Excel bytes
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False, engine='openpyxl')
        excel_bytes = excel_buffer.getvalue()
        
        filename = object_name.split('/')[-1].replace('.arrow', '').replace('.csv', '') + '.xlsx'
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        print(f"⚠️ data_upload_validate export_excel error for {object_name}: {e}")
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/delete_dataframe")
async def delete_dataframe(object_name: str):
    """Delete a single saved dataframe"""
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    csv_name = object_name.rsplit('.', 1)[0] + '.csv'
    try:
        try:
            minio_client.remove_object(MINIO_BUCKET, object_name)
        except S3Error as e:
            if getattr(e, "code", "") not in {"NoSuchKey", "NoSuchBucket"}:
                raise
        try:
            minio_client.remove_object(MINIO_BUCKET, csv_name)
        except S3Error as e:
            if getattr(e, "code", "") not in {"NoSuchKey", "NoSuchBucket"}:
                raise
        _delete_workbook_artifacts(object_name)
        redis_client.delete(object_name)
        redis_client.delete(csv_name)
        remove_arrow_object(object_name)
        await delete_arrow_dataset(object_name)
        mark_operation_log_deleted(object_name)
        return {"deleted": object_name}
    except S3Error as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/delete_all_dataframes")
async def delete_all_dataframes():
    """Delete all saved dataframes for the current project"""
    prefix = await get_object_prefix()
    deleted = []
    try:
        objects = list(minio_client.list_objects(MINIO_BUCKET, prefix=prefix, recursive=True))
        for obj in objects:
            obj_name = obj.object_name
            try:
                minio_client.remove_object(MINIO_BUCKET, obj_name)
            except S3Error as e:
                if getattr(e, "code", "") not in {"NoSuchKey", "NoSuchBucket"}:
                    raise
            redis_client.delete(obj_name)
            if obj_name.endswith('.arrow'):
                remove_arrow_object(obj_name)
                await delete_arrow_dataset(obj_name)
                mark_operation_log_deleted(obj_name)
            deleted.append(obj_name)
        return {"deleted": deleted}
    except S3Error as e:
        if getattr(e, "code", "") == "NoSuchBucket":
            return {"deleted": []}
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/rename_dataframe")
async def rename_dataframe(object_name: str = Form(...), new_filename: str = Form(...)):
    """Rename a saved dataframe"""
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    new_object = f"{prefix}{new_filename}"
    if new_object == object_name:
        # Nothing to do if the name hasn't changed
        return {"old_name": object_name, "new_name": object_name}
    try:
        minio_client.copy_object(
            MINIO_BUCKET,
            new_object,
            CopySource(MINIO_BUCKET, object_name),
        )
        try:
            minio_client.remove_object(MINIO_BUCKET, object_name)
        except S3Error:
            pass
        content = redis_client.get(object_name)
        if content is not None:
            redis_client.setex(new_object, 3600, content)
            redis_client.delete(object_name)
        rename_arrow_object(object_name, new_object)
        await rename_arrow_dataset(object_name, new_object)
        _copy_workbook_artifacts(object_name, new_object)
        return {"old_name": object_name, "new_name": new_object}
    except S3Error as e:
        code = getattr(e, "code", "")
        if code in {"NoSuchKey", "NoSuchBucket"}:
            redis_client.delete(object_name)
            raise HTTPException(status_code=404, detail="File not found")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/copy_dataframe")
async def copy_dataframe(object_name: str = Form(...), new_filename: str = Form(...)):
    """Copy a saved dataframe to a new file"""
    prefix = await get_object_prefix()
    if not object_name.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    new_object = f"{prefix}{new_filename}"
    if new_object == object_name:
        raise HTTPException(status_code=400, detail="New filename must be different from original")
    
    # Check if the new file already exists
    try:
        minio_client.stat_object(MINIO_BUCKET, new_object)
        raise HTTPException(status_code=400, detail="File with this name already exists")
    except S3Error as e:
        if e.code != "NoSuchKey":
            raise HTTPException(status_code=500, detail=str(e))
    
    try:
        from minio.commonconfig import CopySource
        # Copy the MinIO object
        minio_client.copy_object(
            MINIO_BUCKET,
            new_object,
            CopySource(MINIO_BUCKET, object_name),
        )
        
        # Copy Redis cache if it exists
        content = redis_client.get(object_name)
        if content is not None:
            redis_client.setex(new_object, 3600, content)
        
        # Copy arrow object registration (flight registry)
        try:
            from app.DataStorageRetrieval.flight_registry import get_arrow_for_flight_path, set_ticket, FILEKEY_TO_CSV
            original_arrow = get_arrow_for_flight_path(object_name)
            if original_arrow:
                # Register the new arrow object with the same arrow path
                csv_name = FILEKEY_TO_CSV.get(object_name, new_filename.replace('.arrow', '.csv'))
                set_ticket(new_object, csv_name, original_arrow)
        except Exception as e:
            logger.warning(f"Failed to copy arrow object registration: {e}")
        
        # Note: Arrow dataset registration in the database is typically handled
        # when the file is first accessed/validated, so we don't need to copy it here
        
        # Get file stats for response
        try:
            stat = minio_client.stat_object(MINIO_BUCKET, new_object)
            return {
                "old_name": object_name,
                "new_name": new_object,
                "last_modified": stat.last_modified.isoformat() if hasattr(stat.last_modified, 'isoformat') else str(stat.last_modified),
                "size": stat.size
            }
        except Exception:
            return {
                "old_name": object_name,
                "new_name": new_object,
                "last_modified": None,
                "size": None
            }
    except S3Error as e:
        code = getattr(e, "code", "")
        if code in {"NoSuchKey", "NoSuchBucket"}:
            raise HTTPException(status_code=404, detail="File not found")
        raise HTTPException(status_code=500, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/workbook_metadata")
async def get_workbook_metadata_endpoint(object_name: str = Query(...)):
    decoded = unquote(object_name)
    prefix = await get_object_prefix()
    if not decoded.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    try:
        metadata = _load_workbook_metadata(decoded)
    except FileNotFoundError:
        # Return default metadata for files without workbook metadata (e.g., single-sheet files)
        # This prevents 404 errors in logs for normal operation
        return {
            "has_multiple_sheets": False,
            "sheet_names": [],
            "selected_sheet": None,
            "workbook_path": None,
            "flight_path": None
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return metadata


@router.post("/change_sheet")
async def change_workbook_sheet(
    object_name: str = Form(...),
    sheet_name: str = Form(...),
):
    decoded = unquote(object_name)
    sheet_name_clean = sheet_name.strip()
    if not sheet_name_clean:
        raise HTTPException(status_code=400, detail="sheet_name is required")
    prefix = await get_object_prefix()
    if not decoded.startswith(prefix):
        raise HTTPException(status_code=400, detail="Invalid object name")
    try:
        metadata = _load_workbook_metadata(decoded)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Workbook metadata not found")

    workbook_path = metadata.get("workbook_path")
    if not workbook_path:
        raise HTTPException(status_code=400, detail="Workbook not associated with dataframe")

    try:
        workbook_bytes = read_minio_object(workbook_path)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read workbook: {exc}")

    try:
        # Use ExcelReader to preserve all columns (single approach, no duplication)
        from app.features.data_upload_validate.file_ingestion.readers.excel_reader import ExcelReader
        dfs_dict, excel_metadata = ExcelReader.read(
            content=workbook_bytes,
            sheet_name=sheet_name_clean,
            auto_detect_header=True,
            return_raw=False,
        )
        
        # Get sheet names from metadata
        sheet_names = excel_metadata.get("sheet_names", [])
        if sheet_name_clean not in sheet_names:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name_clean}' not found")
        
        # Handle both single DataFrame and dict (multiple sheets)
        if isinstance(dfs_dict, dict):
            df = dfs_dict[sheet_name_clean]
        else:
            df = dfs_dict
        df_pl = pl.from_pandas(df)
        arrow_buf = io.BytesIO()
        df_pl.write_ipc(arrow_buf)
        arrow_bytes = arrow_buf.getvalue()
        minio_client.put_object(
            MINIO_BUCKET,
            decoded,
            io.BytesIO(arrow_bytes),
            len(arrow_bytes),
            content_type="application/octet-stream",
        )
        flight_path = metadata.get("flight_path")
        if flight_path:
            upload_dataframe(df, flight_path)
        metadata["selected_sheet"] = sheet_name_clean
        metadata["sheet_names"] = sheet_names
        metadata["has_multiple_sheets"] = len(sheet_names) > 1
        _save_workbook_metadata(decoded, metadata)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to update sheet: {exc}")

    return {"status": "success", "selected_sheet": sheet_name_clean}


def _load_dataframe_for_processing(object_name: str, content: bytes) -> pd.DataFrame:
    if object_name.lower().endswith(".arrow"):
        reader = pa.ipc.RecordBatchFileReader(pa.BufferReader(content))
        table = reader.read_all()
        return table.to_pandas()
    if object_name.lower().endswith(".csv"):
        # Use RobustFileReader to preserve all columns
        filename = Path(object_name).name
        df_result, _ = RobustFileReader.read_file_to_pandas(
            content=content,
            filename=filename,
            auto_detect_header=True,
            return_raw=False,
        )
        # Handle both single DataFrame and dict (multiple sheets)
        if isinstance(df_result, dict):
            return list(df_result.values())[0]  # Use first sheet
        return df_result
    if object_name.lower().endswith(".parquet"):
        return pd.read_parquet(io.BytesIO(content))
    raise ValueError("Unsupported file format for processing")


def _cast_series_dtype(series: pd.Series, dtype: str, datetime_format: str | None = None) -> pd.Series:
    dtype_lower = dtype.lower()
    if dtype_lower in {"string", "str", "text", "object"}:
        return series.astype("string")
    if dtype_lower in {"int", "integer", "int64"}:
        numeric = pd.to_numeric(series, errors="coerce")
        return numeric.round().astype("Int64")
    if dtype_lower in {"float", "double", "float64"}:
        return pd.to_numeric(series, errors="coerce")
    if dtype_lower in {"bool", "boolean"}:
        return series.astype("boolean")
    if dtype_lower in {"datetime", "timestamp", "datetime64"}:
        # Normalize separators: replace all '/', '.' with '-' to handle mixed separators
        normalized_series = series.astype(str).str.replace('/', '-', regex=False).str.replace('.', '-', regex=False)
        # Normalize format string if provided (replace '/' and '.' with '-') to match normalized data
        normalized_format = datetime_format.replace('/', '-').replace('.', '-') if datetime_format else None
        # Use the provided format if available, otherwise fall back to auto-detection
        return pd.to_datetime(normalized_series, format=normalized_format, errors="coerce")
        # Normalize format string if provided (replace '/' and '.' with '-') to match normalized data
        normalized_format = datetime_format.replace('/', '-').replace('.', '-') if datetime_format else None
        # Use the provided format if available, otherwise fall back to auto-detection
        return pd.to_datetime(normalized_series, format=normalized_format, errors="coerce")
    if dtype_lower == "date":
        # Normalize separators: replace all '/', '.' with '-' to handle mixed separators
        normalized_series = series.astype(str).str.replace('/', '-', regex=False).str.replace('.', '-', regex=False)
        # Normalize format string if provided (replace '/' and '.' with '-')
        normalized_format = datetime_format.replace('/', '-').replace('.', '-') if datetime_format else None
        parsed = pd.to_datetime(normalized_series, format=normalized_format, errors="coerce")
        if datetime_format:
            # When a format was supplied (typically via auto-detect), keep full datetime precision
            return parsed
        return parsed.dt.date
    return series


def _apply_missing_strategy(
    df: pd.DataFrame, column: str, strategy: str, custom_value: Optional[Any]
) -> pd.DataFrame:
    series = df[column]
    strategy_lower = strategy.lower()
    if strategy_lower == "drop":
        return df[series.notna()]

    fill_value: Any | None = None
    if strategy_lower == "mean":
        fill_value = pd.to_numeric(series, errors="coerce").mean()
    elif strategy_lower == "median":
        fill_value = pd.to_numeric(series, errors="coerce").median()
    elif strategy_lower == "zero":
        fill_value = 0
    elif strategy_lower == "mode":
        mode_series = series.mode(dropna=True)
        fill_value = mode_series.iloc[0] if not mode_series.empty else ""
    elif strategy_lower == "empty":
        fill_value = ""
    elif strategy_lower == "custom":
        fill_value = custom_value
    elif strategy_lower == "ffill":
        df[column] = series.ffill()
        return df
    elif strategy_lower == "bfill":
        df[column] = series.bfill()
        return df
    else:
        fill_value = custom_value

    if fill_value is not None:
        df[column] = series.fillna(fill_value)
    return df


@router.post("/process_saved_dataframe", response_model=ProcessDataframeResponse)
async def process_saved_dataframe(payload: ProcessDataframeRequest):
    """Apply column-level processing (rename, dtype conversion, missing value handling) to a saved dataframe."""
    decoded = unquote(payload.object_name)
    prefix = await get_object_prefix()
    logger.info(f"🔧 [process_saved_dataframe] Received object_name: {decoded}, current prefix: {prefix}")
    
    if not decoded.startswith(prefix):
        logger.warning(f"⚠️ [process_saved_dataframe] Object name {decoded} does not start with prefix {prefix}")
        raise HTTPException(status_code=400, detail="Invalid object name")
    if not payload.instructions:
        raise HTTPException(status_code=400, detail="No processing instructions supplied")

    try:
        content = read_minio_object(decoded)
    except Exception as exc:
        raise HTTPException(status_code=404, detail=f"Dataframe not found: {exc}") from exc

    try:
        df = _load_dataframe_for_processing(decoded, content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to load dataframe: {exc}") from exc

    df_processed = df.copy()
    for instruction in payload.instructions:
        source_col = instruction.column
        if instruction.new_name and instruction.new_name in df_processed.columns and source_col != instruction.new_name:
            # Avoid clobbering existing column names by appending suffix
            logger.warning("process_saved_dataframe: target column %s already exists, skipping rename", instruction.new_name)
            continue
        if source_col not in df_processed.columns:
            logger.warning("process_saved_dataframe: column %s not found", source_col)
            continue
        if getattr(instruction, "drop_column", False):
            df_processed.drop(columns=[source_col], inplace=True)
            continue
        target_col = source_col
        if instruction.new_name and instruction.new_name != source_col:
            df_processed.rename(columns={source_col: instruction.new_name}, inplace=True)
            target_col = instruction.new_name
        if instruction.dtype:
            try:
                df_processed[target_col] = _cast_series_dtype(
                    df_processed[target_col],
                    instruction.dtype,
                    getattr(instruction, "datetime_format", None),
                )
            except Exception as exc:
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to cast column '{target_col}' to {instruction.dtype}: {exc}",
                ) from exc
        missing_strategy = getattr(instruction, "missing_strategy", None)
        custom_value = getattr(instruction, "custom_value", None)
        fill_value = (
            custom_value
            if custom_value is not None
            else getattr(instruction, "fill_value", None)
        )
        if missing_strategy:
            df_processed = _apply_missing_strategy(
                df_processed, target_col, missing_strategy, fill_value
            )
        elif fill_value is not None:
            df_processed[target_col] = df_processed[target_col].fillna(fill_value)

    arrow_buffer = io.BytesIO()
    pl.from_pandas(df_processed).write_ipc(arrow_buffer)
    arrow_bytes = arrow_buffer.getvalue()
    
    # CRITICAL: Write back to the exact same path (decoded) to preserve folder structure
    # For Excel sheets in folders, decoded should be: "client/app/project/folder_name/sheets/Sheet1.arrow"
    logger.info(f"💾 [process_saved_dataframe] Writing processed dataframe to: {decoded}")
    minio_client.put_object(
        MINIO_BUCKET,
        decoded,  # Use exact same path to preserve folder structure
        io.BytesIO(arrow_bytes),
        len(arrow_bytes),
        content_type="application/octet-stream",
    )
    logger.info(f"✅ [process_saved_dataframe] Successfully wrote to: {decoded}")

    flight_path = get_flight_path_for_csv(decoded)
    if flight_path:
        upload_dataframe(df_processed, flight_path)

    redis_client.setex(decoded, 3600, arrow_bytes)

    columns_meta = []
    for name in df_processed.columns:
        column_series = df_processed[name]
        columns_meta.append(
            {
                "name": name,
                "dtype": str(column_series.dtype),
                "missing_count": int(column_series.isna().sum()),
            }
        )

    return ProcessDataframeResponse(
        status="success",
        object_name=decoded,
        rows=len(df_processed),
        columns=columns_meta,
    )


@router.post("/file-metadata")
async def get_file_metadata(request: Request):
    """
    Get metadata for a file including column dtypes, missing values, and sample data.
    Expects JSON body with 'file_path' key.
    """
    try:
        body = await request.json()
        file_path = body.get("file_path")
        
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path is required")
        
        # Read file from MinIO
        data = read_minio_object(file_path)
        filename = Path(file_path).name
        
        # Parse based on file type using RobustFileReader
        if filename.lower().endswith(".csv") or filename.lower().endswith((".xls", ".xlsx")):
            # Use RobustFileReader which handles column preservation automatically
            df_result, _ = RobustFileReader.read_file_to_pandas(
                content=data,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
        elif filename.lower().endswith(".arrow"):
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")
        
        # Collect column metadata
        columns_info = []
        for col in df.columns:
            col_data = df[col]
            missing_count = int(col_data.isna().sum())
            total_rows = len(df)
            missing_percentage = (missing_count / total_rows * 100) if total_rows > 0 else 0
            
            # Get sample values (non-null)
            sample_values = col_data.dropna().head(5).tolist()
            
            columns_info.append({
                "name": str(col),
                "dtype": str(col_data.dtype),
                "missing_count": missing_count,
                "missing_percentage": round(missing_percentage, 2),
                "sample_values": sample_values,
            })
        
        return {
            "columns": columns_info,
            "total_rows": len(df),
            "total_columns": len(df.columns),
        }
        
    except Exception as e:
        logger.error(f"Error getting file metadata: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


def _find_stable_column_count(rows: List[List[str]], min_consistency: int = 3) -> tuple[int, int]:
    """
    Find where column count stabilizes (becomes consistent).
    Data rows typically have the same number of columns consistently.
    
    Example:
        - Description rows: 5 columns each
        - Data rows: 30 columns each (consistent)
        - Returns: (30, index_where_30_columns_start)
    
    Args:
        rows: List of rows, each row is a list of cell values
        min_consistency: Minimum number of consecutive rows with same column count to consider it stable
    
    Returns:
        Tuple of (stable_column_count, data_start_index)
        - stable_column_count: The consistent column count in data rows
        - data_start_index: Index where data rows start (after description rows)
    """
    if not rows:
        return 0, 0
    
    # Count columns in each row (use actual length of row list, including empty cells)
    # Also count non-empty cells for better detection
    column_counts = []
    non_empty_counts = []
    for row in rows:
        col_count = len(row)
        column_counts.append(col_count)
        # Count non-empty cells (cells that have actual content)
        non_empty = sum(1 for cell in row if cell and str(cell).strip())
        non_empty_counts.append(non_empty)
    
    if not column_counts:
        return 0, 0
    
    logger.info(f"Column counts (first 20): {column_counts[:20]}")
    logger.info(f"Non-empty counts (first 20): {non_empty_counts[:20]}")
    
    # Find the most common column count (excluding 0 and 1, which are likely empty/trivial rows)
    non_trivial_counts = [c for c in column_counts if c > 1]
    
    if not non_trivial_counts:
        # All rows have 0 or 1 columns, return first row as data start
        max_count = max(column_counts) if column_counts else 0
        logger.info(f"No non-trivial column counts found, using max_count={max_count}, data_start=0")
        return max_count, 0
    
    # Find the stable column count (most common count that appears consistently)
    count_freq = Counter(non_trivial_counts)
    stable_count = count_freq.most_common(1)[0][0] if count_freq else max(non_trivial_counts)
    
    logger.info(f"Stable column count detected: {stable_count} (appears {count_freq[stable_count]} times out of {len(non_trivial_counts)} rows)")
    
    # Find where this stable count starts appearing consistently
    # Look for first occurrence where we have min_consistency consecutive rows with stable_count
    data_start_idx = 0
    consecutive_count = 0
    
    for idx, col_count in enumerate(column_counts):
        if col_count == stable_count:
            consecutive_count += 1
            if consecutive_count >= min_consistency:
                # Found stable data block - data starts min_consistency rows back
                data_start_idx = idx - min_consistency + 1
                break
        else:
            consecutive_count = 0
    
    # If we didn't find a stable block with min_consistency, try to find where stable_count first appears
    if data_start_idx == 0:
        # Check if first few rows have the stable count
        if column_counts[0] == stable_count:
            # First row has stable count - might be header, data starts at row 0
            data_start_idx = 0
        else:
            # Find first occurrence of stable_count
            for idx, col_count in enumerate(column_counts):
                if col_count == stable_count:
                    data_start_idx = idx
                    break
    
    # Ensure data_start_idx is valid
    if data_start_idx < 0:
        data_start_idx = 0
    if data_start_idx >= len(rows):
        data_start_idx = 0
    
    logger.info(f"Column count analysis: stable_count={stable_count}, data_start_idx={data_start_idx}, "
                f"column_counts_sample={column_counts[:10]}")
    
    return stable_count, data_start_idx


def _detect_delimiter(content: bytes) -> str:
    """Detect CSV delimiter by sampling first few lines."""
    try:
        # Try to decode as UTF-8
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = content.decode('latin-1')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='ignore')
    
    # Sample first 5 lines
    lines = text.split('\n')[:5]
    if not lines:
        return ','
    
    # Count occurrences of common delimiters
    delimiter_counts = {',': 0, ';': 0, '\t': 0, '|': 0}
    for line in lines:
        if line.strip():
            delimiter_counts[','] += line.count(',')
            delimiter_counts[';'] += line.count(';')
            delimiter_counts['\t'] += line.count('\t')
            delimiter_counts['|'] += line.count('|')
    
    # Return delimiter with highest count
    return max(delimiter_counts, key=delimiter_counts.get) if max(delimiter_counts.values()) > 0 else ','


def _read_csv_rows_simple(content: bytes, pad_rows: bool = True) -> List[List[str]]:
    """
    Read CSV file row by row using Python's csv module.
    This ensures we get ALL rows including headers as data (not column names).
    Returns list of rows, each row is a list of cell values.
    
    Args:
        content: File content as bytes
        pad_rows: If True, pad all rows to same length. If False, keep original row lengths.
    """
    try:
        # Try to decode as UTF-8
        text = content.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = content.decode('latin-1')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='ignore')
    
    # Detect delimiter
    delimiter = _detect_delimiter(content)
    
    # Read rows using csv module (this reads ALL rows including headers as data)
    rows = []
    csv_reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    
    for row in csv_reader:
        # Convert tuple to list
        row_list = list(row)
        
        if pad_rows:
            # Pad rows to have same length as max columns seen so far
            if rows:
                max_cols = max(len(r) for r in rows)
                if len(row_list) < max_cols:
                    row_list.extend([''] * (max_cols - len(row_list)))
                elif len(row_list) > max_cols:
                    # Pad previous rows
                    for i, prev_row in enumerate(rows):
                        if len(prev_row) < len(row_list):
                            rows[i] = prev_row + [''] * (len(row_list) - len(prev_row))
        
        rows.append(row_list)
    
    return rows


def _read_excel_rows_simple(content: bytes, sheet_name: Optional[str] = None) -> List[List[str]]:
    """
    Read Excel file row by row using pandas with header=None to get ALL rows as data.
    This ensures we get actual column headers from the file, not pandas-generated names.
    Returns list of rows, each row is a list of cell values.
    
    IMPORTANT: We preserve original dtypes during reading, then convert to strings only for
    the row-by-row representation. This allows proper dtype detection later.
    """
    excel_file = io.BytesIO(content)
    
    # Use pandas with header=None to read ALL rows as data (including headers)
    # This ensures we get the actual column names from the Excel file, not "col_0", "col_1", etc.
    # CRITICAL: Don't use dtype=str here - let pandas infer types so we can preserve numeric/datetime types
    try:
        df = pd.read_excel(
            excel_file,
            sheet_name=sheet_name,
            header=None,  # Read all rows as data, don't treat first row as column names
            engine='openpyxl',
            # REMOVED dtype=str - let pandas infer types naturally
            # This preserves numeric types (int64, float64) and datetime types
            na_filter=True,  # Convert empty cells to NaN (pandas default)
        )
        
        # Convert DataFrame to list of lists
        # Convert values to strings for row representation, but preserve type information
        rows = []
        for _, row in df.iterrows():
            row_values = []
            for val in row:
                if pd.isna(val):
                    row_values.append('')
                else:
                    # Convert to string but preserve the original value's type information
                    # This allows downstream code to detect numeric/datetime types
                    row_values.append(str(val))
            rows.append(row_values)
        
        return rows
    except Exception as e:
        logger.error(f"Error reading Excel file with pandas: {e}")
        # Fallback to openpyxl if pandas fails
        excel_file.seek(0)  # Reset file pointer
        workbook = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
        
        if sheet_name:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
                logger.warning(f"Sheet '{sheet_name}' not found, using active sheet")
        else:
            sheet = workbook.active
        
        rows = []
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else '' for cell in row]
            rows.append(row_values)
        
        workbook.close()
        return rows


def _separate_rows_simple(all_rows: List[List[str]]) -> tuple[List[List[str]], List[List[str]], int]:
    """
    Separate description rows from data rows using column count stability.
    - Data rows have consistent column count (e.g., all have 30 columns)
    - Description rows have different column count (e.g., 5 columns)
    - Find where column count stabilizes - that's where data starts
    
    Returns: (description_rows, data_rows, data_start_index)
    """
    if not all_rows:
        return [], [], 0
    
    # Find stable column count and where data starts
    stable_column_count, data_start_idx = _find_stable_column_count(all_rows, min_consistency=3)
    
    # Ensure data_start_idx is valid
    if data_start_idx < 0:
        data_start_idx = 0
    if data_start_idx >= len(all_rows):
        data_start_idx = 0
    
    # Description rows are everything before data_start_idx
    description_rows = all_rows[:data_start_idx]
    data_rows = all_rows[data_start_idx:]
    
    logger.info(f"Separated rows: {len(description_rows)} description rows, {len(data_rows)} data rows, "
                f"stable column count: {stable_column_count}, data starts at index: {data_start_idx}")
    
    return description_rows, data_rows, data_start_idx


def _detect_header_row_simple(data_rows: List[List[str]], max_check: int = 20) -> tuple[int, str]:
    """
    Detect header row by comparing rows with each other.
    The row that is most different/unmatched from other rows is likely the header.
    
    Logic:
    1. Compare each row with all other rows
    2. Calculate similarity/difference score for each row
    3. The row with lowest similarity (most different) is the header
    4. Header rows typically have:
       - Text values (not numbers)
       - Different patterns than data rows
       - More consistent structure across columns
    """
    if not data_rows or len(data_rows) < 2:
        return 0, 'low'
    
    # Check first few rows
    check_rows = data_rows[:min(max_check, len(data_rows))]
    
    def _is_numeric(cell_value: str) -> bool:
        """Check if a cell value is numeric."""
        if not cell_value or not str(cell_value).strip():
            return False
        try:
            float(str(cell_value).replace(',', '').replace('$', '').replace('%', ''))
            return True
        except ValueError:
            return False
    
    def _row_similarity(row1: List[str], row2: List[str]) -> float:
        """Calculate similarity between two rows (0-1, higher = more similar)."""
        if not row1 or not row2:
            return 0.0
        
        min_len = min(len(row1), len(row2))
        if min_len == 0:
            return 0.0
        
        matches = 0
        total = 0
        
        for i in range(min_len):
            val1 = str(row1[i]).strip() if row1[i] else ''
            val2 = str(row2[i]).strip() if row2[i] else ''
            
            # Skip empty cells
            if not val1 and not val2:
                continue
            
            total += 1
            
            # Check if both are numeric (data rows often have numeric patterns)
            if _is_numeric(val1) and _is_numeric(val2):
                matches += 1
            # Check if both are non-numeric (could be text data)
            elif not _is_numeric(val1) and not _is_numeric(val2):
                # Check if they have similar length/pattern
                if abs(len(val1) - len(val2)) < 5:
                    matches += 0.5
        
        return matches / total if total > 0 else 0.0
    
    # Calculate average similarity of each row with all other rows
    row_similarities = []
    for idx, row in enumerate(check_rows):
        if not row:
            row_similarities.append(1.0)  # Empty row = high similarity (not header)
            continue
        
        similarities = []
        # Compare with all other rows
        for other_idx, other_row in enumerate(check_rows):
            if idx != other_idx and other_row:
                sim = _row_similarity(row, other_row)
                similarities.append(sim)
        
        # Average similarity (lower = more different = more likely to be header)
        avg_similarity = sum(similarities) / len(similarities) if similarities else 1.0
        row_similarities.append(avg_similarity)
    
    # Find row with lowest similarity (most different from others)
    min_similarity = min(row_similarities)
    best_idx = row_similarities.index(min_similarity)
    
    # Additional checks to boost confidence
    best_row = check_rows[best_idx]
    
    # Check if best row looks like headers (mostly text, not numbers)
    text_count = 0
    non_empty_count = 0
    for cell in best_row[:min(20, len(best_row))]:  # Check first 20 columns
        if cell and str(cell).strip():
            non_empty_count += 1
            if not _is_numeric(str(cell)):
                text_count += 1
    
    text_ratio = text_count / non_empty_count if non_empty_count > 0 else 0
    
    # Calculate confidence based on:
    # 1. How different it is from other rows (low similarity)
    # 2. Whether it looks like headers (high text ratio)
    # 3. Position (earlier rows more likely)
    difference_score = 1.0 - min_similarity  # Higher = more different
    header_likelihood = difference_score * 0.6 + text_ratio * 0.4
    
    # Adjust for position (earlier rows get slight boost)
    if best_idx == 0:
        header_likelihood += 0.1
    elif best_idx < 3:
        header_likelihood += 0.05
    
    # Determine confidence
    if header_likelihood >= 0.7 and best_idx < 3:
        confidence = 'high'
    elif header_likelihood >= 0.5 and best_idx < 5:
        confidence = 'medium'
    else:
        confidence = 'low'
    
    logger.info(f"Header detection: row {best_idx} selected (similarity={min_similarity:.3f}, "
                f"text_ratio={text_ratio:.3f}, likelihood={header_likelihood:.3f}, confidence={confidence})")
    
    return best_idx, confidence


@router.get("/file-preview")
async def get_file_preview(
    object_name: str = Query(..., description="MinIO object name/path"),
    client_id: str = Query(""),
    app_id: str = Query(""),
    project_id: str = Query(""),
    sheet_name: Optional[str] = Query(None, description="Sheet name for Excel files"),
):
    """
    Get preview of file with raw rows (no headers applied).
    Uses simple row-by-row reading with Python's csv module.
    Separates description rows from data rows using column-count matching algorithm:
    - Finds where column count stabilizes (becomes consistent)
    - Rows before stable count = description rows
    - Rows with stable count = data rows
    Returns separated description_rows and data_rows for user to select header.
    """
    # Validate object_name is not empty
    if not object_name or not object_name.strip():
        raise HTTPException(
            status_code=400,
            detail="object_name is required and cannot be empty. Please ensure the file was uploaded successfully."
        )
    
    try:
        # CRITICAL: If we receive an Arrow file path, try to find the original CSV/Excel file
        # Arrow files are processed files with "col_0", "col_1" column names - we need the original!
        original_object_name = object_name
        if object_name.lower().endswith(".arrow"):
            logger.warning(f"Received Arrow file path for preview: {object_name} - attempting to find original file")
            path_parts = object_name.split("/")
            
            # Strategy 1: Check originals/ folder (for files uploaded via /upload-file)
            # Path format: {prefix}tmp/{filename}.arrow -> {prefix}tmp/originals/{original_filename}
            if "tmp" in path_parts:
                tmp_idx = path_parts.index("tmp")
                original_filename_base = path_parts[-1].replace(".arrow", "")
                
                # Try common extensions
                for ext in [".csv", ".xlsx", ".xls"]:
                    test_path = "/".join(path_parts[:tmp_idx+1]) + "/originals/" + original_filename_base + ext
                    try:
                        test_data = read_minio_object(test_path)
                        original_object_name = test_path
                        logger.info(f"✓ Found original file in originals/: {object_name} -> {original_object_name}")
                        break
                    except Exception as e:
                        logger.debug(f"  Tried {test_path}: {e}")
                        continue
            
            # Strategy 2: Check uploads/ folder (for Excel multi-sheet uploads)
            # Excel multi-sheet saves original as: {prefix}uploads/{session_id}/original.xlsx
            if original_object_name == object_name and "tmp" in path_parts:
                # Extract session ID from path (might be in different positions)
                # Try to find uploads folder at same level as tmp
                prefix_parts = path_parts[:tmp_idx] if tmp_idx > 0 else []
                # Session ID might be the filename without extension
                session_id = path_parts[-1].replace(".arrow", "")
                test_path = "/".join(prefix_parts) + "/uploads/" + session_id + "/original.xlsx"
                try:
                    test_data = read_minio_object(test_path)
                    original_object_name = test_path
                    logger.info(f"✓ Found original Excel file in uploads/: {object_name} -> {original_object_name}")
                except Exception as e:
                    logger.debug(f"  Tried {test_path}: {e}")
                    pass
            
            # If still no original found, log warning but proceed with Arrow file
            if original_object_name == object_name:
                logger.warning(f"⚠ Could not find original file for {object_name}, will read Arrow file (may show col_0, col_1)")
        
        # Read file from MinIO (use original file if found, otherwise use provided path)
        data = read_minio_object(original_object_name)
        filename = Path(original_object_name).name
        
        logger.info(f"Reading file for preview: original_object_name={original_object_name}, filename: {filename}, size: {len(data)} bytes")
        
        # Read rows based on file type (WITHOUT padding to preserve original column counts)
        # IMPORTANT: Always read original CSV/Excel files, not processed Arrow files
        # Arrow files have column names like "col_0", "col_1" which we don't want
        if filename.lower().endswith(".arrow"):
            # Handle Arrow files - convert to rows
            # Arrow files are processed files, so extract column names as first row
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
            # Extract column names as first row
            column_names = df.columns.tolist()
            logger.info(f"Arrow file column names (first 10): {column_names[:10]}")
            # Reset columns to numeric indices
            df.columns = range(len(df.columns))
            # Prepend column names as first row
            first_row_df = pd.DataFrame([column_names], columns=df.columns)
            df = pd.concat([first_row_df, df], ignore_index=True)
            # Convert DataFrame to list of lists (preserve original lengths)
            all_rows = [list(row) for row in df.values]
        elif filename.lower().endswith(".csv"):
            # Use Python's csv module - reads ALL rows including headers as data
            # This ensures we get actual column headers from CSV, not "col_0", "col_1"
            logger.info("Reading CSV file with csv.reader() - will read all rows including headers")
            all_rows = _read_csv_rows_simple(data, pad_rows=False)
            if all_rows:
                first_row_preview = all_rows[0][:10]
                logger.info(f"CSV first row (headers): {first_row_preview}")
                # Check if first row looks like pandas-generated column names (col_0, col_1, etc.)
                if len(first_row_preview) > 0 and all(
                    str(cell).strip().lower().startswith('col_') and 
                    str(cell).strip().lower()[4:].isdigit() 
                    for cell in first_row_preview if cell
                ):
                    logger.warning("CSV first row looks like pandas-generated column names (col_0, col_1, etc.) - file may have been processed incorrectly")
        elif filename.lower().endswith((".xls", ".xlsx")):
            # Use pandas with header=None - reads ALL rows including headers as data
            # This ensures we get actual column headers from Excel, not "col_0", "col_1"
            logger.info(f"Reading Excel file with pandas (header=None) - will read all rows including headers")
            all_rows = _read_excel_rows_simple(data, sheet_name)
            if all_rows:
                first_row_preview = all_rows[0][:10]
                logger.info(f"Excel first row (headers): {first_row_preview}")
                # Check if first row looks like pandas-generated column names (col_0, col_1, etc.)
                if len(first_row_preview) > 0 and all(
                    str(cell).strip().lower().startswith('col_') and 
                    str(cell).strip().lower()[4:].isdigit() 
                    for cell in first_row_preview if cell
                ):
                    logger.warning("Excel first row looks like pandas-generated column names (col_0, col_1, etc.) - file may have been processed incorrectly")
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")
        
        if not all_rows:
            raise HTTPException(status_code=400, detail="File appears to be empty")
        
        # Separate description rows from data rows using simple column-count logic
        # This works because rows have their original column counts (not padded)
        description_rows, data_rows, data_start_idx = _separate_rows_simple(all_rows)
        
        # Log separation results for debugging
        logger.info(f"File preview separation results:")
        logger.info(f"  Total rows read: {len(all_rows)}")
        logger.info(f"  Description rows: {len(description_rows)} (rows 0-{data_start_idx-1})")
        logger.info(f"  Data rows: {len(data_rows)} (starting from row {data_start_idx})")
        if all_rows:
            logger.info(f"  Column counts sample (first 10 rows): {[len(row) for row in all_rows[:10]]}")
        
        # NOW pad rows for consistent structure (after separation)
        if data_rows:
            max_cols = max(len(row) for row in data_rows) if data_rows else 0
            if max_cols > 0:
                # Pad data rows to have same number of columns
                padded_data_rows = []
                for row in data_rows:
                    padded_row = row + [''] * (max_cols - len(row))
                    padded_data_rows.append(padded_row)
                data_rows = padded_data_rows
        
        if description_rows:
            max_desc_cols = max(len(row) for row in description_rows) if description_rows else 0
            if max_desc_cols > 0:
                # Pad description rows to have same number of columns
                padded_desc_rows = []
                for row in description_rows:
                    padded_row = row + [''] * (max_desc_cols - len(row))
                    padded_desc_rows.append(padded_row)
                description_rows = padded_desc_rows
        
        # Get preview rows (first 15 rows of data)
        preview_rows = data_rows[:15]
        preview_row_count = len(preview_rows)
        
        # Detect suggested header row in the data rows (relative to data rows, 0-indexed)
        suggested_header_row_relative, suggested_header_confidence = _detect_header_row_simple(data_rows)
        
        # Calculate absolute row index (1-indexed for display)
        # data_start_idx is 0-indexed, suggested_header_row_relative is 0-indexed relative to data rows
        # To get 1-indexed absolute: data_start_idx (0-indexed) + suggested_header_row_relative (0-indexed) + 1
        # This matches how data_rows_structured calculates row_index: data_start_idx + idx + 1
        suggested_header_row_absolute = data_start_idx + suggested_header_row_relative + 1
        
        logger.info(f"Header detection results: relative={suggested_header_row_relative}, "
                   f"data_start_idx={data_start_idx}, absolute={suggested_header_row_absolute}, "
                   f"confidence={suggested_header_confidence}")
        
        # Convert description rows to structured format
        # Use actual absolute row indices (1-indexed for display)
        description_rows_structured = []
        for idx, row in enumerate(description_rows):
            absolute_row_index = idx + 1  # 1-indexed (first description row is row 1)
            description_rows_structured.append({
                "row_index": absolute_row_index,  # 1-indexed for display
                "cells": [str(cell) if cell else "" for cell in row]
            })
        
        # Convert data rows to structured format with row_index and relative_index
        data_rows_structured = []
        for idx, row in enumerate(preview_rows):
            data_rows_structured.append({
                "row_index": data_start_idx + idx + 1,  # 1-indexed absolute row number
                "relative_index": idx,  # 0-indexed relative to data rows
                "cells": [str(cell) if cell else "" for cell in row]
            })
        
        # Get column count (use max columns across all rows)
        column_count = max(len(row) for row in all_rows) if all_rows else 0
        
        response_data = {
            "data_rows": data_rows_structured,
            "description_rows": description_rows_structured,
            "data_rows_count": len(data_rows),
            "description_rows_count": len(description_rows),
            "data_rows_start": data_start_idx,
            "preview_row_count": preview_row_count,
            "column_count": column_count,
            "total_rows": len(data_rows),
            "suggested_header_row": suggested_header_row_relative,  # Relative to data rows (0-indexed)
            "suggested_header_row_absolute": suggested_header_row_absolute,  # Absolute including description rows
            "suggested_header_confidence": suggested_header_confidence,
            "has_description_rows": len(description_rows) > 0,
        }
        
        # Log response for debugging
        logger.info(f"File preview response: {len(description_rows_structured)} description rows, "
                   f"{len(data_rows_structured)} data rows in preview, "
                   f"data_rows_start={data_start_idx}, "
                   f"has_description_rows={len(description_rows) > 0}")
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting file preview: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/row-issues")
async def get_row_issues(
    object_name: str = Query(..., description="MinIO object name/path"),
    client_id: str = Query(""),
    app_id: str = Query(""),
    project_id: str = Query(""),
    sheet_name: Optional[str] = Query(None, description="Sheet name for Excel files"),
    limit: int = Query(200, ge=1, le=2000, description="Max issues to return"),
    offset: int = Query(0, ge=0, description="Offset into the issues list"),
):
    """
    Scan the full file for structural row issues (delimiter spillover, extra/missing columns, sparse rows).
    Returns total issue counts and a paginated slice of problematic rows to avoid loading everything in the UI.
    """
    try:
        # Resolve original file if an Arrow path is provided
        original_object_name = object_name
        if object_name.lower().endswith(".arrow"):
            path_parts = object_name.split("/")
            if "tmp" in path_parts:
                tmp_idx = path_parts.index("tmp")
                original_filename_base = path_parts[-1].replace(".arrow", "")
                for ext in [".csv", ".xlsx", ".xls"]:
                    test_path = "/".join(path_parts[:tmp_idx+1]) + "/originals/" + original_filename_base + ext
                    try:
                        read_minio_object(test_path)
                        original_object_name = test_path
                        break
                    except Exception:
                        continue
            if original_object_name == object_name and "tmp" in path_parts:
                tmp_idx = path_parts.index("tmp")
                prefix_parts = path_parts[:tmp_idx] if tmp_idx > 0 else []
                session_id = path_parts[-1].replace(".arrow", "")
                test_path = "/".join(prefix_parts) + "/uploads/" + session_id + "/original.xlsx"
                try:
                    read_minio_object(test_path)
                    original_object_name = test_path
                except Exception:
                    pass

        data = read_minio_object(original_object_name)
        filename = Path(original_object_name).name

        # Read all rows (un-padded) to keep true column counts
        if filename.lower().endswith(".arrow"):
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
            df.columns = range(len(df.columns))
            all_rows = [list(row) for row in df.values]
        elif filename.lower().endswith(".csv"):
            all_rows = _read_csv_rows_simple(data, pad_rows=False)
        elif filename.lower().endswith((".xls", ".xlsx")):
            all_rows = _read_excel_rows_simple(data, sheet_name)
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")

        if not all_rows:
            raise HTTPException(status_code=400, detail="File appears to be empty")

        # Separate description/data rows (keeps raw lengths)
        description_rows, data_rows, data_start_idx = _separate_rows_simple(all_rows)
        if not data_rows:
            raise HTTPException(status_code=400, detail="No data rows found in file")

        expected_columns = max(len(row) for row in data_rows) if data_rows else 0

        issues: List[Dict[str, Any]] = []
        for idx, row in enumerate(data_rows):
            row_len = len(row)
            non_empty = sum(1 for cell in row if cell and str(cell).strip())
            row_issues: List[str] = []

            if row_len > expected_columns:
                row_issues.append(f"Has {row_len - expected_columns} extra column(s) - possible delimiter spillover")
            elif row_len < expected_columns and non_empty > 0:
                row_issues.append(f"Missing {expected_columns - row_len} column(s)")

            # Unescaped delimiters / unconventional content
            has_comma = any(cell and "," in str(cell) and not str(cell).startswith('"') for cell in row)
            has_semicolon = any(cell and str(cell).count(";") >= 2 for cell in row)
            has_tab = any(cell and "\t" in str(cell) for cell in row)
            if has_comma or has_semicolon or has_tab:
                row_issues.append("Contains potential unescaped delimiters")

            # Sparse row
            if expected_columns > 0 and non_empty > 0 and non_empty < expected_columns * 0.2:
                row_issues.append("Mostly empty - may be metadata or formatting")

            if row_issues:
                issues.append({
                    "row_index": data_start_idx + idx + 1,  # absolute, 1-indexed
                    "column_count": row_len,
                    "non_empty_cells": non_empty,
                    "issues": row_issues,
                    "severity": "warning",
                })

        total_issues = len(issues)
        paged_issues = issues[offset:offset + limit] if issues else []
        has_more = offset + limit < total_issues

        return {
            "total_data_rows": len(data_rows),
            "expected_columns": expected_columns,
            "issues_count": total_issues,
            "issues": paged_issues,
            "limit": limit,
            "offset": offset,
            "has_more": has_more,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error scanning row issues: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/apply-header-selection")
async def apply_header_selection(
    object_name: str = Form(..., description="MinIO object name/path"),
    header_row: int = Form(..., description="Row index to use as header (0-based)"),
    header_row_count: int = Form(1, description="Number of header rows to merge (default: 1)"),
    sheet_name: Optional[str] = Form(None, description="Sheet name for Excel files"),
    client_id: str = Form(""),
    app_id: str = Form(""),
    project_id: str = Form(""),
):
    """
    Apply header row selection to file and save processed version.
    Reads file in raw mode, applies selected header row, and saves as Arrow file.
    """
    try:
        # Read file from MinIO
        data = read_minio_object(object_name)
        filename = Path(object_name).name
        
        # Read rows based on file type (WITHOUT padding to preserve original column counts)
        if filename.lower().endswith(".arrow"):
            # Handle Arrow files - convert to rows
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
            # Extract column names as first row
            column_names = df.columns.tolist()
            # Reset columns to numeric indices
            df.columns = range(len(df.columns))
            # Prepend column names as first row
            first_row_df = pd.DataFrame([column_names], columns=df.columns)
            df = pd.concat([first_row_df, df], ignore_index=True)
            # Convert DataFrame to list of lists (preserve original lengths)
            all_rows = [list(row) for row in df.values]
        elif filename.lower().endswith(".csv"):
            # Use Python's csv module - DON'T pad rows so we can detect column count differences
            all_rows = _read_csv_rows_simple(data, pad_rows=False)
        elif filename.lower().endswith((".xls", ".xlsx")):
            # Use openpyxl for Excel files (already preserves original lengths)
            all_rows = _read_excel_rows_simple(data, sheet_name)
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")
        
        if not all_rows:
            raise HTTPException(status_code=400, detail="File appears to be empty")
        
        # Separate description rows from data rows using simple column-count logic
        # This works because rows have their original column counts (not padded)
        description_rows, data_rows, data_start_idx = _separate_rows_simple(all_rows)
        
        # Convert data rows to DataFrame for processing
        # Find max columns to ensure consistent DataFrame structure
        max_cols = max(len(row) for row in data_rows) if data_rows else 0
        if max_cols == 0:
            raise HTTPException(status_code=400, detail="No data rows found in file")
        
        # NOW pad rows to have same number of columns (after separation)
        padded_data_rows = []
        for row in data_rows:
            padded_row = row + [''] * (max_cols - len(row))
            padded_data_rows.append(padded_row[:max_cols])  # Ensure exact length
        
        # CRITICAL: Extract headers from RAW STRING ROWS before type inference
        # This preserves numeric/date headers exactly as they appear in the file
        # Validate header row index (header_row is relative to padded_data_rows after description separation)
        if header_row < 0 or header_row >= len(padded_data_rows):
            raise HTTPException(status_code=400, detail=f"Header row index {header_row} is out of range (0-{len(padded_data_rows)-1})")
        
        # Validate header_row_count
        if header_row_count < 1:
            header_row_count = 1
        if header_row_count > 10:
            header_row_count = 10  # Limit to 10 rows max
        
        # Check if we have enough rows for multi-row header
        if header_row + header_row_count > len(padded_data_rows):
            logger.warning(f"Requested {header_row_count} header rows starting at {header_row}, but only {len(padded_data_rows) - header_row} rows available. Using available rows.")
            header_row_count = len(padded_data_rows) - header_row
        
        # Extract headers from RAW STRING ROWS (before type inference)
        # This ensures numeric/date headers are preserved as strings exactly as they appear
        if header_row_count > 1:
            # Get header rows from raw string data
            header_rows = []
            for i in range(header_row_count):
                row_idx = header_row + i
                if row_idx < len(padded_data_rows):
                    # Get raw string values directly from padded rows
                    header_row_data = padded_data_rows[row_idx]
                    header_rows.append(header_row_data)
            
            # Merge header rows: combine values from each row
            # Example: ["Sales", "", "2024"] + ["", "Q1", ""] → ["Sales", "Q1", "2024"]
            # Or: ["Sales", "", ""] + ["", "", "2024"] → ["Sales_2024", "", ""]
            merged_headers = []
            max_cols = max(len(row) for row in header_rows) if header_rows else 0
            
            for col_idx in range(max_cols):
                column_values = []
                for header_row_data in header_rows:
                    if col_idx < len(header_row_data):
                        val = header_row_data[col_idx]
                        # Convert to string and strip - preserve exact representation
                        val_str = str(val).strip() if val else ""
                        # Only add non-empty values
                        if val_str:
                            column_values.append(val_str)
                
                # Combine column values
                if column_values:
                    # Join with underscore, e.g., "Sales" + "2024" → "Sales_2024"
                    merged_header = "_".join(column_values)
                    merged_headers.append(merged_header)
                else:
                    merged_headers.append("")  # Empty column
            
            headers = merged_headers
            logger.info(f"Merged {header_row_count} header rows into: {headers[:5]}...")
        else:
            # Single header row - extract from raw string data
            header_row_raw = padded_data_rows[header_row]
            headers = []
            for val in header_row_raw:
                # Convert to string and preserve exact representation (including numeric like "2021.0")
                val_str = str(val).strip() if val else ""
                headers.append(val_str)
        
        # Now create DataFrame from data rows (skip header rows)
        # CRITICAL: Let pandas infer types naturally instead of forcing all to string
        # This preserves numeric and datetime types from the original file
        data_rows_only = padded_data_rows[header_row + header_row_count:]
        df_data = pd.DataFrame(data_rows_only)
        
        # Try to infer types for each column (preserve numeric/datetime types)
        # This is important because _read_excel_rows_simple and _read_csv_rows_simple
        # return string values, but we want pandas to detect the actual types
        for col_idx in range(len(df_data.columns)):
            col_data = df_data.iloc[:, col_idx]
            
            # Skip if column is all empty
            if col_data.astype(str).str.strip().eq('').all():
                continue
            
            # Try to convert to numeric first (handles int and float)
            try:
                numeric_col = pd.to_numeric(col_data, errors='coerce')
                # If most values converted successfully, use numeric type
                non_null_count = numeric_col.notna().sum()
                total_count = len(col_data)
                if total_count > 0 and non_null_count >= total_count * 0.8:  # At least 80% numeric
                    df_data.iloc[:, col_idx] = numeric_col
                    logger.debug(f"Column {col_idx}: inferred numeric type (converted {non_null_count}/{total_count} values)")
                    continue
            except Exception as e:
                logger.debug(f"Column {col_idx}: numeric conversion failed: {e}")
                pass
            
            # Try to convert to datetime
            try:
                datetime_col = pd.to_datetime(col_data, errors='coerce')
                # If most values converted successfully, use datetime type
                non_null_count = datetime_col.notna().sum()
                total_count = len(col_data)
                if total_count > 0 and non_null_count >= total_count * 0.8:  # At least 80% datetime
                    df_data.iloc[:, col_idx] = datetime_col
                    logger.debug(f"Column {col_idx}: inferred datetime type (converted {non_null_count}/{total_count} values)")
                    continue
            except Exception as e:
                logger.debug(f"Column {col_idx}: datetime conversion failed: {e}")
                pass
            
            # Otherwise keep as object/string (default)
            # This preserves text columns
        
        # df_data already contains only data rows (header rows were excluded above)
        df_processed = df_data.copy()
        
        # Ensure we have enough column names (pad if needed)
        if len(headers) < len(df_processed.columns):
            headers.extend([f"Column_{i}" for i in range(len(headers), len(df_processed.columns))])
        elif len(headers) > len(df_processed.columns):
            headers = headers[:len(df_processed.columns)]
        
        # Apply headers directly - preserve exact names as selected by user
        df_processed.columns = headers[:len(df_processed.columns)]
        
        # Only normalize empty column names (don't standardize/modify non-empty names)
        # This preserves numeric headers and exact user-selected names
        from app.features.data_upload_validate.file_ingestion.processors.cleaning import DataCleaner
        # Only normalize truly empty column names, don't modify existing ones
        columns_list = df_processed.columns.tolist()
        new_columns = []
        unnamed_counter = 0
        for col in columns_list:
            col_str = str(col).strip()
            if not col_str or col_str == "":
                new_columns.append(f"Column_{unnamed_counter}")
                unnamed_counter += 1
            else:
                # Preserve exact column name as-is
                new_columns.append(col_str)
        
        # Handle duplicate column names by appending suffix
        seen = {}
        final_columns = []
        for col in new_columns:
            if col in seen:
                seen[col] += 1
                final_columns.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                final_columns.append(col)
        
        df_processed.columns = final_columns
        
        # Remove empty rows/columns
        df_processed = DataCleaner.remove_empty_rows(df_processed)
        df_processed = DataCleaner.remove_empty_columns(df_processed)
        df_processed = df_processed.reset_index(drop=True)
        
        # CRITICAL FIX: Replace empty strings with NaN before converting to Polars
        # Polars will fail if it tries to convert empty strings to numeric types
        # Replace empty strings and whitespace-only strings with NaN for all object columns
        for col in df_processed.columns:
            if df_processed[col].dtype == 'object':
                # Replace empty strings, whitespace-only strings, and null-like strings with NaN
                df_processed[col] = df_processed[col].replace(['', ' ', '  ', 'None', 'null', 'NULL', 'nan', 'NaN', 'N/A', 'n/a'], pd.NA)
        
        # Convert to Polars and save as Arrow
        # Use nan_to_null=True to convert pandas NaN to Polars null
        # This prevents Polars from trying to convert empty strings to numeric types
        try:
            df_pl = pl.from_pandas(df_processed, nan_to_null=True)
        except Exception as e:
            # If conversion fails due to type inference issues, convert all to string first
            logger.warning(f"Direct Polars conversion failed: {e}. Converting all columns to string first, then inferring types.")
            # Convert all columns to string first (this prevents type inference errors)
            df_processed_str = df_processed.astype(str)
            # Replace string representations of NaN/None/empty with None
            df_processed_str = df_processed_str.replace(['nan', 'None', 'null', 'NULL', 'NaN', '<NA>', 'NaT', ''], None)
            # Convert to Polars with all columns as string
            df_pl = pl.from_pandas(df_processed_str, nan_to_null=True)
            # Now let Polars infer proper types - it will handle None/null values gracefully
            # Polars can infer types from string data when null values are properly set
        
        # Save to MinIO
        arrow_buf = io.BytesIO()
        df_pl.write_ipc(arrow_buf)
        arrow_name = Path(filename).stem + ".arrow"
        
        # Determine upload path (use same prefix as original if possible)
        tmp_prefix = "temp_uploads/"
        if "/" in object_name:
            parts = object_name.split("/")
            if len(parts) > 1:
                tmp_prefix = "/".join(parts[:-1]) + "/"
        
        result = upload_to_minio(arrow_buf.getvalue(), arrow_name, tmp_prefix)
        
        if result.get("status") != "success":
            raise HTTPException(status_code=500, detail=result.get("error_message", "Failed to save processed file"))
        
        return {
            "status": "success",
            "file_path": result["object_name"],
            "file_name": arrow_name,
            "total_rows": len(df_processed),
            "total_columns": len(df_processed.columns),
            "header_row_applied": header_row,
            "header_row_count": header_row_count,
            "description_rows_count": len(description_rows),
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error applying header selection: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/file-columns")
async def get_file_columns(
    object_name: str = Query(..., description="MinIO object name/path"),
    client_id: str = Query(""),
    app_id: str = Query(""),
    project_id: str = Query(""),
):
    """
    Get column names and sample values from a file.
    Returns column names, sample values, and AI/historical suggestions.
    """
    try:
        # Read file from MinIO
        data = read_minio_object(object_name)
        filename = Path(object_name).name
        
        # Parse based on file type using RobustFileReader
        if filename.lower().endswith(".csv") or filename.lower().endswith((".xls", ".xlsx")):
            # Use RobustFileReader which handles column preservation automatically
            df_result, _ = RobustFileReader.read_file_to_pandas(
                content=data,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
        elif filename.lower().endswith(".arrow"):
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")
        
        # Get column names and sample values
        columns = list(df.columns)
        sample_values = []
        rule_based_suggestions = []
        historical_matches = []
        
        for col in columns:
            # Get sample values (non-null, first 5)
            samples = df[col].dropna().head(5).tolist()
            sample_values.append([str(s) for s in samples])
            
            # Rule-based column name cleaning
            cleaned = str(col).strip()
            # Remove special characters (keep alphanumeric and underscores)
            cleaned = re.sub(r'[^a-zA-Z0-9_]', '_', cleaned)
            # Remove leading/trailing underscores
            cleaned = cleaned.strip('_')
            # Convert to snake_case if contains spaces or camelCase
            if ' ' in str(col) or any(c.isupper() for c in str(col) if c.isalpha()):
                cleaned = re.sub(r'([a-z])([A-Z])', r'\1_\2', cleaned).replace(' ', '_').lower()
            # Ensure it starts with a letter
            if cleaned and cleaned[0].isdigit():
                cleaned = 'col_' + cleaned
            # If empty after cleaning, use default
            if not cleaned:
                cleaned = 'unnamed_column'
            
            rule_based_suggestions.append(cleaned if cleaned != str(col) else None)
            
            # TODO: Query historical matches from memory service
            historical_matches.append(None)
        
        return {
            "columns": columns,
            "sample_values": sample_values,
            "rule_based_suggestions": rule_based_suggestions,
            "historical_matches": historical_matches,
        }
        
    except Exception as e:
        logger.error(f"Error getting file columns: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/detect-datetime-format")
async def detect_datetime_format(request: Request):
    """
    Auto-detect datetime format for a column.
    Expects JSON body with:
    - file_path: str
    - column_name: str
    """
    try:
        body = await request.json()
        file_path = body.get("file_path")
        column_name = body.get("column_name")
        
        if not file_path or not column_name:
            raise HTTPException(status_code=400, detail="file_path and column_name are required")
        
        # Read file from MinIO
        data = read_minio_object(file_path)
        filename = Path(file_path).name
        
        # Parse based on file type using RobustFileReader
        if filename.lower().endswith(".csv") or filename.lower().endswith((".xls", ".xlsx")):
            # Use RobustFileReader which handles column preservation automatically
            df_result, _ = RobustFileReader.read_file_to_pandas(
                content=data,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
        elif filename.lower().endswith(".arrow"):
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")
        
        if column_name not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{column_name}' not found")
        
        # Get non-null values from entire column
        column_data = df[column_name].dropna()
        if len(column_data) == 0:
            return {
                "detected_format": None,
                "can_detect": False,
                "sample_values": []
            }
        
        # Use entire column for detection (not just sample values)
        all_values = column_data.astype(str).tolist()
        sample_values = all_values[:5]  # Keep for response display only
        
        # Detect which separator the original data uses (check entire column)
        # Check if data primarily uses '/', '-', or '.' as separator
        has_slash = any('/' in str(val) for val in all_values)
        has_dash = any('-' in str(val) and not str(val).startswith('-') for val in all_values)
        has_dot = any('.' in str(val) and not str(val).endswith('.') for val in all_values)
        
        # Normalize separators for detection (handle mixed / and -)
        # Convert all / to - for standardization during testing
        # Also normalize dots to dashes for consistency
        normalized_all = [str(val).replace('/', '-').replace('.', '-') for val in all_values]
        
        # Try common datetime formats (test with normalized data, but return format matching original separator)
        # Format pairs: (format_with_dash, format_with_slash)
        common_format_pairs = [
            # Date-only formats
            ('%Y-%m-%d', '%Y/%m/%d'),
            ('%d-%m-%Y', '%d/%m/%Y'),
            ('%m-%d-%Y', '%m/%d/%Y'),
            ('%Y-%m-%d', '%Y.%m.%d'),  # Dot separator
            ('%d-%m-%Y', '%d.%m.%Y'),
            ('%m-%d-%Y', '%m.%d.%Y'),
            # Two-digit year formats
            ('%d-%m-%y', '%d/%m/%y'),
            ('%m-%d-%y', '%m/%d/%y'),
            ('%y-%m-%d', '%y/%m/%d'),
            ('%d-%m-%y', '%d.%m.%y'),
            ('%m-%d-%y', '%m.%d.%y'),
            # Date with time (hours:minutes:seconds)
            ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S'),
            ('%d-%m-%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S'),
            ('%m-%d-%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S'),
            ('%Y-%m-%d %H:%M:%S', '%Y.%m.%d %H:%M:%S'),
            ('%d-%m-%Y %H:%M:%S', '%d.%m.%Y %H:%M:%S'),
            # Date with time (hours:minutes only)
            ('%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M'),
            ('%d-%m-%Y %H:%M', '%d/%m/%Y %H:%M'),
            ('%m-%d-%Y %H:%M', '%m/%d/%Y %H:%M'),
            # ISO 8601 formats
            ('%Y-%m-%dT%H:%M:%S', '%Y/%m/%dT%H:%M:%S'),
            ('%Y-%m-%dT%H:%M:%S.%f', '%Y/%m/%dT%H:%M:%S.%f'),  # With microseconds
            ('%Y-%m-%dT%H:%M:%SZ', '%Y/%m/%dT%H:%M:%SZ'),  # UTC timezone
            ('%Y-%m-%dT%H:%M:%S%z', '%Y/%m/%dT%H:%M:%S%z'),  # With timezone offset
            ('%Y-%m-%dT%H:%M:%S.%f%z', '%Y/%m/%dT%H:%M:%S.%f%z'),  # Microseconds + timezone
            # Compact formats (no separators)
            ('%Y%m%d', '%Y%m%d'),  # Same for both (no separator to normalize)
            ('%d%m%Y', '%d%m%Y'),
            ('%m%d%Y', '%m%d%Y'),
            ('%Y%m%d %H%M%S', '%Y%m%d %H%M%S'),
            # Text-based month formats (no separator normalization needed)
            ('%d %B %Y', '%d %B %Y'),  # "01 January 2021"
            ('%B %d, %Y', '%B %d, %Y'),  # "January 01, 2021"
            ('%d %b %Y', '%d %b %Y'),  # "01 Jan 2021"
            ('%b %d, %Y', '%b %d, %Y'),  # "Jan 01, 2021"
            ('%d-%B-%Y', '%d-%B-%Y'),  # "01-January-2021"
            ('%d-%b-%Y', '%d-%b-%Y'),  # "01-Jan-2021"
            ('%B %d %Y', '%B %d %Y'),  # "January 01 2021"
            ('%b %d %Y', '%b %d %Y'),  # "Jan 01 2021"
            # Text formats with time
            ('%d %B %Y %H:%M:%S', '%d %B %Y %H:%M:%S'),
            ('%B %d, %Y %H:%M:%S', '%B %d, %Y %H:%M:%S'),
            ('%d %b %Y %H:%M:%S', '%d %b %Y %H:%M:%S'),
        ]
        
        detected_format = None
        for fmt_dash, fmt_slash in common_format_pairs:
            try:
                # Test with normalized values from entire column
                success_count = 0
                for val in normalized_all:
                    try:
                        pd.to_datetime(val, format=fmt_dash)
                        success_count += 1
                    except:
                        break
                
                if success_count >= len(normalized_all):
                    # Format matched! Now return the format that matches original data separator
                    # Check for dot separator first (most specific)
                    if has_dot and not has_slash and not has_dash:
                        # Original data uses '.', return format with '.' (convert dash format to dot)
                        detected_format = fmt_dash.replace('-', '.')
                    elif has_slash and not has_dash and not has_dot:
                        # Original data uses '/', return format with '/'
                        detected_format = fmt_slash
                    elif has_dash and not has_slash and not has_dot:
                        # Original data uses '-', return format with '-'
                        detected_format = fmt_dash
                    else:
                        # Mixed or unclear: prefer the format that matches the first value
                        first_val = str(all_values[0])
                        if '.' in first_val and '.' not in first_val.split()[0] if ' ' in first_val else True:
                            # Check if dot is a date separator (not decimal or end of sentence)
                            detected_format = fmt_dash.replace('-', '.')
                        elif '/' in first_val:
                            detected_format = fmt_slash
                        else:
                            detected_format = fmt_dash
                    break
            except:
                continue
        
        return {
            "detected_format": detected_format,
            "can_detect": detected_format is not None,
            "sample_values": sample_values
        }
        
    except Exception as e:
        logger.error(f"Error detecting datetime format: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/apply-data-transformations")
async def apply_data_transformations(request: Request):
    """
    Apply column drops, renames, dtype changes and missing value strategies to a file.
    Expects JSON body with:
    - file_path: str
    - columns_to_drop: list[str] (column names to remove) - OPTIONAL
    - column_renames: dict[str, str] (old_name -> new_name) - OPTIONAL
    - dtype_changes: dict[str, str | dict] (column_name -> new_dtype or {dtype: str, format: str})
    - missing_value_strategies: dict[str, dict] (column_name -> {strategy: str, value?: str})
    """
    try:
        body = await request.json()
        logger.info("=" * 80)
        logger.info("apply_data_transformations endpoint called")
        logger.info(f"Full request body: {body}")
        logger.info("=" * 80)
        
        file_path = body.get("file_path")
        columns_to_drop = body.get("columns_to_drop", [])
        column_renames = body.get("column_renames", {})
        dtype_changes = body.get("dtype_changes", {})
        missing_value_strategies = body.get("missing_value_strategies", {})
        
        logger.info(f"Extracted file_path: {file_path}")
        logger.info(f"Extracted columns_to_drop: {columns_to_drop}")
        logger.info(f"Extracted column_renames: {column_renames}")
        logger.info(f"Extracted dtype_changes: {dtype_changes}")
        logger.info(f"Extracted missing_value_strategies: {missing_value_strategies}")
        
        # Log what transformations are being applied
        if len(columns_to_drop) > 0:
            logger.info(f"✅ COLUMN DROPS: {len(columns_to_drop)} columns to drop")
        if len(column_renames) > 0:
            logger.info(f"✅ COLUMN RENAMES: {len(column_renames)} columns to rename")
        if len(dtype_changes) == 0 and len(missing_value_strategies) > 0:
            logger.warning("⚠️  MISSING VALUE ONLY REQUEST - No dtype changes found!")
        elif len(dtype_changes) > 0 and len(missing_value_strategies) == 0:
            logger.warning("⚠️  DTYPE ONLY REQUEST - No missing value strategies found!")
        elif len(dtype_changes) > 0 and len(missing_value_strategies) > 0:
            logger.info("✅ COMBINED REQUEST - Both dtype and missing value changes found!")
        
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path is required")
        
        # Read file from MinIO
        data = read_minio_object(file_path)
        filename = Path(file_path).name
        
        # Parse based on file type using RobustFileReader
        if filename.lower().endswith(".csv") or filename.lower().endswith((".xls", ".xlsx")):
            # Use RobustFileReader which handles column preservation automatically
            df_result, _ = RobustFileReader.read_file_to_pandas(
                content=data,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            # Handle both single DataFrame and dict (multiple sheets)
            if isinstance(df_result, dict):
                df = list(df_result.values())[0]  # Use first sheet
            else:
                df = df_result
        elif filename.lower().endswith(".arrow"):
            df_pl = pl.read_ipc(io.BytesIO(data))
            df = df_pl.to_pandas()
        else:
            raise HTTPException(status_code=400, detail="Only CSV, XLSX and Arrow files supported")
        
        # 1. Drop columns first (before renames)
        if columns_to_drop:
            logger.info(f"Dropping columns: {columns_to_drop}")
            cols_to_drop_existing = [col for col in columns_to_drop if col in df.columns]
            if cols_to_drop_existing:
                df = df.drop(columns=cols_to_drop_existing)
                logger.info(f"✅ Dropped columns: {cols_to_drop_existing}")
            else:
                logger.info("No columns to drop found in dataframe")
        
        # 2. Apply column renames (after drops)
        if column_renames:
            logger.info(f"Applying column renames: {column_renames}")
            # Filter out renames where old_name == new_name
            valid_renames = {old: new for old, new in column_renames.items() if old != new and old in df.columns}
            if valid_renames:
                df = df.rename(columns=valid_renames)
                logger.info(f"✅ Renamed columns: {valid_renames}")
            else:
                logger.info("No valid column renames to apply")
        
        # Apply missing value strategies
        for col_name, strategy_config in missing_value_strategies.items():
            if col_name not in df.columns:
                continue
                
            strategy = strategy_config.get("strategy", "none")
            
            if strategy == "none":
                continue
            elif strategy == "drop":
                df = df.dropna(subset=[col_name])
            elif strategy == "mean":
                if pd.api.types.is_numeric_dtype(df[col_name]):
                    df[col_name].fillna(df[col_name].mean(), inplace=True)
            elif strategy == "median":
                if pd.api.types.is_numeric_dtype(df[col_name]):
                    df[col_name].fillna(df[col_name].median(), inplace=True)
            elif strategy == "mode":
                mode_val = df[col_name].mode()
                if len(mode_val) > 0:
                    df[col_name].fillna(mode_val[0], inplace=True)
            elif strategy == "zero":
                df[col_name].fillna(0, inplace=True)
            elif strategy == "empty":
                df[col_name].fillna("", inplace=True)
            elif strategy == "custom":
                custom_value = strategy_config.get("value", "")
                logger.info(f"Applying custom value '{custom_value}' to column '{col_name}' (dtype: {df[col_name].dtype})")
                
                # Convert custom value to match column dtype
                if pd.api.types.is_numeric_dtype(df[col_name]):
                    try:
                        # Try to convert to numeric
                        numeric_value = pd.to_numeric(custom_value, errors='coerce')
                        if pd.notna(numeric_value):
                            df[col_name].fillna(numeric_value, inplace=True)
                            logger.info(f"Converted custom value '{custom_value}' to numeric: {numeric_value}")
                        else:
                            logger.warning(f"Could not convert custom value '{custom_value}' to numeric for column '{col_name}'")
                    except Exception as e:
                        logger.warning(f"Error converting custom value '{custom_value}' to numeric: {str(e)}")
                else:
                    # For non-numeric columns, use as string
                    df[col_name].fillna(str(custom_value), inplace=True)
                    logger.info(f"Applied custom string value '{custom_value}' to column '{col_name}'")
            elif strategy == "ffill":
                df[col_name] = df[col_name].ffill()
            elif strategy == "bfill":
                df[col_name] = df[col_name].bfill()
        
        # Apply dtype changes
        logger.info(f"Starting dtype changes. Total dtype_changes to apply: {len(dtype_changes)}")
        logger.info(f"dtype_changes received: {dtype_changes}")
        
        for col_name, dtype_config in dtype_changes.items():
            logger.info(f"Processing dtype change for column: {col_name}, config: {dtype_config}")
            
            if col_name not in df.columns:
                logger.warning(f"Column '{col_name}' not found in dataframe. Skipping.")
                continue
            
            # Handle both string dtype and dict with {dtype, format}
            if isinstance(dtype_config, dict):
                new_dtype = dtype_config.get('dtype')
                datetime_format = dtype_config.get('format')
                logger.info(f"Dict config detected - dtype: {new_dtype}, format: {datetime_format}")
            else:
                new_dtype = dtype_config
                datetime_format = None
                logger.info(f"String config detected - dtype: {new_dtype}")
                
            try:
                if new_dtype == "int64":
                    # logger.info(f"Converting column '{col_name}' to int64")
                    # logger.info(f"Sample values before conversion: {df[col_name].head(5).tolist()}")
                    # logger.info(f"Column dtype before conversion: {df[col_name].dtype}")
                    
                    # Convert to numeric first, then round to remove decimals, then to Int64
                    numeric_col = pd.to_numeric(df[col_name], errors='coerce')
                    df[col_name] = numeric_col.round().astype('Int64')
                    
                    # logger.info(f"Sample values after conversion: {df[col_name].head(5).tolist()}")
                    # logger.info(f"Column dtype after conversion: {df[col_name].dtype}")
                    # logger.info(f"Non-null count: {df[col_name].notna().sum()} out of {len(df[col_name])}")
                elif new_dtype == "float64":
                    df[col_name] = pd.to_numeric(df[col_name], errors='coerce')
                elif new_dtype == "object":
                    df[col_name] = df[col_name].astype(str)
                elif new_dtype == "datetime64":
                    # Mirror process_saved_dataframe behavior: single-step conversion with optional format
                    if datetime_format:
                        df[col_name] = pd.to_datetime(df[col_name], format=datetime_format, errors='coerce')
                    else:
                        logger.info(f"Converting column '{col_name}' to datetime64 without specific format")
                        df[col_name] = pd.to_datetime(df[col_name], errors='coerce')
                elif new_dtype == "bool":
                    df[col_name] = df[col_name].astype(bool)
            except Exception as e:
                logger.warning(f"Could not convert {col_name} to {new_dtype}: {str(e)}")
        
        # Save back to MinIO (overwrite the temp file)
        buffer = io.BytesIO()
        if filename.lower().endswith(".csv"):
            df.to_csv(buffer, index=False)
        elif filename.lower().endswith((".xls", ".xlsx")):
            df.to_excel(buffer, index=False)
        elif filename.lower().endswith(".arrow"):
            df_pl_updated = pl.from_pandas(df)
            df_pl_updated.write_ipc(buffer)
        
        buffer.seek(0)
        
        # Upload back to MinIO
        minio_client.put_object(
            MINIO_BUCKET,
            file_path,
            buffer,
            length=buffer.getbuffer().nbytes,
            content_type="application/octet-stream",
        )
        
        return {
            "status": "success",
            "message": "Transformations applied successfully",
            "rows_affected": len(df),
        }
        
    except Exception as e:
        logger.error(f"Error applying transformations: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/get-guided-flow-state")
async def get_guided_flow_state(
    client_name: str = Query(""),
    app_name: str = Query(""),
    project_name: str = Query(""),
):
    """Retrieve persisted guided flow state from Redis."""
    try:
        key_parts = ("guided_flow_state", client_name, app_name, project_name)
        cached_state_str = redis_client.get(key_parts)
        
        if cached_state_str:
            try:
                cached_state = json.loads(cached_state_str)
                return {"state": cached_state}
            except json.JSONDecodeError:
                logger.warning(f"Invalid JSON in cached state for {key_parts}")
                return {"state": None}
        
        return {"state": None}
    except Exception as e:
        logger.error(f"Error retrieving guided flow state: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/save-guided-flow-state")
async def save_guided_flow_state(request: Request):
    """Save guided flow state to Redis."""
    try:
        body = await request.json()
        client_name = body.get("client_name", "")
        app_name = body.get("app_name", "")
        project_name = body.get("project_name", "")
        state = body.get("state", {})
        
        key_parts = ("guided_flow_state", client_name, app_name, project_name)
        # Store JSON as string with 24 hours TTL
        state_json = json.dumps(state)
        redis_client.set(key_parts, state_json, ex=86400)  # 24 hours TTL
        
        return {"status": "success", "message": "Flow state saved"}
    except Exception as e:
        logger.error(f"Error saving guided flow state: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/check-priming-status")
async def check_priming_status(
    client_name: str = Query(""),
    app_name: str = Query(""),
    project_name: str = Query(""),
    file_name: str = Query(""),
):
    """Check if a file has completed priming (all 7 steps) and return step-by-step status."""
    try:
        # Check Redis for completion flag
        primed_key_parts = ("primed_files", client_name, app_name, project_name, file_name)
        is_primed = redis_client.get(primed_key_parts)
        
        # Check for flow state to get current stage and step completion
        flow_key_parts = ("guided_flow_state", client_name, app_name, project_name)
        flow_state_str = redis_client.get(flow_key_parts)
        flow_state = None
        if flow_state_str:
            try:
                flow_state = json.loads(flow_state_str)
            except json.JSONDecodeError:
                logger.warning(f"Invalid JSON in flow state for {flow_key_parts}")
        
        current_stage = None
        completed_steps: list[str] = []
        all_steps = ["U0", "U1", "U2", "U3", "U4", "U5", "U6", "U7"]
        
        if flow_state and isinstance(flow_state, dict):
            current_stage = flow_state.get("currentStage")
            # Check if this file is in the uploaded files
            uploaded_files = flow_state.get("uploadedFiles", [])
            file_in_flow = any(
                f.get("path") == file_name or f.get("name") == file_name
                for f in uploaded_files
            )
            
            if file_in_flow:
                current_stage = flow_state.get("currentStage")
                # Determine completed steps based on current stage
                if current_stage:
                    stage_index = all_steps.index(current_stage) if current_stage in all_steps else -1
                    if stage_index >= 0:
                        # All steps up to and including current stage are completed
                        completed_steps = all_steps[:stage_index + 1]
                    elif current_stage == "U7":
                        # All steps completed
                        completed_steps = all_steps
                else:
                    # If no current stage but file is in flow, assume U0 completed
                    completed_steps = ["U0"]
        
        # If primed flag exists, all steps are completed
        if is_primed:
            completed_steps = all_steps
            current_stage = "U7"
        
        # Determine completion status
        completed = bool(is_primed) or (current_stage == "U7")
        missing_steps = [s for s in all_steps if s not in completed_steps]
        # Status colors:
        # - Red: U0 or U1 (step 1 or earlier) - not in progress
        # - Yellow: U2-U6 (step 2 to step 7) - in progress
        # - Green: U7 (step 8 - completed) - primed
        is_in_progress = bool(current_stage) and current_stage not in ["U0", "U1", "U7"]
        
        return {
            "completed": completed,
            "current_stage": current_stage,
            "is_primed": bool(is_primed),
            "completed_steps": completed_steps,
            "missing_steps": missing_steps,
            "is_in_progress": is_in_progress,
        }
    except Exception as e:
        logger.error(f"Error checking priming status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/mark-file-primed")
async def mark_file_primed(request: Request):
    """Mark a file as primed (completed all 7 steps)."""
    try:
        body = await request.json()
        client_name = body.get("client_name", "")
        app_name = body.get("app_name", "")
        project_name = body.get("project_name", "")
        file_name = body.get("file_name", "")
        
        # Set primed flag in Redis
        primed_key_parts = ("primed_files", client_name, app_name, project_name, file_name)
        redis_client.set(primed_key_parts, "true", ttl=86400 * 30)  # 30 days TTL
        
        return {"status": "success", "message": "File marked as primed"}
    except Exception as e:
        logger.error(f"Error marking file as primed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/finalize-primed-file")
async def finalize_primed_file(request: Request):
    """
    Finalize a primed file by moving it from tmp location to saved dataframes location.
    This should be called after U7 completion to persist the transformed data.
    
    Expects JSON body with:
    - file_path: str (current path in tmp/)
    - file_name: str (desired file name for saved dataframe)
    - client_name, app_name, project_name: str (for prefix)
    - validator_atom_id: str (optional, for flight path)
    - column_classifications: list[dict] (optional, from guided flow U4 stage)
      Each dict has: columnName, columnRole ('identifier' | 'measure')
    """
    try:
        body = await request.json()
        logger.info("=" * 80)
        logger.info("finalize_primed_file endpoint called")
        logger.info(f"Request body: {body}")
        logger.info("=" * 80)
        
        file_path = body.get("file_path", "")
        file_name = body.get("file_name", "")
        client_name = body.get("client_name", "")
        app_name = body.get("app_name", "")
        project_name = body.get("project_name", "")
        validator_atom_id = body.get("validator_atom_id", "guided-upload")
        # Column classifications from guided flow (U4 stage)
        column_classifications = body.get("column_classifications", [])
        
        if not file_path:
            raise HTTPException(status_code=400, detail="file_path is required")
        if not file_name:
            # Use the original file name from path
            file_name = Path(file_path).stem
        
        # Read the transformed file from MinIO
        try:
            data = read_minio_object(file_path)
        except Exception as e:
            logger.error(f"Failed to read file from MinIO: {file_path}, error: {str(e)}")
            raise HTTPException(status_code=404, detail=f"File not found: {file_path}")
        
        filename = Path(file_path).name
        
        # Parse the file based on type
        if filename.lower().endswith(".csv"):
            df_result, _ = RobustFileReader.read_file_to_polars(
                content=data,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            if isinstance(df_result, dict):
                df_pl = list(df_result.values())[0]
            else:
                df_pl = df_result
        elif filename.lower().endswith((".xls", ".xlsx")):
            df_result, _ = RobustFileReader.read_file_to_polars(
                content=data,
                filename=filename,
                auto_detect_header=True,
                return_raw=False,
            )
            if isinstance(df_result, dict):
                df_pl = list(df_result.values())[0]
            else:
                df_pl = df_result
        elif filename.lower().endswith(".arrow"):
            df_pl = pl.read_ipc(io.BytesIO(data))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type")
        
        # Convert to Arrow format
        arrow_buf = io.BytesIO()
        arrow_table = df_pl.to_arrow(use_pyarrow=True)
        with pa.ipc.new_file(arrow_buf, arrow_table.schema) as writer:
            writer.write(arrow_table)
        arrow_bytes = arrow_buf.getvalue()
        
        # Get the proper prefix for saved dataframes
        os.environ["CLIENT_NAME"] = client_name
        os.environ["APP_NAME"] = app_name
        os.environ["PROJECT_NAME"] = project_name
        prefix = await get_object_prefix()
        
        # Create the arrow file name
        arrow_name = Path(file_name).stem + ".arrow"
        
        # Upload to MinIO with proper prefix (not tmp/)
        result = upload_to_minio(arrow_bytes, arrow_name, prefix)
        saved_object_name = result.get("object_name", "")
        
        logger.info(f"✅ Saved primed file to: {saved_object_name}")
        
        # Upload to Flight for fast access
        flight_path = f"{validator_atom_id}/{arrow_name}"
        try:
            upload_dataframe(df_pl.to_pandas(), flight_path)
            logger.info(f"✅ Uploaded to Flight: {flight_path}")
        except Exception as e:
            logger.warning(f"Failed to upload to Flight: {str(e)}")
        
        # Set ticket for Flight access
        set_ticket(
            file_name,
            saved_object_name,
            flight_path,
            filename,
        )
        redis_client.set(f"flight:{flight_path}", saved_object_name)
        
        # Mark as primed in Redis
        primed_key_parts = ("primed_files", client_name, app_name, project_name, file_name)
        redis_client.set(primed_key_parts, "true", ttl=86400 * 30)
        
        # Save column classifications to MongoDB if provided from guided flow
        if column_classifications:
            try:
                # Extract identifiers and measures from guided flow classifications
                identifiers = []
                measures = []
                unclassified = []
                
                for col_class in column_classifications:
                    col_name = str(col_class.get("columnName", "")).strip().lower()
                    col_role = col_class.get("columnRole", "")
                    
                    if col_role == "identifier":
                        identifiers.append(col_name)
                    elif col_role == "measure":
                        measures.append(col_name)
                    else:
                        unclassified.append(col_name)
                
                # Get project_id from env
                project_id = None
                env = None
                try:
                    env = await get_env_vars(
                        client_name=client_name,
                        app_name=app_name,
                        project_name=project_name,
                    )
                    if env:
                        project_id_str = env.get("PROJECT_ID")
                        if project_id_str:
                            project_id = int(project_id_str)
                except Exception as e:
                    logger.warning(f"Failed to get project_id: {e}")
                
                # Save to MongoDB using same pattern as column classifier atom
                config_data = {
                    "project_id": project_id,
                    "client_name": client_name,
                    "app_name": app_name,
                    "project_name": project_name,
                    "identifiers": identifiers,
                    "measures": measures,
                    "unclassified": unclassified,
                    "dimensions": {},  # Empty dimensions object (same as column classifier)
                    "file_name": saved_object_name,
                }
                
                # Add env if available
                if env:
                    config_data["env"] = env
                
                mongo_result = save_classifier_config_to_mongo(config_data)
                logger.info(f"✅ Saved guided flow classification: {saved_object_name} | {len(identifiers)} identifiers, {len(measures)} measures, {len(unclassified)} unclassified")
            except Exception as e:
                logger.warning(f"Failed to save column classifications: {e}")
                # Don't fail the whole operation if classification save fails
        else:
            # No classifications provided - run auto-classification
            try:
                project_id = None
                try:
                    env = await get_env_vars(
                        client_name=client_name,
                        app_name=app_name,
                        project_name=project_name,
                    )
                    if env:
                        project_id_str = env.get("PROJECT_ID")
                        if project_id_str:
                            project_id = int(project_id_str)
                except Exception:
                    pass
                
                await _auto_classify_and_save_file(
                    object_name=saved_object_name,
                    client_name=client_name,
                    app_name=app_name,
                    project_name=project_name,
                    project_id=project_id,
                )
            except Exception as e:
                logger.warning(f"Failed to auto-classify file: {e}")
        
        # Remove the temp file
        if file_path.startswith(prefix + "tmp/") or "/tmp/" in file_path:
            try:
                minio_client.remove_object(MINIO_BUCKET, file_path)
                logger.info(f"🗑️ Removed temp file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to remove temp file: {str(e)}")
        
        return {
            "status": "success",
            "message": "File finalized and saved",
            "saved_path": saved_object_name,
            "flight_path": flight_path,
            "rows": len(df_pl),
            "columns": len(df_pl.columns),
            "classification_saved": bool(column_classifications),
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error finalizing primed file: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))
