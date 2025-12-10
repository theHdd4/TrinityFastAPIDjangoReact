"""Robust Excel reader with multiple sheet support and header detection."""

import logging
import io
from typing import Dict, Tuple, Optional, List
import pandas as pd

from app.features.data_upload_validate.file_ingestion.processors.header_detector import HeaderDetector
from app.features.data_upload_validate.file_ingestion.processors.cleaning import DataCleaner

logger = logging.getLogger(__name__)


class ExcelReader:
    """Robust Excel reader that handles various edge cases."""

    @staticmethod
    def read(
        content: bytes,
        sheet_name: Optional[str] = None,
        auto_detect_header: bool = True,
        skip_empty_rows: bool = True,
        return_raw: bool = False,
    ) -> Tuple[Dict[str, pd.DataFrame], dict]:
        """
        Read Excel file with robust handling.
        
        Args:
            content: File content bytes
            sheet_name: Optional specific sheet name (if None, reads all sheets)
            auto_detect_header: Whether to auto-detect header row
            skip_empty_rows: Whether to skip mostly empty rows
            return_raw: If True, return raw DataFrame without header detection (all rows as data)
        
        Returns:
            Tuple of (dict of {sheet_name: DataFrame}, metadata dict)
        """
        metadata = {
            "sheet_names": [],
            "selected_sheet": None,
            "has_multiple_sheets": False,
            "header_rows": {},
            "data_start_rows": {},
        }
        
        try:
            excel_file = pd.ExcelFile(io.BytesIO(content), engine='openpyxl')
            sheet_names = excel_file.sheet_names or ["Sheet1"]
            metadata["sheet_names"] = sheet_names
            metadata["has_multiple_sheets"] = len(sheet_names) > 1
            
            # Determine which sheets to read
            sheets_to_read = [sheet_name] if sheet_name and sheet_name in sheet_names else sheet_names
            
            if sheet_name and sheet_name not in sheet_names:
                logger.warning(f"Sheet '{sheet_name}' not found, using '{sheet_names[0]}'")
                sheets_to_read = [sheet_names[0]]
            
            if sheet_name:
                metadata["selected_sheet"] = sheet_name if sheet_name in sheet_names else sheet_names[0]
            else:
                metadata["selected_sheet"] = sheets_to_read[0] if sheets_to_read else None
            
            dfs = {}
            
            for sheet in sheets_to_read:
                try:
                    # CRITICAL: Find max columns FIRST by scanning all rows
                    # This prevents truncation when description rows have fewer columns than data rows
                    max_cols = ExcelReader._find_max_columns(excel_file, sheet, sample_rows=0)  # Scan ALL rows
                    logger.debug(f"Sheet '{sheet}' max columns detected: {max_cols}")
                    
                    # CRITICAL FIX: Use usecols parameter to force pandas to read all columns
                    # Convert column index to Excel column letters (A, B, C, ..., Z, AA, AB, ...)
                    if max_cols > 0:
                        # Generate column letters up to max_cols
                        def num_to_col_letter(n):
                            """Convert 0-based column index to Excel column letter (A=0, B=1, ..., Z=25, AA=26, ...)"""
                            result = ""
                            n += 1  # Convert to 1-based
                            while n > 0:
                                n -= 1
                                result = chr(65 + (n % 26)) + result
                                n //= 26
                            return result
                        
                        # Create usecols range: "A:XXX" where XXX is the last column
                        last_col_letter = num_to_col_letter(max_cols - 1)
                        usecols_range = f"A:{last_col_letter}"
                        logger.debug(f"Using usecols range: {usecols_range} to preserve all {max_cols} columns")
                    else:
                        usecols_range = None
                    
                    df_raw = pd.read_excel(
                        excel_file,
                        sheet_name=sheet,
                        header=None,
                        engine='openpyxl',
                        keep_default_na=False,  # Don't convert empty strings to NaN
                        na_values=[],  # Don't treat any values as NaN - preserve everything
                        usecols=usecols_range,  # Force reading all columns up to max_cols
                    )
                    
                    # CRITICAL FIX: If detected max columns is greater than what pandas read,
                    # we need to expand the DataFrame to include all columns
                    if max_cols > len(df_raw.columns):
                        logger.warning(f"Pandas read {len(df_raw.columns)} columns but detected {max_cols} columns. Expanding DataFrame.")
                        # Add missing columns filled with NaN
                        for i in range(len(df_raw.columns), max_cols):
                            df_raw[f"col_{i}"] = None
                        # Reorder columns to ensure correct order
                        df_raw = df_raw.reindex(columns=[f"col_{i}" for i in range(max_cols)], fill_value=None)
                    
                    logger.info(f"Read sheet '{sheet}': {len(df_raw)} rows, {len(df_raw.columns)} columns")
                    
                    if df_raw.empty:
                        logger.warning(f"Sheet '{sheet}' is empty (no rows)")
                        dfs[sheet] = pd.DataFrame()
                        continue
                    
                    # If return_raw is True, return raw DataFrame without any header processing
                    # IMPORTANT: Return ALL rows, even if they appear empty - user needs to see everything
                    if return_raw:
                        metadata["header_rows"][sheet] = None
                        metadata["data_start_rows"][sheet] = 0
                        # Don't filter rows - return everything as-is, including empty rows
                        dfs[sheet] = df_raw.reset_index(drop=True)
                        logger.info(f"Returning raw sheet '{sheet}': {len(dfs[sheet])} rows (including empty rows)")
                        continue
                    
                    # Detect header row if requested
                    if auto_detect_header:
                        header_row = HeaderDetector.find_header_row(df_raw)
                        data_start = HeaderDetector.detect_table_start(df_raw, header_row)
                        metadata["header_rows"][sheet] = header_row
                        metadata["data_start_rows"][sheet] = data_start
                        
                        # Extract headers and data
                        headers = df_raw.iloc[header_row].fillna("").astype(str).tolist()
                        df = df_raw.iloc[data_start:].copy()
                        df.columns = headers
                    else:
                        # Use first row as header
                        headers = df_raw.iloc[0].fillna("").astype(str).tolist()
                        df = df_raw.iloc[1:].copy()
                        df.columns = headers
                    
                    # Clean headers
                    df = DataCleaner.normalize_column_names(df)
                    df = DataCleaner.standardize_headers(df)
                    
                    # Remove empty rows/columns if requested
                    if skip_empty_rows:
                        df = DataCleaner.remove_empty_rows(df)
                    
                    df = DataCleaner.remove_empty_columns(df)
                    
                    # Reset index
                    df = df.reset_index(drop=True)
                    
                    dfs[sheet] = df
                    
                except Exception as e:
                    logger.error(f"Error reading sheet '{sheet}': {e}")
                    dfs[sheet] = pd.DataFrame()
            
            return dfs, metadata
            
        except Exception as e:
            logger.exception(f"Error reading Excel file: {e}")
            raise ValueError(f"Error parsing Excel file: {e}") from e
    
    @staticmethod
    def _find_max_columns(excel_file: pd.ExcelFile, sheet_name: str, sample_rows: int = 0) -> int:
        """
        Find the maximum number of columns across all rows in an Excel sheet.
        This prevents truncation when description rows have fewer columns than data rows.
        
        Uses openpyxl's max_column property as the primary method (most reliable),
        with row scanning as verification/fallback.
        
        Args:
            excel_file: pd.ExcelFile object
            sheet_name: Name of the sheet to scan
            sample_rows: Number of rows to sample (0 = all rows, default: scan all)
        
        Returns:
            Maximum column count found
        """
        try:
            import openpyxl
            
            # Get the workbook from the ExcelFile
            workbook = excel_file.book
            
            # Get the worksheet
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return 0
            
            worksheet = workbook[sheet_name]
            
            # PRIMARY METHOD: Use openpyxl's max_column property
            # This tracks the maximum column that has ever been used in the sheet,
            # which is more reliable than scanning rows, especially for sparse data
            max_cols_from_property = worksheet.max_column
            logger.debug(f"Sheet '{sheet_name}' max_column property: {max_cols_from_property}")
            
            # VERIFICATION METHOD: Scan rows to find rightmost non-empty cell per row
            # This helps catch cases where max_column might be inaccurate
            max_cols_from_scan = 0
            row_count = 0
            
            # CRITICAL: Scan ALL rows (or sample_rows if specified) to find true maximum
            # This is essential because description rows at the start may have fewer columns
            # than data rows that come later
            for row in worksheet.iter_rows(values_only=True):
                if sample_rows > 0 and row_count >= sample_rows:
                    break
                
                # Count columns in row - find the rightmost non-empty cell
                # This gives us the actual column count for this row
                last_col_idx = -1
                for idx, val in enumerate(row):
                    if val is not None and str(val).strip():
                        last_col_idx = idx
                
                # Column count is last_col_idx + 1 (0-based to 1-based)
                # If no non-empty cells, count all non-None cells
                if last_col_idx >= 0:
                    col_count = last_col_idx + 1
                else:
                    col_count = len([v for v in row if v is not None])
                
                if col_count > max_cols_from_scan:
                    max_cols_from_scan = col_count
                
                row_count += 1
            
            # Take the maximum of both methods to ensure we don't miss any columns
            max_cols = max(max_cols_from_property, max_cols_from_scan)
            
            logger.info(
                f"Sheet '{sheet_name}' column detection: "
                f"max_column property={max_cols_from_property}, "
                f"scan result={max_cols_from_scan}, "
                f"final max_cols={max_cols} "
                f"(scanned {row_count} rows)"
            )
            
            # Ensure we return at least 1 column
            return max_cols if max_cols > 0 else 1
            
        except Exception as e:
            logger.warning(f"Error finding max columns for sheet '{sheet_name}': {e}")
            # Fallback: return 0 to let pandas infer naturally
            return 0

